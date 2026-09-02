"""Vacancy Request bulk upload -- validate / preview / commit / undo
(2026-08-30).

This joins the existing six-entity BulkUpload framework rather than adding a
third one-shot importer, but it is the odd one out in that framework and these
tests exist mostly to pin the two ways it differs:

1. **Create-only.** Every other entity there is master data, where a
   re-uploaded row should upsert. A vacancy request is an EVENT -- two
   identical rows are two genuine requests, not a duplicate -- so re-uploading
   the same file creates more requests rather than matching existing ones, and
   `updated_count`/`unchanged_count` are always 0.

2. **Undo cancels DRAFTS, it does not soft-delete.** VacancyRequest has no
   `is_active`; it has a status lifecycle. And unlike a Location, a row can
   move on by itself inside the 24h window, so anything past DRAFT is left
   alone rather than being yanked out of an approval chain.
"""

import io
import uuid

import pytest
from openpyxl import Workbook

from app.models.enums import (
    BulkUploadEntityTypeEnum,
    StaffRoleCategoryEnum,
    UserRoleEnum,
    VacancyRequestSourceEnum,
    VacancyRequestStatusEnum,
)
from app.models.vacancy_request import VacancyRequest

from tests.conftest import auth_headers

TEMPLATE = "/api/v1/vacancy-requests/bulk-upload/template"
VALIDATE = "/api/v1/vacancy-requests/bulk-upload/validate"
COMMIT = "/api/v1/vacancy-requests/bulk-upload/commit"

HEADERS = (
    "Campus Code",
    "Department Name",
    "Designation",
    "Number of Positions",
    "Priority",
    "Required By (DD-MM-YYYY)",
    "Justification",
)

# The current template. HEADERS above is deliberately left as the ORIGINAL
# seven columns and every pre-existing test still uses it -- that is the
# backward-compatibility check: a workbook downloaded before the referrer
# columns existed must still import.
HEADERS_WITH_REQUESTER = (
    *HEADERS,
    "Requester Name",
    "Requester Email",
    "Requester Mobile",
)

# The full current template. Kept separate from HEADERS_WITH_REQUESTER above
# so the tests that upload the OLDER header sets keep proving that a workbook
# downloaded before a column was appended still parses.
HEADERS_WITH_LOCATION = (*HEADERS_WITH_REQUESTER, "Location")


def _workbook(rows: list[tuple], headers: tuple = HEADERS) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(client, actor, url, rows, headers: tuple = HEADERS):
    return client.post(
        url,
        headers=auth_headers(client, actor),
        files={
            "file": (
                "vacancy_requests.xlsx",
                _workbook(rows, headers),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


@pytest.fixture()
def bulk_setup(campus_factory, department_factory, designation_factory, user_factory, db_session):
    campus = campus_factory("SSE")
    department = department_factory("SSE", name=f"Bulk Dept {uuid.uuid4().hex[:6]}")
    department.supported_categories = [StaffRoleCategoryEnum.TEACHING]
    designation = designation_factory(
        StaffRoleCategoryEnum.TEACHING, name=f"Bulk Desig {uuid.uuid4().hex[:6]}", department=department
    )
    actor = user_factory(UserRoleEnum.SUPER_ADMIN)
    db_session.commit()
    return campus, department, designation, actor


# --- template -----------------------------------------------------------------


def test_template_downloads_as_a_workbook(client, bulk_setup):
    _c, _d, _des, actor = bulk_setup
    response = client.get(TEMPLATE, headers=auth_headers(client, actor))

    assert response.status_code == 200
    assert response.content[:2] == b"PK"  # xlsx is a zip
    assert "vacancy_request_bulk_upload_template.xlsx" in response.headers["content-disposition"]


# --- validate is a pure preview ----------------------------------------------


def test_validate_previews_without_writing_anything(client, bulk_setup, db_session):
    campus, department, designation, actor = bulk_setup
    before = db_session.query(VacancyRequest).count()

    response = _upload(
        client, actor, VALIDATE, [("SSE", department.name, designation.name, 2, "NORMAL", "01-04-2026", "Growth")]
    )

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["total"] == 1
    assert body["created_count"] == 1
    assert body["rejected_count"] == 0
    # The whole point of a preview.
    assert db_session.query(VacancyRequest).count() == before


def test_validate_never_reports_updated_or_unchanged(client, bulk_setup):
    """Create-only: those two counts exist for shape parity with the
    master-data importers and can never be non-zero here."""
    campus, department, designation, actor = bulk_setup
    row = ("SSE", department.name, designation.name, 1, "NORMAL", "", "Need one")

    _upload(client, actor, COMMIT, [row])
    # The SAME row again -- a master-data importer would call this "unchanged".
    body = _upload(client, actor, VALIDATE, [row]).json()

    assert body["created_count"] == 1
    assert body["updated_count"] == 0
    assert body["unchanged_count"] == 0


# --- row-level validation -----------------------------------------------------


@pytest.mark.parametrize(
    "row,fragment",
    [
        (("XXX", "Any Dept", "Any Desig", 1, "NORMAL", "", "x"), "unknown campus code"),
        (("SSE", "No Such Dept", "Any Desig", 1, "NORMAL", "", "x"), "unknown department"),
        (("SSE", None, None, 1, "NORMAL", "", "x"), "are all required"),
    ],
)
def test_rejects_bad_master_data_with_a_readable_reason(client, bulk_setup, row, fragment):
    campus, department, designation, actor = bulk_setup
    # Substitute the real department/designation names where the case needs them.
    concrete = tuple(
        department.name if v == "Any Dept" else designation.name if v == "Any Desig" else v for v in row
    )

    body = _upload(client, actor, VALIDATE, [concrete]).json()

    assert body["rejected_count"] == 1
    assert fragment in body["rows"][0]["error_reason"].lower()


def test_rejects_a_designation_the_department_does_not_support(
    client, bulk_setup, designation_factory, db_session
):
    """MEMBERSHIP, not equality -- the rule CLAUDE.md calls out as the
    original bug."""
    campus, department, _designation, actor = bulk_setup
    housekeeping = designation_factory(
        StaffRoleCategoryEnum.HOUSEKEEPING, name=f"HK {uuid.uuid4().hex[:6]}", department=department
    )
    db_session.commit()

    body = _upload(
        client, actor, VALIDATE, [("SSE", department.name, housekeeping.name, 1, "NORMAL", "", "x")]
    ).json()

    assert body["rejected_count"] == 1
    assert "does not support" in body["rows"][0]["error_reason"].lower()


@pytest.mark.parametrize("count,fragment", [(0, "1 or more"), (101, "cannot exceed"), ("abc", "1 or more")])
def test_rejects_an_impossible_position_count(client, bulk_setup, count, fragment):
    campus, department, designation, actor = bulk_setup

    body = _upload(
        client, actor, VALIDATE, [("SSE", department.name, designation.name, count, "NORMAL", "", "x")]
    ).json()

    assert body["rejected_count"] == 1
    assert fragment in body["rows"][0]["error_reason"].lower()


def test_rejects_an_unreadable_date(client, bulk_setup):
    campus, department, designation, actor = bulk_setup

    body = _upload(
        client, actor, VALIDATE, [("SSE", department.name, designation.name, 1, "NORMAL", "not-a-date", "x")]
    ).json()

    assert body["rejected_count"] == 1
    assert "date" in body["rows"][0]["error_reason"].lower()


# --- commit -------------------------------------------------------------------


def test_commit_creates_drafts_attributed_to_the_uploader(client, bulk_setup, db_session):
    campus, department, designation, actor = bulk_setup

    response = _upload(
        client, actor, COMMIT, [("SSE", department.name, designation.name, 3, "HIGH", "01-04-2026", "Expansion")]
    )
    assert response.status_code == 200, response.text
    assert response.json()["created_count"] == 1

    vr = db_session.query(VacancyRequest).filter(VacancyRequest.department_id == department.id).one()
    # DRAFT, not submitted -- submit() enforces the sanction ceiling per
    # request and would leave a batch half in the approval queue.
    assert vr.status == VacancyRequestStatusEnum.DRAFT
    assert vr.source == VacancyRequestSourceEnum.BULK_UPLOAD
    assert vr.requested_by_id == actor.id
    assert vr.requested_count == 3
    assert vr.priority.value == "HIGH"
    # Taken from Designation Master, as the QR intake also does.
    assert vr.position_title == designation.name
    assert vr.qualification == designation.qualification


def test_commit_writes_nothing_for_rejected_rows(client, bulk_setup, db_session):
    campus, department, designation, actor = bulk_setup

    body = _upload(
        client,
        actor,
        COMMIT,
        [
            ("SSE", department.name, designation.name, 1, "NORMAL", "", "Good row"),
            ("XXX", department.name, designation.name, 1, "NORMAL", "", "Bad campus"),
        ],
    ).json()

    assert body["created_count"] == 1
    assert body["rejected_count"] == 1
    assert db_session.query(VacancyRequest).filter(VacancyRequest.department_id == department.id).count() == 1


def test_commit_records_the_batch_against_the_right_entity_type(client, bulk_setup, db_session):
    from app.models.bulk_upload_log import BulkUploadLog

    campus, department, designation, actor = bulk_setup

    log_id = _upload(
        client, actor, COMMIT, [("SSE", department.name, designation.name, 1, "NORMAL", "", "x")]
    ).json()["bulk_upload_log_id"]

    log = db_session.get(BulkUploadLog, uuid.UUID(log_id))
    assert log.entity_type == BulkUploadEntityTypeEnum.VACANCY_REQUEST
    assert log.rows_created == 1


def test_reuploading_the_same_file_creates_more_requests(client, bulk_setup, db_session):
    """The create-only property, end to end: a request is an event, so the
    second upload is two more real requests, not a no-op."""
    campus, department, designation, actor = bulk_setup
    row = ("SSE", department.name, designation.name, 1, "NORMAL", "", "Twice")

    _upload(client, actor, COMMIT, [row, row])
    _upload(client, actor, COMMIT, [row, row])

    assert db_session.query(VacancyRequest).filter(VacancyRequest.department_id == department.id).count() == 4


# --- undo ---------------------------------------------------------------------


def _undo(client, actor, log_id):
    return client.post(f"/api/v1/sanctioned-strength/bulk-uploads/{log_id}/undo", headers=auth_headers(client, actor))


def test_undo_cancels_the_drafts_it_created(client, bulk_setup, db_session):
    campus, department, designation, actor = bulk_setup
    log_id = _upload(
        client, actor, COMMIT, [("SSE", department.name, designation.name, 1, "NORMAL", "", "x")]
    ).json()["bulk_upload_log_id"]

    response = _undo(client, actor, log_id)
    assert response.status_code == 200, response.text
    assert response.json()["reverted_history_count"] == 1
    assert response.json()["not_reverted_count"] == 0

    vr = db_session.query(VacancyRequest).filter(VacancyRequest.department_id == department.id).one()
    # CANCELLED, not soft-deleted -- there is no is_active on this model.
    assert vr.status == VacancyRequestStatusEnum.CANCELLED
    assert vr.cancelled_by_id == actor.id
    assert "undone" in (vr.cancellation_reason or "").lower()


def test_undo_leaves_a_request_that_has_already_moved_on(client, bulk_setup, db_session):
    """A row can be submitted inside the 24h undo window. Cancelling it then
    would be destructive, not an undo -- it is counted, not touched."""
    campus, department, designation, actor = bulk_setup
    log_id = _upload(
        client,
        actor,
        COMMIT,
        [
            ("SSE", department.name, designation.name, 1, "NORMAL", "", "Stays draft"),
            ("SSE", department.name, designation.name, 1, "NORMAL", "", "Gets submitted"),
        ],
    ).json()["bulk_upload_log_id"]

    moved_on = (
        db_session.query(VacancyRequest)
        .filter(VacancyRequest.remarks == "Gets submitted")
        .one()
    )
    moved_on.status = VacancyRequestStatusEnum.SUBMITTED
    db_session.commit()

    body = _undo(client, actor, log_id).json()

    assert body["reverted_history_count"] == 1
    assert body["not_reverted_count"] == 1
    db_session.refresh(moved_on)
    assert moved_on.status == VacancyRequestStatusEnum.SUBMITTED


# --- referrer details ---------------------------------------------------------
#
# Optional columns appended to the template so an uploader can say who they are
# raising a row ON BEHALF OF. They fill the same
# `requester_name`/`_email`/`_mobile` columns the public QR intake writes;
# `requested_by_id` stays the uploader either way.


def test_template_carries_the_requester_columns(client, bulk_setup):
    from openpyxl import load_workbook

    _c, _d, _des, actor = bulk_setup
    response = client.get(TEMPLATE, headers=auth_headers(client, actor))

    ws = load_workbook(io.BytesIO(response.content))["Vacancy Requests"]
    header_row = tuple(cell.value for cell in ws[1])
    assert header_row == HEADERS_WITH_LOCATION


def test_a_seven_column_workbook_still_imports(client, bulk_setup, db_session):
    """The template that was downloadable before these columns existed. Cells
    are keyed by header name, so the three new keys are simply absent."""
    campus, department, designation, actor = bulk_setup

    body = _upload(
        client, actor, COMMIT, [("SSE", department.name, designation.name, 1, "NORMAL", "", "Old template")]
    ).json()

    assert body["created_count"] == 1
    vr = db_session.query(VacancyRequest).filter(VacancyRequest.remarks == "Old template").one()
    assert vr.requester_name is None
    assert vr.requester_email is None
    assert vr.requester_mobile is None


def test_commit_records_the_referrer_without_reassigning_the_row(client, bulk_setup, db_session):
    campus, department, designation, actor = bulk_setup

    response = _upload(
        client,
        actor,
        COMMIT,
        [
            (
                "SSE",
                department.name,
                designation.name,
                1,
                "NORMAL",
                "",
                "Referred vacancy",
                "Dr Referrer",
                "referrer@simats.ac.in",
                "+91 90000 00000",
            )
        ],
        HEADERS_WITH_REQUESTER,
    )
    assert response.status_code == 200, response.text

    vr = db_session.query(VacancyRequest).filter(VacancyRequest.remarks == "Referred vacancy").one()
    assert vr.requester_name == "Dr Referrer"
    assert vr.requester_email == "referrer@simats.ac.in"
    assert vr.requester_mobile == "+91 90000 00000"
    # Ownership does NOT move: requested_by_id is NOT NULL and five
    # notification sites dereference `.requested_by`.
    assert vr.requested_by_id == actor.id


def test_each_requester_column_is_independently_optional(client, bulk_setup, db_session):
    """A staff member may only have a name and a mobile for the person who
    referred the vacancy. Rejecting the row would throw away the detail they
    do have."""
    campus, department, designation, actor = bulk_setup

    body = _upload(
        client,
        actor,
        COMMIT,
        [("SSE", department.name, designation.name, 1, "NORMAL", "", "Partial", "Only A Name", "", "")],
        HEADERS_WITH_REQUESTER,
    ).json()

    assert body["created_count"] == 1
    vr = db_session.query(VacancyRequest).filter(VacancyRequest.remarks == "Partial").one()
    assert vr.requester_name == "Only A Name"
    assert vr.requester_email is None
    assert vr.requester_mobile is None


@pytest.mark.parametrize(
    "name, email, mobile, fragment",
    [
        ("A", "", "", "requester name"),
        ("", "not-an-email", "", "email"),
        ("", "", "call me", "mobile"),
    ],
)
def test_rejects_unreadable_requester_details(client, bulk_setup, name, email, mobile, fragment):
    campus, department, designation, actor = bulk_setup

    body = _upload(
        client,
        actor,
        VALIDATE,
        [("SSE", department.name, designation.name, 1, "NORMAL", "", "x", name, email, mobile)],
        HEADERS_WITH_REQUESTER,
    ).json()

    assert body["rejected_count"] == 1
    assert fragment in body["rows"][0]["error_reason"].lower()


def test_a_rejected_row_still_echoes_back_the_requester_it_was_given(client, bulk_setup):
    """So the preview table and the error report show what was typed rather
    than blanking it next to the reason it failed."""
    campus, department, designation, actor = bulk_setup

    body = _upload(
        client,
        actor,
        VALIDATE,
        [("XXX", department.name, designation.name, 1, "NORMAL", "", "x", "Named Referrer", "", "")],
        HEADERS_WITH_REQUESTER,
    ).json()

    assert body["rejected_count"] == 1
    assert body["rows"][0]["requester_name"] == "Named Referrer"


def test_error_report_for_a_vacancy_request_batch(client, bulk_setup):
    """The dialog's "Download error report" button reuses the SHARED endpoint
    on the sanctioned-strength router, whose `_import_module_for` dispatch had
    no VACANCY_REQUEST branch and so answered 500 for these batches. Pinned
    here because that button is the only caller.
    """
    from openpyxl import load_workbook

    campus, department, designation, actor = bulk_setup
    log_id = _upload(
        client,
        actor,
        COMMIT,
        [
            ("SSE", department.name, designation.name, 1, "NORMAL", "", "Good row", "", "", ""),
            ("ZZZ", department.name, designation.name, 1, "NORMAL", "", "Bad campus", "Named Referrer", "", ""),
        ],
        HEADERS_WITH_REQUESTER,
    ).json()["bulk_upload_log_id"]

    response = client.get(
        f"/api/v1/sanctioned-strength/bulk-upload/{log_id}/error-report",
        headers=auth_headers(client, actor),
    )
    assert response.status_code == 200, response.text

    ws = load_workbook(io.BytesIO(response.content)).active
    assert tuple(cell.value for cell in ws[1]) == ("Row", *HEADERS_WITH_LOCATION, "Error Reason")
    rejected = [row for row in ws.iter_rows(min_row=2, values_only=True) if any(cell is not None for cell in row)]
    assert len(rejected) == 1
    # Requester columns land under their own headings, and the reason stays in
    # the LAST column -- the pairing the report writer's strict zip enforces.
    assert rejected[0][1] == "ZZZ"
    assert rejected[0][8] == "Named Referrer"
    assert "campus" in rejected[0][-1].lower()


# --- the Location column (2026-09-02) ----------------------------------------
#
# Appended last, by the same rule as the referrer columns: an older workbook
# simply has no "Location" key and reads as blank. Whether blank is ACCEPTABLE
# is decided per campus, matching `vacancy_request_rules.validate_location` --
# which is why `bulk_setup` above (an SSE campus with no Location rows) still
# imports without one, and every pre-existing test in this file still passes.


def _location_row(department, designation, location_cell, justification="With location"):
    return ("SSE", department.name, designation.name, 1, "NORMAL", "", justification, "", "", "", location_cell)


@pytest.fixture()
def bulk_setup_with_location(bulk_setup, location_factory, db_session):
    campus, department, designation, actor = bulk_setup
    location = location_factory(
        "SSE", name="CB Block", block_building="Circular Building", floor_venue="Ground Floor"
    )
    db_session.commit()
    return campus, department, designation, actor, location


@pytest.mark.parametrize(
    "cell",
    [
        "Circular Building - Ground Floor",   # as the template's example writes it
        "Circular Building — Ground Floor",   # as the UI renders it, with an em dash
        "circular  building   ground floor",  # sloppy casing and spacing, no dash
        "CB Block Ground Floor",              # by `name` rather than `block_building`
    ],
)
def test_location_is_matched_forgivingly_and_stored(client, bulk_setup_with_location, db_session, cell):
    """Nobody should have to type an em dash into Excel."""
    _campus, department, designation, actor, location = bulk_setup_with_location

    body = _upload(
        client, actor, COMMIT,
        [_location_row(department, designation, cell, justification=f"Row {cell}")],
        HEADERS_WITH_LOCATION,
    ).json()

    assert body["created_count"] == 1, body
    vr = db_session.query(VacancyRequest).filter(VacancyRequest.remarks == f"Row {cell}").one()
    assert vr.location_id == location.id


def test_preview_echoes_the_canonical_label_not_what_was_typed(client, bulk_setup_with_location):
    _campus, department, designation, actor, _location = bulk_setup_with_location

    body = _upload(
        client, actor, VALIDATE,
        [_location_row(department, designation, "cb block ground floor")],
        HEADERS_WITH_LOCATION,
    ).json()

    assert body["rows"][0]["status"] == "created"
    assert body["rows"][0]["location_name"] == "Circular Building - Ground Floor"


def test_location_is_required_when_the_campus_has_locations(client, bulk_setup_with_location):
    _campus, department, designation, actor, _location = bulk_setup_with_location

    body = _upload(
        client, actor, VALIDATE,
        [_location_row(department, designation, "")],
        HEADERS_WITH_LOCATION,
    ).json()

    assert body["rejected_count"] == 1
    assert body["rows"][0]["error_reason"] == "Location is required on campus SSE."


def test_location_stays_optional_on_a_campus_with_none(client, bulk_setup, db_session):
    """`bulk_setup`'s SSE campus has no Location rows. Five of seven
    production campuses are in exactly this state."""
    _campus, department, designation, actor = bulk_setup

    body = _upload(
        client, actor, COMMIT,
        [_location_row(department, designation, "", justification="No locations here")],
        HEADERS_WITH_LOCATION,
    ).json()

    assert body["created_count"] == 1, body
    vr = db_session.query(VacancyRequest).filter(VacancyRequest.remarks == "No locations here").one()
    assert vr.location_id is None


def test_an_unknown_location_is_rejected_with_the_campus_named(client, bulk_setup_with_location):
    _campus, department, designation, actor, _location = bulk_setup_with_location

    body = _upload(
        client, actor, VALIDATE,
        [_location_row(department, designation, "Nowhere Block")],
        HEADERS_WITH_LOCATION,
    ).json()

    assert body["rejected_count"] == 1
    assert body["rows"][0]["error_reason"] == "Unknown location 'Nowhere Block' on campus SSE."


def test_a_location_on_another_campus_does_not_resolve(
    client, bulk_setup_with_location, location_factory, db_session
):
    """The index is keyed on (campus, alias): two campuses may each have a
    "Main Block - Ground Floor" and those must never resolve to each other."""
    _campus, department, designation, actor, _location = bulk_setup_with_location
    location_factory("SCAD", name="Design Wing", block_building="Design Wing", floor_venue="First Floor")
    db_session.commit()

    body = _upload(
        client, actor, VALIDATE,
        [_location_row(department, designation, "Design Wing - First Floor")],
        HEADERS_WITH_LOCATION,
    ).json()

    assert body["rejected_count"] == 1
    assert "Unknown location" in body["rows"][0]["error_reason"]


def test_a_block_name_covering_several_floors_is_rejected_as_ambiguous(
    client, bulk_setup_with_location, location_factory, db_session
):
    """Real data has six rows named "CB Block" differing only by floor.
    Naming the block alone does not say which one is meant."""
    _campus, department, designation, actor, _location = bulk_setup_with_location
    location_factory("SSE", name="CB Block", block_building="Circular Building", floor_venue="First Floor")
    db_session.commit()

    body = _upload(
        client, actor, VALIDATE,
        [_location_row(department, designation, "Circular Building")],
        HEADERS_WITH_LOCATION,
    ).json()

    assert body["rejected_count"] == 1
    reason = body["rows"][0]["error_reason"]
    assert "matches more than one location" in reason
    assert "Circular Building - Ground Floor" in reason
    assert "Circular Building - First Floor" in reason
    assert "Include the floor" in reason


def test_duplicate_master_data_for_one_place_resolves_rather_than_rejecting(
    client, bulk_setup_with_location, location_factory, db_session
):
    """Two Location rows describing the SAME physical place is duplicate
    master data, not a choice -- the UI's picker collapses those too, and
    rejecting the row would push a data-quality problem onto the uploader."""
    _campus, department, designation, actor, location = bulk_setup_with_location
    twin = location_factory(
        "SSE", name="CB Block", block_building="circular building", floor_venue="ground floor"
    )
    db_session.commit()

    body = _upload(
        client, actor, COMMIT,
        [_location_row(department, designation, "Circular Building - Ground Floor", justification="Twin")],
        HEADERS_WITH_LOCATION,
    ).json()

    assert body["created_count"] == 1, body
    vr = db_session.query(VacancyRequest).filter(VacancyRequest.remarks == "Twin").one()
    # Deterministic tie-break on the smallest id, matching
    # dedupeLocationsForPicker, so a re-upload always picks the same row.
    assert vr.location_id == min([location.id, twin.id], key=str)


def test_template_lists_the_valid_locations_per_campus(client, bulk_setup_with_location):
    """Location Master is not otherwise visible from a spreadsheet, so without
    this the uploader has no way to know what to type."""
    from openpyxl import load_workbook

    _campus, _department, _designation, actor, _location = bulk_setup_with_location
    response = client.get(TEMPLATE, headers=auth_headers(client, actor))

    reference = load_workbook(io.BytesIO(response.content))["Reference"]
    listed = {
        (reference[f"F{r}"].value, reference[f"G{r}"].value)
        for r in range(2, reference.max_row + 1)
        if reference[f"G{r}"].value
    }
    assert ("SSE", "Circular Building - Ground Floor") in listed
