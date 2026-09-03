"""Formula-injection hardening of every workbook the app hands out
(audit H3, 2026-09-03). See app/services/xlsx_safety.py.

Two layers: the rule itself (`safe_cell`), and the sink (`harden_workbook`)
proven through real endpoints -- a master-data export and a report export
-- so the guarantee is checked where user-controlled text actually reaches
a spreadsheet, not only in isolation. Bulk-upload error reports are covered
in test_bulk_upload_shared_endpoints.py.
"""

import io
from datetime import datetime, timezone

import pytest
from openpyxl import Workbook, load_workbook

from app.models.enums import UserRoleEnum
from app.services import exports
from app.services.xlsx_safety import harden_workbook, safe_cell
from tests.conftest import auth_headers

DANGEROUS = ["=1+1", "+123", "-123", "@test", "\tpadded", "\rreturned"]


@pytest.mark.parametrize("value", DANGEROUS)
def test_safe_cell_escapes_every_formula_and_separator_prefix(value):
    assert safe_cell(value) == "'" + value


@pytest.mark.parametrize("value", ["normal text", "Dr. A, Dr. B", " =not a formula", "", "1+1=2", "a-b"])
def test_safe_cell_leaves_ordinary_text_alone(value):
    assert safe_cell(value) == value


@pytest.mark.parametrize("value", [123, -5, 1.5, None, datetime(2026, 9, 3, tzinfo=timezone.utc)])
def test_safe_cell_never_touches_non_strings(value):
    assert safe_cell(value) is value


def _reload(workbook: Workbook):
    buf = io.BytesIO()
    workbook.save(buf)
    return load_workbook(io.BytesIO(buf.getvalue())).active


def test_harden_workbook_retypes_a_formula_cell_as_text():
    wb = Workbook()
    ws = wb.active
    ws.append(DANGEROUS + ["normal text", 42])
    # openpyxl really does write "=1+1" as a formula -- the thing being fixed.
    assert ws.cell(row=1, column=1).data_type == "f"

    sheet = _reload(harden_workbook(wb))
    for col, value in enumerate(DANGEROUS, start=1):
        cell = sheet.cell(row=1, column=col)
        assert cell.data_type == "s", value
        assert cell.value == "'" + value
    assert sheet.cell(row=1, column=7).value == "normal text"
    assert sheet.cell(row=1, column=8).value == 42


def test_harden_workbook_covers_every_sheet():
    wb = Workbook()
    wb.active.append(["=first"])
    wb.create_sheet("Second").append(["=second"])
    harden_workbook(wb)
    assert wb["Sheet"]["A1"].value == "'=first"
    assert wb["Second"]["A1"].value == "'=second"


def test_report_export_escapes_user_text_in_every_column():
    rows = [{"campus_code": "=1+1", "department_name": "+123", "role_category": "@test", "count": -5}]
    sheet = load_workbook(
        io.BytesIO(exports.build_report_excel("resignations", rows, datetime.now(timezone.utc), "Scope: test"))
    ).active
    data = [row for row in sheet.iter_rows(values_only=True) if row and row[0] not in (None, "")]
    last = data[-1]
    assert last[:3] == ("'=1+1", "'+123", "'@test")
    assert last[3] == -5  # numbers stay numbers


def test_locations_export_escapes_a_malicious_location_name(client, user_factory, location_factory, db_session):
    location_factory("SSE", name='=HYPERLINK("http://evil.example","Open")', block_building="+Block", floor_venue="-1")
    db_session.commit()
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)

    response = client.get("/api/v1/locations/export", headers=auth_headers(client, hr_admin))
    assert response.status_code == 200
    sheet = load_workbook(io.BytesIO(response.content)).active

    flat = [cell for row in sheet.iter_rows() for cell in row if isinstance(cell.value, str)]
    dangerous = [c for c in flat if c.value.startswith(("=", "+", "-", "@", "\t", "\r"))]
    assert dangerous == [], [c.value for c in dangerous]
    assert all(c.data_type == "s" for c in flat)
    escaped = {c.value for c in flat}
    assert "'=HYPERLINK(\"http://evil.example\",\"Open\")" in escaped
    assert "'+Block" in escaped and "'-1" in escaped
