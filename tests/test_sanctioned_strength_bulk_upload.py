"""Sanctioned Strength bulk upload (zany-snuggling-pie.md Phase F) --
validate/commit/template/error-report/list/original-file/undo, backed by
app/services/sanctioned_strength_import.py and the new endpoints in
app/api/v1/routers/sanctioned_strength.py.
"""

import csv
import io
import uuid
from datetime import date, datetime, timedelta, timezone

import pytest
from sqlalchemy.exc import IntegrityError

from app.models.bulk_upload_log import BulkUploadLog
from app.models.enums import BulkUploadStatusEnum, StaffRoleCategoryEnum, UserRoleEnum
from app.models.sanctioned_strength import SanctionedStrength, SanctionedStrengthHistory
from app.services import sanctioned_strength_import as ssi

from tests.conftest import auth_headers

ENDPOINT = "/api/v1/sanctioned-strength"

HEADERS = [
    "Campus Code",
    "Department Name",
    "Designation Name",
    "Approved Strength",
    "Effective From (DD-MM-YYYY)",
    "Remarks",
]


def _csv_bytes(rows: list[list]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(HEADERS)
    for row in rows:
        writer.writerow(row)
    return buf.getvalue().encode("utf-8")


def _upload_validate(client, actor, rows):
    return client.post(
        f"{ENDPOINT}/bulk-upload/validate",
        files={"file": ("upload.csv", _csv_bytes(rows), "text/csv")},
        headers=auth_headers(client, actor),
    )


def _upload_commit(client, actor, rows, filename: str = "upload.csv"):
    return client.post(
        f"{ENDPOINT}/bulk-upload/commit",
        files={"file": (filename, _csv_bytes(rows), "text/csv")},
        headers=auth_headers(client, actor),
    )


@pytest.fixture()
def bulk_upload_setup(campus_factory, department_factory, designation_factory, db_session, user_factory):
    campus = campus_factory("SSE")
    department = department_factory("SSE", name=f"Bulk Dept {uuid.uuid4().hex[:6]}")
    department.supported_categories = [StaffRoleCategoryEnum.TEACHING]
    designation = designation_factory(
        StaffRoleCategoryEnum.TEACHING, name=f"Bulk Designation {uuid.uuid4().hex[:6]}", department=department
    )
    db_session.flush()
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)
    return {"campus": campus, "department": department, "designation": designation, "hr_admin": hr_admin}


def _row(setup, strength=5, effective_from="01-04-2026", remarks=""):
    return [
        setup["campus"].code,
        setup["department"].name,
        setup["designation"].name,
        strength,
        effective_from,
        remarks,
    ]


# --- validate: read-only, DB untouched ---------------------------------------------------------


def test_validate_new_row_is_created_and_writes_nothing(client, bulk_upload_setup, db_session):
    response = _upload_validate(client, bulk_upload_setup["hr_admin"], [_row(bulk_upload_setup)])
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["created_count"] == 1
    assert body["updated_count"] == 0
    assert body["unchanged_count"] == 0
    assert body["rejected_count"] == 0
    assert body["rows"][0]["status"] == "created"

    assert db_session.query(SanctionedStrength).count() == 0
    assert db_session.query(SanctionedStrengthHistory).count() == 0
    assert db_session.query(BulkUploadLog).count() == 0


def test_validate_existing_row_same_value_is_unchanged(client, bulk_upload_setup, sanctioned_strength_factory):
    sanctioned_strength_factory(
        campus=bulk_upload_setup["campus"],
        department=bulk_upload_setup["department"],
        designation=bulk_upload_setup["designation"],
        approved_strength=5,
        created_by=bulk_upload_setup["hr_admin"],
    )
    response = _upload_validate(client, bulk_upload_setup["hr_admin"], [_row(bulk_upload_setup, strength=5)])
    assert response.status_code == 200
    body = response.json()
    assert body["unchanged_count"] == 1
    assert body["rows"][0]["status"] == "unchanged"


def test_validate_existing_row_new_value_is_updated(client, bulk_upload_setup, sanctioned_strength_factory):
    sanctioned_strength_factory(
        campus=bulk_upload_setup["campus"],
        department=bulk_upload_setup["department"],
        designation=bulk_upload_setup["designation"],
        approved_strength=5,
        created_by=bulk_upload_setup["hr_admin"],
    )
    response = _upload_validate(client, bulk_upload_setup["hr_admin"], [_row(bulk_upload_setup, strength=9)])
    assert response.status_code == 200
    body = response.json()
    assert body["updated_count"] == 1
    assert body["rows"][0]["status"] == "updated"


def test_validate_unknown_campus_is_rejected(client, bulk_upload_setup):
    row = _row(bulk_upload_setup)
    row[0] = "ZZZ"
    response = _upload_validate(client, bulk_upload_setup["hr_admin"], [row])
    assert response.status_code == 200
    body = response.json()
    assert body["rejected_count"] == 1
    assert "campus" in body["rows"][0]["error_reason"].lower()


def test_validate_unknown_department_is_rejected(client, bulk_upload_setup):
    row = _row(bulk_upload_setup)
    row[1] = "Nonexistent Department XYZ"
    response = _upload_validate(client, bulk_upload_setup["hr_admin"], [row])
    body = response.json()
    assert body["rejected_count"] == 1
    assert "department" in body["rows"][0]["error_reason"].lower()


def test_validate_unknown_designation_is_rejected(client, bulk_upload_setup):
    row = _row(bulk_upload_setup)
    row[2] = "Nonexistent Designation XYZ"
    response = _upload_validate(client, bulk_upload_setup["hr_admin"], [row])
    body = response.json()
    assert body["rejected_count"] == 1
    assert "designation" in body["rows"][0]["error_reason"].lower()


def test_validate_category_mismatch_is_rejected(
    client, campus_factory, department_factory, designation_factory, user_factory, db_session
):
    campus = campus_factory("SSE")
    department = department_factory("SSE", name=f"NonTeach Dept {uuid.uuid4().hex[:6]}")
    department.supported_categories = [StaffRoleCategoryEnum.NON_TEACHING]
    designation = designation_factory(StaffRoleCategoryEnum.TEACHING, department=department)
    db_session.flush()
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)

    row = [campus.code, department.name, designation.name, 5, "01-04-2026", ""]
    response = _upload_validate(client, hr_admin, [row])
    body = response.json()
    assert body["rejected_count"] == 1
    assert "does not support" in body["rows"][0]["error_reason"].lower()


def test_validate_negative_strength_is_rejected(client, bulk_upload_setup):
    response = _upload_validate(client, bulk_upload_setup["hr_admin"], [_row(bulk_upload_setup, strength=-3)])
    body = response.json()
    assert body["rejected_count"] == 1


def test_validate_non_integer_strength_is_rejected(client, bulk_upload_setup):
    row = _row(bulk_upload_setup)
    row[3] = "five"
    response = _upload_validate(client, bulk_upload_setup["hr_admin"], [row])
    body = response.json()
    assert body["rejected_count"] == 1


def test_validate_malformed_date_is_rejected(client, bulk_upload_setup):
    row = _row(bulk_upload_setup, effective_from="2026-04-01")  # wrong format, not DD-MM-YYYY
    response = _upload_validate(client, bulk_upload_setup["hr_admin"], [row])
    body = response.json()
    assert body["rejected_count"] == 1
    assert "date" in body["rows"][0]["error_reason"].lower()


def test_validate_in_file_duplicate_key_is_rejected(client, bulk_upload_setup):
    row = _row(bulk_upload_setup)
    response = _upload_validate(client, bulk_upload_setup["hr_admin"], [row, list(row)])
    body = response.json()
    assert body["created_count"] == 1
    assert body["rejected_count"] == 1
    assert "duplicate" in body["rows"][1]["error_reason"].lower()


def test_validate_forbidden_for_non_write_role(client, bulk_upload_setup, user_factory):
    hod = user_factory(UserRoleEnum.CAMPUS_HOD, campus_code="SSE")
    response = _upload_validate(client, hod, [_row(bulk_upload_setup)])
    assert response.status_code == 403


# --- commit: writes, history, idempotency, all-or-nothing -------------------------------------


def test_commit_creates_row_and_history_and_log(client, bulk_upload_setup, db_session):
    response = _upload_commit(client, bulk_upload_setup["hr_admin"], [_row(bulk_upload_setup)])
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["created_count"] == 1
    log_id = uuid.UUID(body["bulk_upload_log_id"])

    row = db_session.query(SanctionedStrength).one()
    assert row.approved_strength == 5
    assert row.campus_id == bulk_upload_setup["campus"].id

    history = db_session.query(SanctionedStrengthHistory).filter(
        SanctionedStrengthHistory.sanctioned_strength_id == row.id
    ).one()
    assert history.old_value is None
    assert history.new_value == 5
    assert history.source.value == "BULK_UPLOAD"
    assert history.bulk_upload_log_id == log_id

    log = db_session.get(BulkUploadLog, log_id)
    assert log.rows_created == 1
    assert log.rows_total == 1
    assert log.status == BulkUploadStatusEnum.COMPLETED
    assert log.stored_file_object_key is not None
    assert log.undo_deadline > datetime.now(timezone.utc)


def test_commit_succeeds_with_a_storage_warning_when_object_storage_is_unavailable(
    client, bulk_upload_setup, db_session, fake_minio_client
):
    """Same fix as Location's own bulk upload (see
    test_location_bulk_upload.py's identically-named test for the full bug
    writeup) -- object storage being unreachable must never block a
    validated row from actually committing."""
    fake_minio_client.fail_puts = True

    response = _upload_commit(client, bulk_upload_setup["hr_admin"], [_row(bulk_upload_setup)])
    assert response.status_code == 200, response.text
    body = response.json()

    assert body["created_count"] == 1
    assert db_session.query(SanctionedStrength).count() == 1
    assert body["storage_warning"] == (
        "Workbook storage is temporarily unavailable. The file was successfully parsed, "
        "but the original workbook could not be archived."
    )

    log = db_session.get(BulkUploadLog, uuid.UUID(body["bulk_upload_log_id"]))
    assert log.rows_created == 1
    assert log.stored_file_object_key is None


def test_commit_skips_rejected_rows_entirely(client, bulk_upload_setup, db_session):
    good = _row(bulk_upload_setup)
    bad = _row(bulk_upload_setup)
    bad[0] = "ZZZ"
    response = _upload_commit(client, bulk_upload_setup["hr_admin"], [good, bad])
    assert response.status_code == 200
    body = response.json()
    assert body["created_count"] == 1
    assert body["rejected_count"] == 1
    assert db_session.query(SanctionedStrength).count() == 1


def test_commit_updates_existing_row_and_writes_history(client, bulk_upload_setup, sanctioned_strength_factory, db_session):
    existing = sanctioned_strength_factory(
        campus=bulk_upload_setup["campus"],
        department=bulk_upload_setup["department"],
        designation=bulk_upload_setup["designation"],
        approved_strength=5,
        created_by=bulk_upload_setup["hr_admin"],
    )
    db_session.commit()

    response = _upload_commit(client, bulk_upload_setup["hr_admin"], [_row(bulk_upload_setup, strength=9)])
    assert response.status_code == 200
    body = response.json()
    assert body["updated_count"] == 1

    db_session.refresh(existing)
    assert existing.approved_strength == 9

    history = (
        db_session.query(SanctionedStrengthHistory)
        .filter(SanctionedStrengthHistory.sanctioned_strength_id == existing.id)
        .all()
    )
    assert len(history) == 1
    assert history[0].old_value == 5
    assert history[0].new_value == 9


def test_reuploading_identical_file_twice_writes_no_new_history_second_time(client, bulk_upload_setup, db_session):
    rows = [_row(bulk_upload_setup)]
    first = _upload_commit(client, bulk_upload_setup["hr_admin"], rows)
    assert first.status_code == 200
    db_session.commit()

    history_count_after_first = db_session.query(SanctionedStrengthHistory).count()
    assert history_count_after_first == 1

    second = _upload_commit(client, bulk_upload_setup["hr_admin"], rows)
    assert second.status_code == 200, second.text
    body = second.json()
    assert body["unchanged_count"] == 1
    assert body["created_count"] == 0
    assert body["updated_count"] == 0

    db_session.commit()
    assert db_session.query(SanctionedStrengthHistory).count() == history_count_after_first
    assert db_session.query(SanctionedStrength).count() == 1
    assert db_session.query(BulkUploadLog).count() == 2  # one log row per commit call, even when a no-op


def test_commit_forbidden_for_non_write_role(client, bulk_upload_setup, user_factory):
    hod = user_factory(UserRoleEnum.CAMPUS_HOD, campus_code="SSE")
    response = _upload_commit(client, hod, [_row(bulk_upload_setup)])
    assert response.status_code == 403


def test_commit_rows_is_all_or_nothing_defense_in_depth(
    client, bulk_upload_setup, department_factory, designation_factory, db_session, user_factory
):
    """Simulates a second row that slipped past validate_rows (e.g. a
    hand-built ValidationResult referencing a designation_id that doesn't
    actually exist -- an IntegrityError on the FK) to prove commit_rows()
    doesn't partially persist: an earlier, otherwise-valid row in the same
    batch must not survive the rollback triggered by a later row's
    failure."""
    campus = bulk_upload_setup["campus"]
    department = bulk_upload_setup["department"]
    designation = bulk_upload_setup["designation"]
    actor = bulk_upload_setup["hr_admin"]

    other_department = department_factory("SSE", name=f"Defense Dept {uuid.uuid4().hex[:6]}")
    other_department.supported_categories = [StaffRoleCategoryEnum.TEACHING]
    other_designation = designation_factory(StaffRoleCategoryEnum.TEACHING, department=other_department)
    db_session.flush()

    row_ok = ssi.ImportRowResult(
        row_number=2,
        status="created",
        campus_id=campus.id,
        department_id=department.id,
        designation_id=designation.id,
        category=designation.category,
        approved_strength=4,
        effective_from=date(2026, 4, 1),
    )
    row_bad = ssi.ImportRowResult(
        row_number=3,
        status="created",
        campus_id=campus.id,
        department_id=other_department.id,
        designation_id=uuid.uuid4(),  # never inserted -- violates the designation FK
        category=other_designation.category,
        approved_strength=7,
        effective_from=date(2026, 4, 1),
    )
    validation = ssi.ValidationResult(
        rows=[row_ok, row_bad], total=2, created_count=2, updated_count=0, unchanged_count=0, rejected_count=0
    )
    log = BulkUploadLog(
        filename="defense-in-depth.csv",
        uploaded_by_id=actor.id,
        rows_total=2,
        rows_created=2,
        rows_updated=0,
        rows_rejected=0,
        undo_deadline=datetime.now(timezone.utc) + timedelta(hours=24),
    )
    db_session.add(log)
    db_session.flush()

    with pytest.raises(IntegrityError):
        ssi.commit_rows(db_session, validation=validation, actor=actor, bulk_upload_log_id=log.id)
    db_session.rollback()

    assert (
        db_session.query(SanctionedStrength)
        .filter(
            SanctionedStrength.campus_id == campus.id,
            SanctionedStrength.department_id == department.id,
            SanctionedStrength.designation_id == designation.id,
        )
        .count()
        == 0
    )
    assert (
        db_session.query(SanctionedStrength)
        .filter(SanctionedStrength.department_id == other_department.id)
        .count()
        == 0
    )


# --- template / error-report / list / original-file / undo ------------------------------------


def test_template_download_is_xlsx(client, bulk_upload_setup):
    response = client.get(f"{ENDPOINT}/bulk-upload/template", headers=auth_headers(client, bulk_upload_setup["hr_admin"]))
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in response.headers["content-disposition"]


def test_template_forbidden_for_non_write_role(client, bulk_upload_setup, user_factory):
    hod = user_factory(UserRoleEnum.CAMPUS_HOD, campus_code="SSE")
    response = client.get(f"{ENDPOINT}/bulk-upload/template", headers=auth_headers(client, hod))
    assert response.status_code == 403


def test_error_report_contains_only_rejected_rows(client, bulk_upload_setup, db_session):
    good = _row(bulk_upload_setup)
    bad = _row(bulk_upload_setup)
    bad[0] = "ZZZ"
    commit_response = _upload_commit(client, bulk_upload_setup["hr_admin"], [good, bad])
    assert commit_response.status_code == 200, commit_response.text
    log_id = commit_response.json()["bulk_upload_log_id"]

    response = client.get(
        f"{ENDPOINT}/bulk-upload/{log_id}/error-report", headers=auth_headers(client, bulk_upload_setup["hr_admin"])
    )
    assert response.status_code == 200

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    data_rows = list(ws.iter_rows(min_row=5, values_only=True))
    non_empty = [r for r in data_rows if any(c is not None for c in r)]
    assert len(non_empty) == 1
    assert non_empty[0][0] == "ZZZ"


def test_list_bulk_uploads_returns_committed_batches(client, bulk_upload_setup):
    _upload_commit(client, bulk_upload_setup["hr_admin"], [_row(bulk_upload_setup)])
    response = client.get(f"{ENDPOINT}/bulk-uploads", headers=auth_headers(client, bulk_upload_setup["hr_admin"]))
    assert response.status_code == 200
    body = response.json()
    assert body["total"] >= 1


def test_original_file_download_round_trips(client, bulk_upload_setup):
    commit_response = _upload_commit(client, bulk_upload_setup["hr_admin"], [_row(bulk_upload_setup)])
    log_id = commit_response.json()["bulk_upload_log_id"]

    response = client.get(
        f"{ENDPOINT}/bulk-uploads/{log_id}/original-file", headers=auth_headers(client, bulk_upload_setup["hr_admin"])
    )
    assert response.status_code == 200
    assert response.content == _csv_bytes([_row(bulk_upload_setup)])


def test_undo_within_deadline_reverts_created_row_by_deactivating(client, bulk_upload_setup, db_session):
    commit_response = _upload_commit(client, bulk_upload_setup["hr_admin"], [_row(bulk_upload_setup)])
    log_id = commit_response.json()["bulk_upload_log_id"]
    row_id = db_session.query(SanctionedStrength).one().id

    response = client.post(
        f"{ENDPOINT}/bulk-uploads/{log_id}/undo", headers=auth_headers(client, bulk_upload_setup["hr_admin"])
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["status"] == "UNDONE"
    assert body["reverted_history_count"] == 1

    db_session.refresh(db_session.get(SanctionedStrength, row_id))
    row = db_session.get(SanctionedStrength, row_id)
    assert row.is_active is False

    log = db_session.get(BulkUploadLog, uuid.UUID(log_id))
    assert log.status == BulkUploadStatusEnum.UNDONE
    assert log.undone_by_id == bulk_upload_setup["hr_admin"].id


def test_undo_reverts_updated_row_to_old_value(client, bulk_upload_setup, sanctioned_strength_factory, db_session):
    existing = sanctioned_strength_factory(
        campus=bulk_upload_setup["campus"],
        department=bulk_upload_setup["department"],
        designation=bulk_upload_setup["designation"],
        approved_strength=5,
        created_by=bulk_upload_setup["hr_admin"],
    )
    db_session.commit()

    commit_response = _upload_commit(client, bulk_upload_setup["hr_admin"], [_row(bulk_upload_setup, strength=9)])
    log_id = commit_response.json()["bulk_upload_log_id"]

    response = client.post(
        f"{ENDPOINT}/bulk-uploads/{log_id}/undo", headers=auth_headers(client, bulk_upload_setup["hr_admin"])
    )
    assert response.status_code == 200

    db_session.refresh(existing)
    assert existing.approved_strength == 5


def test_undo_after_deadline_is_rejected(client, bulk_upload_setup, db_session):
    commit_response = _upload_commit(client, bulk_upload_setup["hr_admin"], [_row(bulk_upload_setup)])
    log_id = commit_response.json()["bulk_upload_log_id"]

    log = db_session.get(BulkUploadLog, uuid.UUID(log_id))
    log.undo_deadline = datetime.now(timezone.utc) - timedelta(hours=1)
    db_session.flush()

    response = client.post(
        f"{ENDPOINT}/bulk-uploads/{log_id}/undo", headers=auth_headers(client, bulk_upload_setup["hr_admin"])
    )
    assert response.status_code == 409
    assert "expired" in response.json()["detail"].lower()


def test_undo_twice_is_rejected(client, bulk_upload_setup):
    commit_response = _upload_commit(client, bulk_upload_setup["hr_admin"], [_row(bulk_upload_setup)])
    log_id = commit_response.json()["bulk_upload_log_id"]

    first = client.post(f"{ENDPOINT}/bulk-uploads/{log_id}/undo", headers=auth_headers(client, bulk_upload_setup["hr_admin"]))
    assert first.status_code == 200

    second = client.post(f"{ENDPOINT}/bulk-uploads/{log_id}/undo", headers=auth_headers(client, bulk_upload_setup["hr_admin"]))
    assert second.status_code == 409
    assert "already" in second.json()["detail"].lower()


def test_undo_forbidden_for_non_write_role(client, bulk_upload_setup, user_factory):
    commit_response = _upload_commit(client, bulk_upload_setup["hr_admin"], [_row(bulk_upload_setup)])
    log_id = commit_response.json()["bulk_upload_log_id"]
    hod = user_factory(UserRoleEnum.CAMPUS_HOD, campus_code="SSE")

    response = client.post(f"{ENDPOINT}/bulk-uploads/{log_id}/undo", headers=auth_headers(client, hod))
    assert response.status_code == 403


def test_undo_unknown_id_is_404(client, bulk_upload_setup):
    response = client.post(
        f"{ENDPOINT}/bulk-uploads/{uuid.uuid4()}/undo", headers=auth_headers(client, bulk_upload_setup["hr_admin"])
    )
    assert response.status_code == 404
