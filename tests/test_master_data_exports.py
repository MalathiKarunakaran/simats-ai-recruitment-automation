"""xlsx exports for the three master-data screens that had bulk upload but no
export until 2026-08-27: Locations, Housekeeping Staff, and the three
Sanctioned Strength views.

Departments'/Designations'/EligibilityRules' own export tests live in their
respective test modules; these are the newer siblings, grouped here because
they share one export core (`exports._build_master_export_excel`) and one
scope-note helper (`deps.campus_scope_note`).

Every assertion reads the real workbook back with openpyxl rather than
trusting the byte count -- the whole point of an export is that the file
opens and carries the right rows.
"""

import io

import pytest
from openpyxl import load_workbook

from app.models.enums import HousekeepingShiftEnum, StaffRoleCategoryEnum, UserRoleEnum

from tests.conftest import auth_headers

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _sheet(response):
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(XLSX_CONTENT_TYPE)
    workbook = load_workbook(io.BytesIO(response.content))
    return workbook.active


def _data_rows(sheet):
    """Row 1 is `Generated:`, row 2 is `Scope:`, row 3 blank, row 4 headers --
    the shared layout every export in this app uses."""
    rows = list(sheet.iter_rows(min_row=5, values_only=True))
    return [row for row in rows if any(cell is not None for cell in row)]


def _headers(sheet):
    return [cell for cell in next(sheet.iter_rows(min_row=4, max_row=4, values_only=True)) if cell is not None]


# --- Locations --------------------------------------------------------------


def test_export_locations_returns_xlsx_with_scope_line(client, user_factory, location_factory):
    location_factory("SSE", name="Main Block", category=StaffRoleCategoryEnum.TEACHING)
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)

    sheet = _sheet(client.get("/api/v1/locations/export", headers=auth_headers(client, hr_admin)))

    assert sheet.cell(row=1, column=1).value.startswith("Generated: ")
    assert sheet.cell(row=2, column=1).value.startswith("Scope: ")
    assert _headers(sheet)[:2] == ["Campus Code", "Location Name"]
    assert any(row[1] == "Main Block" for row in _data_rows(sheet))


def test_export_locations_respects_category_filter(client, user_factory, location_factory):
    location_factory("SSE", name="Lecture Hall", category=StaffRoleCategoryEnum.TEACHING)
    location_factory("SSE", name="Store Room", category=StaffRoleCategoryEnum.NON_TEACHING)
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)

    sheet = _sheet(
        client.get(
            "/api/v1/locations/export",
            params={"category": "TEACHING"},
            headers=auth_headers(client, hr_admin),
        )
    )

    names = [row[1] for row in _data_rows(sheet)]
    assert "Lecture Hall" in names
    assert "Store Room" not in names


def test_export_locations_excludes_inactive_unless_asked(client, user_factory, location_factory, db_session):
    active = location_factory("SSE", name="Active Hall", category=StaffRoleCategoryEnum.TEACHING)
    retired = location_factory("SSE", name="Retired Hall", category=StaffRoleCategoryEnum.TEACHING)
    retired.is_active = False
    db_session.flush()
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)

    default_names = [
        row[1] for row in _data_rows(_sheet(client.get("/api/v1/locations/export", headers=auth_headers(client, hr_admin))))
    ]
    assert active.name in default_names
    assert "Retired Hall" not in default_names

    with_inactive = [
        row[1]
        for row in _data_rows(
            _sheet(
                client.get(
                    "/api/v1/locations/export",
                    params={"include_inactive": "true"},
                    headers=auth_headers(client, hr_admin),
                )
            )
        )
    ]
    assert "Retired Hall" in with_inactive


def test_export_locations_is_campus_scoped_for_non_global_role(
    client, user_factory, location_factory, campus_factory
):
    """A campus-scoped role must never receive another campus's rows, and the
    Scope line must say so rather than implying the file is complete."""
    location_factory("SSE", name="SSE Hall", category=StaffRoleCategoryEnum.TEACHING)
    location_factory("SCAD", name="SCAD Hall", category=StaffRoleCategoryEnum.TEACHING)
    hod = user_factory(UserRoleEnum.CAMPUS_HOD, campus_code="SSE")

    sheet = _sheet(client.get("/api/v1/locations/export", headers=auth_headers(client, hod)))

    names = [row[1] for row in _data_rows(sheet)]
    assert "SSE Hall" in names
    assert "SCAD Hall" not in names
    assert "home campus" in sheet.cell(row=2, column=1).value


def test_export_locations_forbidden_for_candidate(client, user_factory):
    candidate = user_factory(UserRoleEnum.CANDIDATE)
    response = client.get("/api/v1/locations/export", headers=auth_headers(client, candidate))
    assert response.status_code == 403


# --- Housekeeping Staff -----------------------------------------------------


@pytest.fixture()
def housekeeping_setup(campus_factory, department_factory, designation_factory, location_factory, user_factory):
    campus = campus_factory("SSE")
    department = department_factory("SSE", name="Facilities", code="FAC", category=StaffRoleCategoryEnum.HOUSEKEEPING)
    designation = designation_factory(
        name="Housekeeper", category=StaffRoleCategoryEnum.HOUSEKEEPING, department=department
    )
    location = location_factory("SSE", name="Block A", category=StaffRoleCategoryEnum.HOUSEKEEPING)
    admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    return {"campus": campus, "designation": designation, "location": location, "admin": admin}


def test_export_housekeeping_staff_returns_rows(client, housekeeping_setup, housekeeping_staff_factory):
    housekeeping_staff_factory(
        campus=housekeeping_setup["campus"],
        designation=housekeeping_setup["designation"],
        location=housekeeping_setup["location"],
        created_by=housekeeping_setup["admin"],
        name="Asha R",
        supervisor="Mr Kumar",
        shift=HousekeepingShiftEnum.MORNING,
    )

    sheet = _sheet(
        client.get("/api/v1/housekeeping-staff/export", headers=auth_headers(client, housekeeping_setup["admin"]))
    )

    assert _headers(sheet)[:3] == ["Campus Code", "Bio ID", "Name"]
    rows = _data_rows(sheet)
    assert any(row[2] == "Asha R" and row[8] == "Mr Kumar" for row in rows)


def test_export_housekeeping_staff_respects_shift_filter(
    client, housekeeping_setup, housekeeping_staff_factory
):
    housekeeping_staff_factory(
        campus=housekeeping_setup["campus"],
        designation=housekeeping_setup["designation"],
        location=housekeeping_setup["location"],
        created_by=housekeeping_setup["admin"],
        name="Morning Person",
        shift=HousekeepingShiftEnum.MORNING,
    )
    housekeeping_staff_factory(
        campus=housekeeping_setup["campus"],
        designation=housekeeping_setup["designation"],
        location=housekeeping_setup["location"],
        created_by=housekeeping_setup["admin"],
        name="Evening Person",
        shift=HousekeepingShiftEnum.EVENING,
    )

    sheet = _sheet(
        client.get(
            "/api/v1/housekeeping-staff/export",
            params={"shift": "MORNING"},
            headers=auth_headers(client, housekeeping_setup["admin"]),
        )
    )

    names = [row[2] for row in _data_rows(sheet)]
    assert "Morning Person" in names
    assert "Evening Person" not in names


def test_export_housekeeping_staff_forbidden_for_candidate(client, user_factory):
    candidate = user_factory(UserRoleEnum.CANDIDATE)
    response = client.get("/api/v1/housekeeping-staff/export", headers=auth_headers(client, candidate))
    assert response.status_code == 403


# --- Sanctioned Strength views ----------------------------------------------


@pytest.fixture()
def strength_setup(campus_factory, department_factory, designation_factory, user_factory, sanctioned_strength_factory):
    campus = campus_factory("SSE")
    department = department_factory("SSE", name="Computer Science", code="CSE", category=StaffRoleCategoryEnum.TEACHING)
    designation = designation_factory(
        name="Assistant Professor", category=StaffRoleCategoryEnum.TEACHING, department=department
    )
    admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    sanctioned_strength_factory(
        campus=campus, department=department, designation=designation, approved_strength=7, created_by=admin
    )
    return {"campus": campus, "admin": admin}


def test_export_teaching_strength_view(client, strength_setup):
    sheet = _sheet(
        client.get(
            "/api/v1/sanctioned-strength/views/teaching/export",
            headers=auth_headers(client, strength_setup["admin"]),
        )
    )

    assert sheet.title == "Teaching Strength"
    assert _headers(sheet)[:3] == ["Campus", "Department", "Designation"]
    rows = _data_rows(sheet)
    assert any(row[2] == "Assistant Professor" and row[4] == 7 for row in rows)


def test_export_housekeeping_strength_view_uses_its_own_shape(client, strength_setup):
    """Housekeeping is Location-grained (Required/Available), not the
    Approved/Working shape Teaching and Non-Teaching share."""
    sheet = _sheet(
        client.get(
            "/api/v1/sanctioned-strength/views/housekeeping/export",
            headers=auth_headers(client, strength_setup["admin"]),
        )
    )

    assert sheet.title == "Housekeeping Strength"
    headers = _headers(sheet)
    assert "Required" in headers and "Available" in headers
    assert "Approved" not in headers


def test_export_non_teaching_strength_view_is_a_distinct_sheet(client, strength_setup):
    sheet = _sheet(
        client.get(
            "/api/v1/sanctioned-strength/views/non-teaching/export",
            headers=auth_headers(client, strength_setup["admin"]),
        )
    )
    assert sheet.title == "Non-Teaching Strength"
    # The TEACHING row seeded above must not leak into the NON_TEACHING view.
    assert all(row[2] != "Assistant Professor" for row in _data_rows(sheet))


def test_export_strength_view_rejects_an_unknown_view(client, strength_setup):
    response = client.get(
        "/api/v1/sanctioned-strength/views/nonsense/export",
        headers=auth_headers(client, strength_setup["admin"]),
    )
    assert response.status_code == 422


def test_export_strength_view_forbidden_for_candidate(client, user_factory):
    candidate = user_factory(UserRoleEnum.CANDIDATE)
    response = client.get(
        "/api/v1/sanctioned-strength/views/teaching/export", headers=auth_headers(client, candidate)
    )
    assert response.status_code == 403


# --- Sanctioned Strength "All" tab ------------------------------------------
#
# The last gap in the 2026-08-27 export-parity pass (`c64cf8a`), closed
# 2026-09-01. Unlike the three views above this one is DEPARTMENT-grained --
# a rollup across every category, backed by GET /departments/vacancy-register
# -- so it has its own header set and its own sheet.


def test_export_vacancy_register_all_tab(client, strength_setup):
    sheet = _sheet(
        client.get(
            "/api/v1/departments/vacancy-register/export",
            headers=auth_headers(client, strength_setup["admin"]),
        )
    )

    assert sheet.title == "Sanctioned Strength (All)"
    headers = _headers(sheet)
    assert headers[:4] == ["Campus", "Department", "Category", "Approved"]
    # Department-grained, so no Designation column at all -- that is what
    # distinguishes this sheet from the Teaching/Non-Teaching ones.
    assert "Designation" not in headers
    assert "Recruitment Status" in headers and "Approval Status" in headers

    rows = _data_rows(sheet)
    row = next(r for r in rows if r[1] == "Computer Science")
    assert row[0] == "SSE"
    # Approved rolls the department's sanctioned strength up: one row of 7.
    assert row[3] == 7


def test_export_vacancy_register_renders_categories_as_text_not_enum_reprs(client, strength_setup):
    """`supported_categories` is a list of enum MEMBERS, not strings. Joining
    it naively is the kind of thing that silently ships
    "StaffRoleCategoryEnum.TEACHING" into a customer's spreadsheet."""
    sheet = _sheet(
        client.get(
            "/api/v1/departments/vacancy-register/export",
            headers=auth_headers(client, strength_setup["admin"]),
        )
    )
    row = next(r for r in _data_rows(sheet) if r[1] == "Computer Science")
    assert row[2] == "TEACHING"


def test_export_vacancy_register_honours_the_category_filter(client, strength_setup):
    """The export mirrors whatever the tab is showing, filters included."""
    response = client.get(
        "/api/v1/departments/vacancy-register/export?category=HOUSEKEEPING",
        headers=auth_headers(client, strength_setup["admin"]),
    )
    sheet = _sheet(response)
    assert all(r[1] != "Computer Science" for r in _data_rows(sheet))


def test_export_vacancy_register_rejects_a_bad_sort_field(client, strength_setup):
    # Shares _validate_register_params with the list endpoint, so the two
    # cannot drift into accepting different values for the same param.
    response = client.get(
        "/api/v1/departments/vacancy-register/export?sort_by=nonsense",
        headers=auth_headers(client, strength_setup["admin"]),
    )
    assert response.status_code == 422


def test_export_vacancy_register_is_not_swallowed_by_the_department_id_route(client, strength_setup):
    """`/departments/vacancy-register/export` and
    `/departments/{department_id}/sanctioned-strength-breakdown` are both
    two-segment paths under the same prefix. If the literal is ever registered
    after the parameterised one, "vacancy-register" parses as a department_id
    and this 422s instead of returning a workbook."""
    response = client.get(
        "/api/v1/departments/vacancy-register/export",
        headers=auth_headers(client, strength_setup["admin"]),
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(XLSX_CONTENT_TYPE)


def test_export_vacancy_register_forbidden_for_candidate(client, user_factory):
    candidate = user_factory(UserRoleEnum.CANDIDATE)
    response = client.get(
        "/api/v1/departments/vacancy-register/export", headers=auth_headers(client, candidate)
    )
    assert response.status_code == 403
