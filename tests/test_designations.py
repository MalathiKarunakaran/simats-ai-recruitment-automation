import io
import uuid

from app.models.enums import EmploymentTypeEnum, StaffRoleCategoryEnum, UserRoleEnum, VacancyRequestStatusEnum
from app.models.vacancy_request import VacancyRequest

from tests.conftest import auth_headers


def _payload(**overrides):
    payload = {
        "name": "Assistant Professor",
        "category": "TEACHING",
        "qualification": "PhD in relevant field",
        "min_experience": "3+ years",
        "employment_type": "FULL_TIME",
    }
    payload.update(overrides)
    return payload


def test_super_admin_can_create_and_list_designation(client, user_factory):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)

    response = client.post(
        "/api/v1/designations", headers=auth_headers(client, super_admin), json=_payload()
    )
    assert response.status_code == 201, response.text
    body = response.json()
    assert body["name"] == "Assistant Professor"
    assert body["category"] == "TEACHING"
    assert body["employment_type"] == "FULL_TIME"
    assert body["department_ids"] == []
    assert body["is_active"] is True

    listing = client.get("/api/v1/designations", headers=auth_headers(client, super_admin))
    assert listing.status_code == 200
    assert listing.json()["total"] >= 1


def test_recruitment_coordinator_can_create_designation_with_departments(
    client, user_factory, department_factory
):
    coordinator = user_factory(UserRoleEnum.RECRUITMENT_COORDINATOR)
    department = department_factory("SSE", "Computer Science", category="TEACHING")

    response = client.post(
        "/api/v1/designations",
        headers=auth_headers(client, coordinator),
        json=_payload(department_ids=[str(department.id)]),
    )
    assert response.status_code == 201, response.text
    assert response.json()["department_ids"] == [str(department.id)]


def test_designation_create_rejects_unknown_department_id(client, user_factory):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    import uuid

    response = client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(department_ids=[str(uuid.uuid4())]),
    )
    assert response.status_code == 400


def test_hr_admin_cannot_create_designation(client, user_factory):
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)

    response = client.post(
        "/api/v1/designations", headers=auth_headers(client, hr_admin), json=_payload()
    )
    assert response.status_code == 403


def test_recruitment_officer_cannot_create_designation(client, user_factory):
    officer = user_factory(UserRoleEnum.RECRUITMENT_OFFICER, campus_code="SSE")

    response = client.post(
        "/api/v1/designations", headers=auth_headers(client, officer), json=_payload()
    )
    assert response.status_code == 403


def test_super_admin_can_patch_designation(client, user_factory):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    create_response = client.post(
        "/api/v1/designations", headers=auth_headers(client, super_admin), json=_payload()
    )
    designation_id = create_response.json()["id"]

    patch_response = client.patch(
        f"/api/v1/designations/{designation_id}",
        headers=auth_headers(client, super_admin),
        json={"is_active": False, "employment_type": "ADJUNCT"},
    )
    assert patch_response.status_code == 200, patch_response.text
    body = patch_response.json()
    assert body["is_active"] is False
    assert body["employment_type"] == "ADJUNCT"


def test_hr_admin_cannot_patch_designation(client, user_factory):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    create_response = client.post(
        "/api/v1/designations", headers=auth_headers(client, super_admin), json=_payload()
    )
    designation_id = create_response.json()["id"]

    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)
    patch_response = client.patch(
        f"/api/v1/designations/{designation_id}",
        headers=auth_headers(client, hr_admin),
        json={"is_active": False},
    )
    assert patch_response.status_code == 403


def test_candidate_cannot_list_designations(client, user_factory):
    candidate = user_factory(UserRoleEnum.CANDIDATE)
    response = client.get("/api/v1/designations", headers=auth_headers(client, candidate))
    assert response.status_code == 403


def test_list_designations_filters_by_department_and_category(client, user_factory, department_factory):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    cse_dept = department_factory("SSE", "Computer Science", category="TEACHING")
    other_dept = department_factory("SSE", "Mechanical Engineering", category="NON_TEACHING")

    matching = client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(name="Professor of CSE", department_ids=[str(cse_dept.id)]),
    )
    assert matching.status_code == 201

    non_matching = client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(name="Non-Teaching Clerk", category="NON_TEACHING", department_ids=[str(other_dept.id)]),
    )
    assert non_matching.status_code == 201

    filtered = client.get(
        f"/api/v1/designations?department_id={cse_dept.id}", headers=auth_headers(client, super_admin)
    )
    assert filtered.status_code == 200
    names = {item["name"] for item in filtered.json()["items"]}
    assert "Professor of CSE" in names
    assert "Non-Teaching Clerk" not in names

    category_filtered = client.get(
        "/api/v1/designations?category=NON_TEACHING", headers=auth_headers(client, super_admin)
    )
    names = {item["name"] for item in category_filtered.json()["items"]}
    assert "Non-Teaching Clerk" in names
    assert "Professor of CSE" not in names


def test_category_counts_reflect_full_set_regardless_of_selected_category(client, user_factory):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)

    teaching = client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(name="Counts Teaching Designation", category="TEACHING"),
    )
    assert teaching.status_code == 201
    non_teaching_1 = client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(name="Counts NonTeaching Designation 1", category="NON_TEACHING"),
    )
    assert non_teaching_1.status_code == 201
    non_teaching_2 = client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(name="Counts NonTeaching Designation 2", category="NON_TEACHING"),
    )
    assert non_teaching_2.status_code == 201
    housekeeping = client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(name="Counts Housekeeping Designation", category="HOUSEKEEPING"),
    )
    assert housekeeping.status_code == 201

    unfiltered = client.get("/api/v1/designations", headers=auth_headers(client, super_admin))
    assert unfiltered.status_code == 200
    unfiltered_body = unfiltered.json()
    assert unfiltered_body["category_counts"]["TEACHING"] >= 1
    assert unfiltered_body["category_counts"]["NON_TEACHING"] >= 2
    assert unfiltered_body["category_counts"]["HOUSEKEEPING"] >= 1
    assert unfiltered_body["category_counts"]["ALL"] == unfiltered_body["total"]
    baseline_counts = unfiltered_body["category_counts"]

    # Narrowing items/total via the category filter must NOT change
    # category_counts -- switching tabs shouldn't collapse the other tabs'
    # displayed counts to zero.
    teaching_only = client.get(
        "/api/v1/designations?category=TEACHING", headers=auth_headers(client, super_admin)
    )
    assert teaching_only.status_code == 200
    teaching_body = teaching_only.json()
    assert teaching_body["total"] < unfiltered_body["total"]
    assert all(item["category"] == "TEACHING" for item in teaching_body["items"])
    assert teaching_body["category_counts"] == baseline_counts

    housekeeping_only = client.get(
        "/api/v1/designations?category=HOUSEKEEPING", headers=auth_headers(client, super_admin)
    )
    assert housekeeping_only.status_code == 200
    housekeeping_body = housekeeping_only.json()
    assert housekeeping_body["category_counts"] == baseline_counts


def test_designation_create_rejects_department_category_mismatch(client, user_factory, department_factory):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    housekeeping_dept = department_factory("SSE", "Campus Upkeep", category="HOUSEKEEPING")

    response = client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(department_ids=[str(housekeeping_dept.id)]),
    )
    assert response.status_code == 400
    assert "Campus Upkeep" in response.json()["detail"]
    assert "TEACHING" in response.json()["detail"]


def test_designation_create_with_matching_department_categories_succeeds(client, user_factory, department_factory):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    teaching_dept = department_factory("SSE", "Computer Science", category="TEACHING")

    response = client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(department_ids=[str(teaching_dept.id)]),
    )
    assert response.status_code == 201, response.text
    assert response.json()["department_ids"] == [str(teaching_dept.id)]


def test_designation_update_category_alone_succeeds_when_linked_departments_still_match(
    client, user_factory, department_factory
):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    non_teaching_dept = department_factory("SSE", "Administration", category="NON_TEACHING")

    create_response = client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(category="NON_TEACHING", department_ids=[str(non_teaching_dept.id)]),
    )
    assert create_response.status_code == 201, create_response.text
    designation_id = create_response.json()["id"]

    patch_response = client.patch(
        f"/api/v1/designations/{designation_id}",
        headers=auth_headers(client, super_admin),
        json={"is_active": False},
    )
    assert patch_response.status_code == 200, patch_response.text


def test_designation_update_category_alone_rejected_when_linked_departments_no_longer_match(
    client, user_factory, department_factory
):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    teaching_dept = department_factory("SSE", "Computer Science", category="TEACHING")

    create_response = client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(category="TEACHING", department_ids=[str(teaching_dept.id)]),
    )
    assert create_response.status_code == 201, create_response.text
    designation_id = create_response.json()["id"]

    patch_response = client.patch(
        f"/api/v1/designations/{designation_id}",
        headers=auth_headers(client, super_admin),
        json={"category": "NON_TEACHING"},
    )
    assert patch_response.status_code == 400
    detail = patch_response.json()["detail"]
    assert "NON_TEACHING" in detail
    assert "department_ids" in detail


def test_designation_update_department_ids_rejects_category_mismatch(
    client, user_factory, department_factory
):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    housekeeping_dept = department_factory("SSE", "Facilities", category="HOUSEKEEPING")

    create_response = client.post(
        "/api/v1/designations", headers=auth_headers(client, super_admin), json=_payload()
    )
    assert create_response.status_code == 201, create_response.text
    designation_id = create_response.json()["id"]

    patch_response = client.patch(
        f"/api/v1/designations/{designation_id}",
        headers=auth_headers(client, super_admin),
        json={"department_ids": [str(housekeeping_dept.id)]},
    )
    assert patch_response.status_code == 400
    assert "Facilities" in patch_response.json()["detail"]


def test_designation_update_department_ids_to_valid_set_succeeds(client, user_factory, department_factory):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    teaching_dept = department_factory("SSE", "Computer Science", category="TEACHING")

    create_response = client.post(
        "/api/v1/designations", headers=auth_headers(client, super_admin), json=_payload()
    )
    assert create_response.status_code == 201, create_response.text
    designation_id = create_response.json()["id"]

    patch_response = client.patch(
        f"/api/v1/designations/{designation_id}",
        headers=auth_headers(client, super_admin),
        json={"department_ids": [str(teaching_dept.id)]},
    )
    assert patch_response.status_code == 200, patch_response.text
    assert patch_response.json()["department_ids"] == [str(teaching_dept.id)]


def test_designation_update_department_ids_and_category_together_validated_against_new_category(
    client, user_factory, department_factory
):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    teaching_dept = department_factory("SSE", "Computer Science", category="TEACHING")
    non_teaching_dept = department_factory("SSE", "Administration", category="NON_TEACHING")

    create_response = client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(department_ids=[str(teaching_dept.id)]),
    )
    assert create_response.status_code == 201, create_response.text
    designation_id = create_response.json()["id"]

    # Changing category and department_ids together to a still-mismatched
    # pair (new category NON_TEACHING, but department_ids still points at a
    # TEACHING department) must be validated against the *new* category.
    patch_response = client.patch(
        f"/api/v1/designations/{designation_id}",
        headers=auth_headers(client, super_admin),
        json={"category": "NON_TEACHING", "department_ids": [str(teaching_dept.id)]},
    )
    assert patch_response.status_code == 400

    # Changing both together to a matching pair succeeds.
    patch_response = client.patch(
        f"/api/v1/designations/{designation_id}",
        headers=auth_headers(client, super_admin),
        json={"category": "NON_TEACHING", "department_ids": [str(non_teaching_dept.id)]},
    )
    assert patch_response.status_code == 200, patch_response.text
    assert patch_response.json()["category"] == "NON_TEACHING"
    assert patch_response.json()["department_ids"] == [str(non_teaching_dept.id)]


def test_category_counts_respect_is_active_filter_but_not_category(client, user_factory):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)

    active_teaching = client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(name="Active Filter Teaching"),
    )
    assert active_teaching.status_code == 201

    inactive_non_teaching = client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(name="Inactive Filter NonTeaching", category="NON_TEACHING", is_active=False),
    )
    assert inactive_non_teaching.status_code == 201

    active_only = client.get(
        "/api/v1/designations?is_active=true", headers=auth_headers(client, super_admin)
    )
    assert active_only.status_code == 200
    active_body = active_only.json()
    names = {item["name"] for item in active_body["items"]}
    assert "Active Filter Teaching" in names
    assert "Inactive Filter NonTeaching" not in names
    # The is_active filter narrows counts too (it's applied before the
    # per-category GROUP BY, same as category/department_id/is_active are
    # all "every other active filter" this endpoint respects) -- so the
    # inactive designation must not be reflected here.
    assert active_body["category_counts"]["ALL"] == active_body["total"]


def test_hr_admin_cannot_delete_designation(client, user_factory):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    create_response = client.post(
        "/api/v1/designations", headers=auth_headers(client, super_admin), json=_payload()
    )
    designation_id = create_response.json()["id"]

    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)
    response = client.delete(
        f"/api/v1/designations/{designation_id}", headers=auth_headers(client, hr_admin)
    )
    assert response.status_code == 403


def test_super_admin_can_delete_designation_with_no_dependents(client, user_factory):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    create_response = client.post(
        "/api/v1/designations", headers=auth_headers(client, super_admin), json=_payload()
    )
    designation_id = create_response.json()["id"]

    response = client.delete(
        f"/api/v1/designations/{designation_id}", headers=auth_headers(client, super_admin)
    )
    assert response.status_code == 204

    get_response = client.get(
        f"/api/v1/designations?category=TEACHING", headers=auth_headers(client, super_admin)
    )
    match = next(item for item in get_response.json()["items"] if item["id"] == designation_id)
    assert match["is_active"] is False


def test_delete_designation_blocked_by_in_flight_vacancy_request(
    client, user_factory, department_factory, db_session
):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    department = department_factory("SSE", "Computer Science", category=StaffRoleCategoryEnum.TEACHING)

    create_response = client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(department_ids=[str(department.id)]),
    )
    designation_id = create_response.json()["id"]

    vacancy_request = VacancyRequest(
        campus_id=department.campus_id,
        department_id=department.id,
        designation_id=uuid.UUID(designation_id),
        role_category=StaffRoleCategoryEnum.TEACHING,
        position_title="Assistant Professor",
        employment_type=EmploymentTypeEnum.FULL_TIME,
        requested_count=1,
        qualification="PhD",
        experience_required="3+ years",
        status=VacancyRequestStatusEnum.SUBMITTED,
        requested_by_id=super_admin.id,
    )
    db_session.add(vacancy_request)
    db_session.flush()

    response = client.delete(
        f"/api/v1/designations/{designation_id}", headers=auth_headers(client, super_admin)
    )
    assert response.status_code == 409
    assert "vacancy request" in response.json()["detail"]


def test_delete_designation_not_blocked_by_terminal_vacancy_request(
    client, user_factory, department_factory, db_session
):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    department = department_factory("SSE", "Computer Science", category=StaffRoleCategoryEnum.TEACHING)

    create_response = client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(department_ids=[str(department.id)]),
    )
    designation_id = create_response.json()["id"]

    vacancy_request = VacancyRequest(
        campus_id=department.campus_id,
        department_id=department.id,
        designation_id=uuid.UUID(designation_id),
        role_category=StaffRoleCategoryEnum.TEACHING,
        position_title="Assistant Professor",
        employment_type=EmploymentTypeEnum.FULL_TIME,
        requested_count=1,
        qualification="PhD",
        experience_required="3+ years",
        status=VacancyRequestStatusEnum.CLOSED,
        requested_by_id=super_admin.id,
    )
    db_session.add(vacancy_request)
    db_session.flush()

    response = client.delete(
        f"/api/v1/designations/{designation_id}", headers=auth_headers(client, super_admin)
    )
    assert response.status_code == 204


def test_delete_designation_blocked_by_active_sanctioned_strength(
    client, user_factory, department_factory, campus_factory, sanctioned_strength_factory, db_session
):
    from app.models.designation import Designation

    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    department = department_factory("SSE", "Computer Science", category=StaffRoleCategoryEnum.TEACHING)
    campus = campus_factory("SSE")

    create_response = client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(department_ids=[str(department.id)]),
    )
    designation_id = create_response.json()["id"]
    designation = db_session.get(Designation, uuid.UUID(designation_id))

    sanctioned_strength_factory(
        campus=campus, department=department, designation=designation, created_by=super_admin
    )

    response = client.delete(
        f"/api/v1/designations/{designation_id}", headers=auth_headers(client, super_admin)
    )
    assert response.status_code == 409
    assert "sanctioned strength" in response.json()["detail"]


def test_delete_unknown_designation_returns_404(client, user_factory):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    response = client.delete(
        f"/api/v1/designations/{uuid.uuid4()}", headers=auth_headers(client, super_admin)
    )
    assert response.status_code == 404


# --- required_skills (Designation Master bulk-upload epic, backend Phase 1) -


def test_create_and_read_designation_with_required_skills(client, user_factory):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)

    response = client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(required_skills="Curriculum design, Research methodology"),
    )
    assert response.status_code == 201, response.text
    assert response.json()["required_skills"] == "Curriculum design, Research methodology"


def test_required_skills_defaults_to_null_when_omitted(client, user_factory):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)

    response = client.post(
        "/api/v1/designations", headers=auth_headers(client, super_admin), json=_payload()
    )
    assert response.status_code == 201, response.text
    assert response.json()["required_skills"] is None


def test_patch_designation_required_skills(client, user_factory):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    create_response = client.post(
        "/api/v1/designations", headers=auth_headers(client, super_admin), json=_payload()
    )
    designation_id = create_response.json()["id"]

    patch_response = client.patch(
        f"/api/v1/designations/{designation_id}",
        headers=auth_headers(client, super_admin),
        json={"required_skills": "Updated skill set"},
    )
    assert patch_response.status_code == 200, patch_response.text
    assert patch_response.json()["required_skills"] == "Updated skill set"


# --- search (previously client-side only) -----------------------------------


def test_list_designations_search_filters_by_name(client, user_factory):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)

    client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(name="Search Target Professor"),
    )
    client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(name="Unrelated Designation"),
    )

    response = client.get(
        "/api/v1/designations?search=Search Target", headers=auth_headers(client, super_admin)
    )
    assert response.status_code == 200
    names = {item["name"] for item in response.json()["items"]}
    assert "Search Target Professor" in names
    assert "Unrelated Designation" not in names


def test_list_designations_search_is_case_insensitive(client, user_factory):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(name="Case Insensitive Search Target"),
    )

    response = client.get(
        "/api/v1/designations?search=case insensitive", headers=auth_headers(client, super_admin)
    )
    assert response.status_code == 200
    names = {item["name"] for item in response.json()["items"]}
    assert "Case Insensitive Search Target" in names


# --- export -------------------------------------------------------------


def test_export_designations_returns_xlsx(client, user_factory, department_factory):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    department = department_factory("SSE", "Computer Science", category="TEACHING")

    create_response = client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(name="Export Target Professor", department_ids=[str(department.id)]),
    )
    assert create_response.status_code == 201, create_response.text

    response = client.get("/api/v1/designations/export", headers=auth_headers(client, super_admin))
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=5, values_only=True))
    names = {row[0] for row in rows if row and row[0]}
    assert "Export Target Professor" in names


def test_export_designations_respects_category_filter(client, user_factory):
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(name="Export Filter Teaching", category="TEACHING"),
    )
    client.post(
        "/api/v1/designations",
        headers=auth_headers(client, super_admin),
        json=_payload(name="Export Filter NonTeaching", category="NON_TEACHING"),
    )

    response = client.get(
        "/api/v1/designations/export?category=NON_TEACHING", headers=auth_headers(client, super_admin)
    )
    assert response.status_code == 200

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=5, values_only=True))
    names = {row[0] for row in rows if row and row[0]}
    assert "Export Filter NonTeaching" in names
    assert "Export Filter Teaching" not in names


def test_export_designations_forbidden_for_candidate(client, user_factory):
    candidate = user_factory(UserRoleEnum.CANDIDATE)
    response = client.get("/api/v1/designations/export", headers=auth_headers(client, candidate))
    assert response.status_code == 403
