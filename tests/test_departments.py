import pytest

from app.models.enums import StaffRoleCategoryEnum, UserRoleEnum

from tests.conftest import auth_headers


def test_hod_cannot_create_department_anymore(client, user_factory, campus_factory):
    # Campus HOD lost department-write access with the Department/Designation
    # Master rollout (confirmed decision) -- only SUPER_ADMIN/HR_ADMIN can
    # write departments now.
    sse = campus_factory("SSE")
    hod = user_factory(UserRoleEnum.CAMPUS_HOD, campus_code="SSE")

    response = client.post(
        "/api/v1/departments",
        headers=auth_headers(client, hod),
        json={"campus_id": str(sse.id), "name": "Computer Science"},
    )
    assert response.status_code == 403


def test_hr_admin_can_create_department_with_master_fields(client, user_factory, campus_factory):
    sse = campus_factory("SSE")
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)

    response = client.post(
        "/api/v1/departments",
        headers=auth_headers(client, hr_admin),
        json={
            "campus_id": str(sse.id),
            "name": "Computer Science",
            "code": "CSE",
            "category": "TEACHING",
            "parent_group": "Engineering",
        },
    )
    assert response.status_code == 201, response.text
    body = response.json()
    assert body["campus_id"] == str(sse.id)
    assert body["code"] == "CSE"
    assert body["category"] == "TEACHING"
    assert body["parent_group"] == "Engineering"


def test_super_admin_can_patch_department_master_fields(client, user_factory, campus_factory):
    sse = campus_factory("SSE")
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)

    create_response = client.post(
        "/api/v1/departments",
        headers=auth_headers(client, super_admin),
        json={"campus_id": str(sse.id), "name": "Mechanical Engineering"},
    )
    assert create_response.status_code == 201
    department_id = create_response.json()["id"]

    patch_response = client.patch(
        f"/api/v1/departments/{department_id}",
        headers=auth_headers(client, super_admin),
        json={"code": "MECH", "category": "NON_TEACHING", "parent_group": "Engineering"},
    )
    assert patch_response.status_code == 200, patch_response.text
    body = patch_response.json()
    assert body["code"] == "MECH"
    assert body["category"] == "NON_TEACHING"
    assert body["parent_group"] == "Engineering"


def test_hod_cannot_patch_department_anymore(client, user_factory, campus_factory, department_factory):
    department = department_factory("SSE")
    hod = user_factory(UserRoleEnum.CAMPUS_HOD, campus_code="SSE")

    response = client.patch(
        f"/api/v1/departments/{department.id}",
        headers=auth_headers(client, hod),
        json={"code": "X"},
    )
    assert response.status_code == 403


def test_department_list_is_campus_scoped_for_hod(client, user_factory, campus_factory, db_session):
    from app.models.department import Department

    sse = campus_factory("SSE")
    scad = campus_factory("SCAD")
    db_session.add_all(
        [
            Department(campus_id=sse.id, name="SSE Dept"),
            Department(campus_id=scad.id, name="SCAD Dept"),
        ]
    )
    db_session.flush()

    hod_sse = user_factory(UserRoleEnum.CAMPUS_HOD, campus_code="SSE")
    response = client.get("/api/v1/departments", headers=auth_headers(client, hod_sse))
    names = {d["name"] for d in response.json()["items"]}
    assert "SSE Dept" in names
    assert "SCAD Dept" not in names


def test_candidate_cannot_list_departments(client, user_factory):
    candidate = user_factory(UserRoleEnum.CANDIDATE)
    response = client.get("/api/v1/departments", headers=auth_headers(client, candidate))
    assert response.status_code == 403


def test_create_department_without_category_defaults_to_non_teaching(client, user_factory, campus_factory):
    # Department.category became NOT NULL in the Phase 1 staff-category
    # migrations (see alembic/versions/a1b2c3d4e5f7_..._backfill_department_category.py)
    # -- the API still accepts an omitted category (mirroring that
    # migration's own "ambiguous -> NON_TEACHING" default) rather than 422ing,
    # since the RBAC test above relies on a category-less payload still
    # reaching the role check.
    sse = campus_factory("SSE")
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)

    response = client.post(
        "/api/v1/departments",
        headers=auth_headers(client, hr_admin),
        json={"campus_id": str(sse.id), "name": "New Dept Without Category"},
    )
    assert response.status_code == 201, response.text
    assert response.json()["category"] == StaffRoleCategoryEnum.NON_TEACHING.value


def test_department_factory_departments_have_a_category(department_factory):
    # Model-level Python default (StaffRoleCategoryEnum.NON_TEACHING) keeps
    # every existing Department(...) construction site working now that the
    # column is NOT NULL, without needing to touch every call site.
    department = department_factory("SSE")
    assert department.category == StaffRoleCategoryEnum.NON_TEACHING


def test_hod_cannot_delete_department(client, user_factory, department_factory):
    department = department_factory("SSE")
    hod = user_factory(UserRoleEnum.CAMPUS_HOD, campus_code="SSE")

    response = client.delete(f"/api/v1/departments/{department.id}", headers=auth_headers(client, hod))
    assert response.status_code == 403


def test_super_admin_can_delete_department_with_no_dependents(client, user_factory, department_factory):
    department = department_factory("SSE")
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)

    response = client.delete(f"/api/v1/departments/{department.id}", headers=auth_headers(client, super_admin))
    assert response.status_code == 204

    get_response = client.get("/api/v1/departments", headers=auth_headers(client, super_admin))
    body = get_response.json()
    match = next(item for item in body["items"] if item["id"] == str(department.id))
    assert match["is_active"] is False


def test_delete_department_blocked_by_active_user(client, user_factory, department_factory, db_session):
    department = department_factory("SSE")
    staff = user_factory(UserRoleEnum.RECRUITMENT_OFFICER, campus_code="SSE")
    staff.department_id = department.id
    db_session.flush()
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)

    response = client.delete(f"/api/v1/departments/{department.id}", headers=auth_headers(client, hr_admin))
    assert response.status_code == 409
    assert "user" in response.json()["detail"]


def test_delete_department_blocked_by_active_designation(client, user_factory, department_factory, designation_factory):
    department = department_factory("SSE", "Computer Science", category=StaffRoleCategoryEnum.TEACHING)
    designation_factory(category=StaffRoleCategoryEnum.TEACHING, department=department)
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)

    response = client.delete(f"/api/v1/departments/{department.id}", headers=auth_headers(client, hr_admin))
    assert response.status_code == 409
    assert "designation" in response.json()["detail"]


def test_delete_department_not_blocked_by_inactive_designation(
    client, user_factory, department_factory, designation_factory, db_session
):
    department = department_factory("SSE", "Computer Science", category=StaffRoleCategoryEnum.TEACHING)
    designation = designation_factory(category=StaffRoleCategoryEnum.TEACHING, department=department)
    designation.is_active = False
    db_session.flush()
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)

    response = client.delete(f"/api/v1/departments/{department.id}", headers=auth_headers(client, hr_admin))
    assert response.status_code == 204


def test_delete_unknown_department_returns_404(client, user_factory):
    import uuid

    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)
    response = client.delete(f"/api/v1/departments/{uuid.uuid4()}", headers=auth_headers(client, hr_admin))
    assert response.status_code == 404


# --- Department Master hardening epic (2026-08-25): search/sort/filter -----


def test_list_departments_search_matches_name(client, user_factory, department_factory):
    department_factory("SSE", name="Computer Science", code="CSE")
    department_factory("SSE", name="Mechanical Engineering", code="MECH")
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)

    response = client.get(
        "/api/v1/departments", params={"search": "computer"}, headers=auth_headers(client, hr_admin)
    )
    names = {d["name"] for d in response.json()["items"]}
    assert names == {"Computer Science"}


def test_list_departments_search_matches_code(client, user_factory, department_factory):
    department_factory("SSE", name="Computer Science", code="CSE")
    department_factory("SSE", name="Mechanical Engineering", code="MECH")
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)

    response = client.get("/api/v1/departments", params={"search": "MECH"}, headers=auth_headers(client, hr_admin))
    names = {d["name"] for d in response.json()["items"]}
    assert names == {"Mechanical Engineering"}


def test_list_departments_category_counts_respect_non_category_filters(
    client, user_factory, department_factory
):
    department_factory("SSE", name="Teach A", category=StaffRoleCategoryEnum.TEACHING)
    department_factory("SSE", name="Teach B", category=StaffRoleCategoryEnum.TEACHING, is_active=False)
    department_factory("SSE", name="NonTeach A", category=StaffRoleCategoryEnum.NON_TEACHING)
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)

    response = client.get(
        "/api/v1/departments", params={"is_active": True}, headers=auth_headers(client, hr_admin)
    )
    body = response.json()
    assert body["category_counts"]["TEACHING"] == 1  # Teach B excluded by is_active=True
    assert body["category_counts"]["NON_TEACHING"] == 1


def test_list_departments_category_counts_do_not_change_when_category_filter_applied(
    client, user_factory, department_factory
):
    department_factory("SSE", name="Teach A", category=StaffRoleCategoryEnum.TEACHING)
    department_factory("SSE", name="NonTeach A", category=StaffRoleCategoryEnum.NON_TEACHING)
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)

    all_response = client.get("/api/v1/departments", headers=auth_headers(client, hr_admin))
    filtered_response = client.get(
        "/api/v1/departments", params={"category": "TEACHING"}, headers=auth_headers(client, hr_admin)
    )
    assert all_response.json()["category_counts"] == filtered_response.json()["category_counts"]
    assert len(filtered_response.json()["items"]) == 1


@pytest.mark.parametrize("sort_by", ["name", "code", "category", "campus", "parent_group", "is_active"])
def test_list_departments_sort_by_each_field_does_not_error(client, user_factory, department_factory, sort_by):
    department_factory("SSE", name="B Dept", code="B", category=StaffRoleCategoryEnum.TEACHING, parent_group="Z")
    department_factory("SCAD", name="A Dept", code="A", category=StaffRoleCategoryEnum.NON_TEACHING, parent_group="A")
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)

    response = client.get(
        "/api/v1/departments", params={"sort_by": sort_by, "sort_dir": "desc"}, headers=auth_headers(client, hr_admin)
    )
    assert response.status_code == 200, response.text
    assert len(response.json()["items"]) == 2


def test_list_departments_campus_id_filter_applies_for_global_scope_role(
    client, user_factory, department_factory
):
    department_factory("SSE", name="SSE Dept")
    department_factory("SCAD", name="SCAD Dept")
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)  # global scope role

    sse_id = client.get(
        "/api/v1/campuses", headers=auth_headers(client, hr_admin)
    ).json()
    sse_campus_id = next(c["id"] for c in sse_id["items"] if c["code"] == "SSE")

    response = client.get(
        "/api/v1/departments", params={"campus_id": sse_campus_id}, headers=auth_headers(client, hr_admin)
    )
    names = {d["name"] for d in response.json()["items"]}
    assert names == {"SSE Dept"}


def test_list_departments_campus_id_filter_ignored_for_non_global_scope_role(
    client, user_factory, campus_factory, department_factory
):
    sse_dept = department_factory("SSE", name="SSE Dept")
    scad = campus_factory("SCAD")
    department_factory("SCAD", name="SCAD Dept")
    hod_sse = user_factory(UserRoleEnum.CAMPUS_HOD, campus_code="SSE")

    # A CAMPUS_HOD passing another campus's id must still only see their own
    # campus -- the query param is silently ignored for non-global roles.
    response = client.get(
        "/api/v1/departments", params={"campus_id": str(scad.id)}, headers=auth_headers(client, hod_sse)
    )
    names = {d["name"] for d in response.json()["items"]}
    assert names == {"SSE Dept"}
    assert sse_dept.name in names


# --- Code+Campus uniqueness (application-level, see the migration's own
# docstring for why there is no DB constraint yet) -------------------------


def test_create_department_rejects_duplicate_code_same_campus(client, user_factory, campus_factory):
    sse = campus_factory("SSE")
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)
    first = client.post(
        "/api/v1/departments",
        headers=auth_headers(client, hr_admin),
        json={"campus_id": str(sse.id), "name": "Computer Science", "code": "CSE"},
    )
    assert first.status_code == 201

    second = client.post(
        "/api/v1/departments",
        headers=auth_headers(client, hr_admin),
        json={"campus_id": str(sse.id), "name": "Computer Science Engineering", "code": "cse"},
    )
    assert second.status_code == 400
    assert second.json()["detail"] == 'Department Code "cse" already exists for campus SSE.'


def test_create_department_allows_same_code_on_a_different_campus(client, user_factory, campus_factory):
    sse = campus_factory("SSE")
    scad = campus_factory("SCAD")
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)
    first = client.post(
        "/api/v1/departments",
        headers=auth_headers(client, hr_admin),
        json={"campus_id": str(sse.id), "name": "Computer Science", "code": "CSE"},
    )
    assert first.status_code == 201

    second = client.post(
        "/api/v1/departments",
        headers=auth_headers(client, hr_admin),
        json={"campus_id": str(scad.id), "name": "Computer Science", "code": "CSE"},
    )
    assert second.status_code == 201


def test_create_department_allows_multiple_null_codes_to_coexist(client, user_factory, campus_factory):
    sse = campus_factory("SSE")
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)
    first = client.post(
        "/api/v1/departments",
        headers=auth_headers(client, hr_admin),
        json={"campus_id": str(sse.id), "name": "Dept One"},
    )
    assert first.status_code == 201

    second = client.post(
        "/api/v1/departments",
        headers=auth_headers(client, hr_admin),
        json={"campus_id": str(sse.id), "name": "Dept Two"},
    )
    assert second.status_code == 201


def test_update_department_rejects_duplicate_code_same_campus(client, user_factory, department_factory):
    department_factory("SSE", name="Computer Science", code="CSE")
    other = department_factory("SSE", name="Mechanical Engineering", code="MECH")
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)

    response = client.patch(
        f"/api/v1/departments/{other.id}", headers=auth_headers(client, hr_admin), json={"code": "CSE"}
    )
    assert response.status_code == 400
    assert response.json()["detail"] == 'Department Code "CSE" already exists for campus SSE.'


def test_update_department_allows_re_saving_its_own_unchanged_code(client, user_factory, department_factory):
    department = department_factory("SSE", name="Computer Science", code="CSE")
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)

    response = client.patch(
        f"/api/v1/departments/{department.id}",
        headers=auth_headers(client, hr_admin),
        json={"code": "CSE", "parent_group": "Engineering"},
    )
    assert response.status_code == 200, response.text


def test_update_department_description_field(client, user_factory, department_factory):
    department = department_factory("SSE", name="Computer Science")
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)

    response = client.patch(
        f"/api/v1/departments/{department.id}",
        headers=auth_headers(client, hr_admin),
        json={"description": "Handles all CS programs."},
    )
    assert response.status_code == 200
    assert response.json()["description"] == "Handles all CS programs."


# --- Export ------------------------------------------------------------


def test_export_departments_returns_xlsx_respecting_filters(client, user_factory, department_factory):
    department_factory("SSE", name="Computer Science", code="CSE", category=StaffRoleCategoryEnum.TEACHING)
    department_factory("SSE", name="Housekeeping Staff", code="HK", category=StaffRoleCategoryEnum.HOUSEKEEPING)
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)

    response = client.get(
        "/api/v1/departments/export",
        params={"category": "TEACHING"},
        headers=auth_headers(client, hr_admin),
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    from openpyxl import load_workbook
    import io

    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=5, values_only=True))
    non_empty = [r for r in rows if any(c is not None for c in r)]
    assert len(non_empty) == 1
    assert non_empty[0][1] == "CSE"


def test_export_departments_forbidden_for_candidate(client, user_factory):
    candidate = user_factory(UserRoleEnum.CANDIDATE)
    response = client.get("/api/v1/departments/export", headers=auth_headers(client, candidate))
    assert response.status_code == 403
