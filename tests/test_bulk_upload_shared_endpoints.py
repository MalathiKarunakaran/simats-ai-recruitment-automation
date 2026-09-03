"""Entity-agnostic behavior of the 4 shared bulk-upload endpoints
(glowing-zooming-hamming.md Phase J, extended 2026-08-25 to DEPARTMENT,
extended again for the starter regulatory-eligibility-rules feature, backend
Phase 1, to ELIGIBILITY_RULE, and extended a 5th time for the Designation
Master bulk-upload epic, backend Phase 1, to DESIGNATION) -- list/
error-report/original-file/undo in app/api/v1/routers/sanctioned_strength.py,
now dispatched by `BulkUploadLog.entity_type` to cover LOCATION,
HOUSEKEEPING_STAFF, DEPARTMENT, ELIGIBILITY_RULE, and DESIGNATION as well as
the pre-existing SANCTIONED_STRENGTH. Per-entity validate/commit/template
behavior is covered in tests/test_location_bulk_upload.py,
tests/test_housekeeping_staff_bulk_upload.py, tests/test_department_bulk_upload.py,
tests/test_eligibility_rule_bulk_upload.py, and
tests/test_designation_bulk_upload.py -- not duplicated here.
"""

import csv
import io
import uuid
from datetime import datetime, timedelta, timezone

import pytest

from app.models.bulk_upload_log import BulkUploadLog
from app.models.bulk_upload_row_log import BulkUploadRowLog
from app.models.department import Department
from app.models.designation import Designation
from app.models.eligibility_rule import EligibilityRule
from app.models.enums import StaffRoleCategoryEnum, UserRoleEnum
from app.models.housekeeping_staff import HousekeepingStaff
from app.models.location import Location

from tests.conftest import auth_headers

SHARED_ENDPOINT = "/api/v1/sanctioned-strength"
LOCATION_ENDPOINT = "/api/v1/locations"
HOUSEKEEPING_ENDPOINT = "/api/v1/housekeeping-staff"
DEPARTMENT_ENDPOINT = "/api/v1/departments"
ELIGIBILITY_RULE_ENDPOINT = "/api/v1/eligibility-rules"
DESIGNATION_ENDPOINT = "/api/v1/designations"

LOCATION_HEADERS = ["Campus Code", "Location Name", "Block/Building", "Floor/Venue", "Category"]
DEPARTMENT_HEADERS = [
    "Campus Code",
    "Department Code",
    "Department Name",
    "Supported Staff Categories",
    "Parent Group",
    "Description",
    "Active",
]
HOUSEKEEPING_HEADERS = [
    "Campus Code",
    "Bio ID",
    "Name",
    "Designation Name",
    "Location Name",
    "Block",
    "Floor/Venue",
    "Shift",
    "Supervisor",
]
DESIGNATION_HEADERS = [
    "Designation Name",
    "Category",
    "Department Codes",
    "Minimum Qualification",
    "Minimum Experience",
    "Employment Type",
    "Required Skills",
    "Active",
]
ELIGIBILITY_RULE_HEADERS = [
    "Campus Code", "Department Code", "Staff Category", "Position Title",
    "Required Qualification Keyword", "NET/SET/SLET Required", "Subject", "Skills Keyword",
    "ID Proof Required", "Shift Preference", "Regulatory Authority", "School/College",
    "Programme/Discipline", "Minimum Qualification", "Minimum Percentage", "Required Experience",
    "Required Credential", "Required Keywords (informational only)", "Preferred Keywords (informational only)",
    "PhD Required", "Professional Registration", "Industry Experience", "Priority",
    "Effective From (YYYY-MM-DD)", "Effective To (YYYY-MM-DD)", "Source Regulation", "Status",
    "Verification Required", "Active", "Notes",
]


def _csv_bytes(headers: list[str], rows: list[list]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return buf.getvalue().encode("utf-8")


def _commit_location(client, actor, rows):
    return client.post(
        f"{LOCATION_ENDPOINT}/bulk-upload/commit",
        files={"file": ("upload.csv", _csv_bytes(LOCATION_HEADERS, rows), "text/csv")},
        headers=auth_headers(client, actor),
    )


def _commit_housekeeping(client, actor, rows):
    return client.post(
        f"{HOUSEKEEPING_ENDPOINT}/bulk-upload/commit",
        files={"file": ("upload.csv", _csv_bytes(HOUSEKEEPING_HEADERS, rows), "text/csv")},
        headers=auth_headers(client, actor),
    )


def _commit_department(client, actor, rows):
    return client.post(
        f"{DEPARTMENT_ENDPOINT}/bulk-upload/commit",
        files={"file": ("upload.csv", _csv_bytes(DEPARTMENT_HEADERS, rows), "text/csv")},
        headers=auth_headers(client, actor),
    )


def _commit_eligibility_rule(client, actor, rows):
    return client.post(
        f"{ELIGIBILITY_RULE_ENDPOINT}/bulk-upload/commit",
        files={"file": ("upload.csv", _csv_bytes(ELIGIBILITY_RULE_HEADERS, rows), "text/csv")},
        headers=auth_headers(client, actor),
    )


def _commit_designation(client, actor, rows):
    return client.post(
        f"{DESIGNATION_ENDPOINT}/bulk-upload/commit",
        files={"file": ("upload.csv", _csv_bytes(DESIGNATION_HEADERS, rows), "text/csv")},
        headers=auth_headers(client, actor),
    )


@pytest.fixture()
def location_setup(campus_factory, user_factory):
    campus = campus_factory("SSE")
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)
    recruitment_officer = user_factory(UserRoleEnum.RECRUITMENT_OFFICER, campus_code="SSE")
    return {"campus": campus, "hr_admin": hr_admin, "recruitment_officer": recruitment_officer}


def _location_row(setup, name="Shared Block", block="Block A", floor="Ground", category="TEACHING"):
    return [setup["campus"].code, name, block, floor, category]


@pytest.fixture()
def department_setup(campus_factory, user_factory):
    campus = campus_factory("SSE")
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)
    return {"campus": campus, "hr_admin": hr_admin}


def _department_row(setup, code="SHDEPT", name="Shared Department", category="TEACHING"):
    return [setup["campus"].code, code, name, category, "Engineering", "", "TRUE"]


@pytest.fixture()
def designation_setup(campus_factory, department_factory, user_factory):
    campus = campus_factory("SSE")
    department = department_factory(
        "SSE", name="Shared Designation Dept", code="SHDESIG", category=StaffRoleCategoryEnum.TEACHING
    )
    # DESIGNATION_WRITE_ROLES = {SUPER_ADMIN, RECRUITMENT_COORDINATOR} --
    # SUPER_ADMIN is used as the write actor here (same as department_setup
    # uses HR_ADMIN, its own write role).
    super_admin = user_factory(UserRoleEnum.SUPER_ADMIN)
    return {"campus": campus, "department": department, "super_admin": super_admin}


def _designation_row(setup, name="Shared Designation", category="TEACHING"):
    return [name, category, setup["department"].code, "PhD in relevant field", "3+ years", "FULL_TIME", "", "TRUE"]


@pytest.fixture()
def eligibility_rule_setup(campus_factory, department_factory, user_factory):
    campus = campus_factory("SSE")
    department = department_factory("SSE", name="Shared Eligibility Dept", code="SHELIG")
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)
    return {"campus": campus, "department": department, "hr_admin": hr_admin}


def _eligibility_rule_row(setup, position_title="Shared Position", regulatory_authority="AICTE_UGC"):
    return [
        setup["campus"].code, setup["department"].code, "TEACHING", position_title,
        "PHD", "TRUE", "Physics", "", "", "", regulatory_authority, "", "", "", "", "", "", "", "",
        "TRUE", "", "", "", "2026-01-01", "", "", "DRAFT", "TRUE", "TRUE", "",
    ]


@pytest.fixture()
def housekeeping_setup(campus_factory, designation_factory, location_factory, user_factory):
    campus = campus_factory("SSE")
    designation = designation_factory(
        StaffRoleCategoryEnum.HOUSEKEEPING, name=f"Shared HK Designation {uuid.uuid4().hex[:6]}"
    )
    location = location_factory("SSE", name=f"Shared HK Location {uuid.uuid4().hex[:6]}")
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)
    recruitment_officer = user_factory(UserRoleEnum.RECRUITMENT_OFFICER, campus_code="SSE")
    return {
        "campus": campus,
        "designation": designation,
        "location": location,
        "hr_admin": hr_admin,
        "recruitment_officer": recruitment_officer,
    }


def _housekeeping_row(setup, bio_id="BIO-2001", name="Jane Roe", shift="MORNING"):
    return [
        setup["campus"].code,
        bio_id,
        name,
        setup["designation"].name,
        setup["location"].name,
        "",
        "",
        shift,
        "",
    ]


# --- list: entity_type filter -----------------------------------------------


def test_list_bulk_uploads_entity_type_filter(client, location_setup, housekeeping_setup):
    _commit_location(client, location_setup["hr_admin"], [_location_row(location_setup)])
    _commit_housekeeping(client, housekeeping_setup["hr_admin"], [_housekeeping_row(housekeeping_setup)])

    response = client.get(
        f"{SHARED_ENDPOINT}/bulk-uploads",
        params={"entity_type": "LOCATION"},
        headers=auth_headers(client, location_setup["hr_admin"]),
    )
    assert response.status_code == 200
    body = response.json()
    assert body["total"] >= 1
    assert all(item["entity_type"] == "LOCATION" for item in body["items"])

    response = client.get(
        f"{SHARED_ENDPOINT}/bulk-uploads",
        params={"entity_type": "HOUSEKEEPING_STAFF"},
        headers=auth_headers(client, housekeeping_setup["hr_admin"]),
    )
    body = response.json()
    assert all(item["entity_type"] == "HOUSEKEEPING_STAFF" for item in body["items"])


def test_list_bulk_uploads_allowed_for_recruitment_officer(client, location_setup):
    _commit_location(client, location_setup["hr_admin"], [_location_row(location_setup)])
    response = client.get(
        f"{SHARED_ENDPOINT}/bulk-uploads", headers=auth_headers(client, location_setup["recruitment_officer"])
    )
    assert response.status_code == 200


# --- error report: dispatches by entity_type --------------------------------


def test_error_report_for_location_batch(client, location_setup):
    good = _location_row(location_setup)
    bad = _location_row(location_setup, name="Other")
    bad[0] = "ZZZ"
    commit_response = _commit_location(client, location_setup["hr_admin"], [good, bad])
    log_id = commit_response.json()["bulk_upload_log_id"]

    response = client.get(
        f"{SHARED_ENDPOINT}/bulk-upload/{log_id}/error-report",
        headers=auth_headers(client, location_setup["hr_admin"]),
    )
    assert response.status_code == 200

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    data_rows = list(ws.iter_rows(min_row=5, values_only=True))
    non_empty = [r for r in data_rows if any(c is not None for c in r)]
    assert len(non_empty) == 1
    assert non_empty[0][0] == "ZZZ"


def test_error_report_for_housekeeping_batch(client, housekeeping_setup):
    good = _housekeeping_row(housekeeping_setup)
    bad = _housekeeping_row(housekeeping_setup, bio_id="BIO-2002")
    bad[3] = "Nonexistent Designation"
    commit_response = _commit_housekeeping(client, housekeeping_setup["hr_admin"], [good, bad])
    log_id = commit_response.json()["bulk_upload_log_id"]

    response = client.get(
        f"{SHARED_ENDPOINT}/bulk-upload/{log_id}/error-report",
        headers=auth_headers(client, housekeeping_setup["hr_admin"]),
    )
    assert response.status_code == 200


# --- original file: entity-agnostic proxy ----------------------------------


def test_original_file_download_for_location_batch(client, location_setup):
    rows = [_location_row(location_setup)]
    commit_response = _commit_location(client, location_setup["hr_admin"], rows)
    log_id = commit_response.json()["bulk_upload_log_id"]

    response = client.get(
        f"{SHARED_ENDPOINT}/bulk-uploads/{log_id}/original-file",
        headers=auth_headers(client, location_setup["hr_admin"]),
    )
    assert response.status_code == 200
    assert response.content == _csv_bytes(LOCATION_HEADERS, rows)


# --- undo: created-vs-updated split + not_reverted_count --------------------


def test_undo_location_batch_deactivates_created_row(client, location_setup, db_session):
    commit_response = _commit_location(client, location_setup["hr_admin"], [_location_row(location_setup)])
    log_id = commit_response.json()["bulk_upload_log_id"]
    row_id = db_session.query(Location).one().id

    response = client.post(
        f"{SHARED_ENDPOINT}/bulk-uploads/{log_id}/undo", headers=auth_headers(client, location_setup["hr_admin"])
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["status"] == "UNDONE"
    assert body["reverted_history_count"] == 1
    assert body["not_reverted_count"] == 0

    location = db_session.get(Location, row_id)
    assert location.is_active is False


def test_undo_location_batch_skips_updated_row_and_counts_it(
    client, location_setup, location_factory, db_session
):
    # "Updated" (2026-08-25 fix, see location_import.py's own docstring) now
    # only fires when the composite match key (campus+name+block+floor+
    # category, normalized) already equals an existing row -- i.e. the SAME
    # real-world location -- and only the raw text differs, purely in case/
    # whitespace. A genuinely different block/floor is a different location
    # entirely (CREATED, not UPDATED) -- see
    # test_undo_location_batch_deactivates_created_row for that case.
    existing = location_factory(
        "SSE", name="shared  block", category=StaffRoleCategoryEnum.TEACHING,
        block_building="block a", floor_venue="ground",
    )
    db_session.commit()

    commit_response = _commit_location(client, location_setup["hr_admin"], [_location_row(location_setup)])
    log_id = commit_response.json()["bulk_upload_log_id"]

    response = client.post(
        f"{SHARED_ENDPOINT}/bulk-uploads/{log_id}/undo", headers=auth_headers(client, location_setup["hr_admin"])
    )
    assert response.status_code == 200
    body = response.json()
    assert body["reverted_history_count"] == 0
    assert body["not_reverted_count"] == 1

    db_session.refresh(existing)
    # Not reverted -- the re-stamped canonical text stays (no prior-value
    # history to restore, and the row still exists/still active).
    assert existing.name == "Shared Block"
    assert existing.block_building == "Block A"
    assert existing.floor_venue == "Ground"
    assert existing.is_active is True


def test_undo_housekeeping_batch_deactivates_created_row(client, housekeeping_setup, db_session):
    commit_response = _commit_housekeeping(
        client, housekeeping_setup["hr_admin"], [_housekeeping_row(housekeeping_setup)]
    )
    log_id = commit_response.json()["bulk_upload_log_id"]
    row_id = db_session.query(HousekeepingStaff).one().id

    response = client.post(
        f"{SHARED_ENDPOINT}/bulk-uploads/{log_id}/undo", headers=auth_headers(client, housekeeping_setup["hr_admin"])
    )
    assert response.status_code == 200
    body = response.json()
    assert body["reverted_history_count"] == 1
    assert body["not_reverted_count"] == 0

    staff = db_session.get(HousekeepingStaff, row_id)
    assert staff.is_active is False


def test_undo_writes_row_log_entries_for_location_batch(client, location_setup, db_session):
    commit_response = _commit_location(client, location_setup["hr_admin"], [_location_row(location_setup)])
    log_id = uuid.UUID(commit_response.json()["bulk_upload_log_id"])

    row_logs = db_session.query(BulkUploadRowLog).filter(BulkUploadRowLog.bulk_upload_log_id == log_id).all()
    assert len(row_logs) == 1
    assert row_logs[0].was_created is True
    assert row_logs[0].entity_type.value == "LOCATION"


def test_undo_after_deadline_is_rejected_for_location_batch(client, location_setup, db_session):
    commit_response = _commit_location(client, location_setup["hr_admin"], [_location_row(location_setup)])
    log_id = commit_response.json()["bulk_upload_log_id"]

    log = db_session.get(BulkUploadLog, uuid.UUID(log_id))
    log.undo_deadline = datetime.now(timezone.utc) - timedelta(hours=1)
    db_session.flush()

    response = client.post(
        f"{SHARED_ENDPOINT}/bulk-uploads/{log_id}/undo", headers=auth_headers(client, location_setup["hr_admin"])
    )
    assert response.status_code == 409


def test_undo_twice_is_rejected_for_housekeeping_batch(client, housekeeping_setup):
    commit_response = _commit_housekeeping(
        client, housekeeping_setup["hr_admin"], [_housekeeping_row(housekeeping_setup)]
    )
    log_id = commit_response.json()["bulk_upload_log_id"]

    first = client.post(
        f"{SHARED_ENDPOINT}/bulk-uploads/{log_id}/undo", headers=auth_headers(client, housekeeping_setup["hr_admin"])
    )
    assert first.status_code == 200

    second = client.post(
        f"{SHARED_ENDPOINT}/bulk-uploads/{log_id}/undo", headers=auth_headers(client, housekeeping_setup["hr_admin"])
    )
    assert second.status_code == 409


def test_undo_allowed_for_recruitment_officer_on_location_batch(client, location_setup):
    """The exact case that would have been broken if the shared endpoints
    stayed gated on the narrower SANCTIONED_STRENGTH_WRITE_ROLES: a
    RECRUITMENT_OFFICER who committed a Location batch must be able to undo
    it through this shared endpoint too."""
    commit_response = _commit_location(
        client, location_setup["recruitment_officer"], [_location_row(location_setup)]
    )
    log_id = commit_response.json()["bulk_upload_log_id"]

    response = client.post(
        f"{SHARED_ENDPOINT}/bulk-uploads/{log_id}/undo",
        headers=auth_headers(client, location_setup["recruitment_officer"]),
    )
    assert response.status_code == 200


def test_undo_forbidden_for_non_write_role_on_location_batch(client, location_setup, user_factory):
    commit_response = _commit_location(client, location_setup["hr_admin"], [_location_row(location_setup)])
    log_id = commit_response.json()["bulk_upload_log_id"]
    hod = user_factory(UserRoleEnum.CAMPUS_HOD, campus_code="SSE")

    response = client.post(f"{SHARED_ENDPOINT}/bulk-uploads/{log_id}/undo", headers=auth_headers(client, hod))
    assert response.status_code == 403


def test_sanctioned_strength_undo_still_reports_zero_not_reverted(
    client, campus_factory, department_factory, designation_factory, user_factory
):
    """Backward-compat check: Sanctioned Strength's own undo response now
    carries the new `not_reverted_count` field too, always 0 (every touched
    row has a real SanctionedStrengthHistory.old_value to replay)."""
    campus = campus_factory("SSE")
    department = department_factory("SSE", name=f"Compat Dept {uuid.uuid4().hex[:6]}")
    department.supported_categories = [StaffRoleCategoryEnum.TEACHING]
    designation = designation_factory(StaffRoleCategoryEnum.TEACHING, department=department)
    hr_admin = user_factory(UserRoleEnum.HR_ADMIN)

    rows = [[campus.code, department.name, designation.name, 5, "01-04-2026", ""]]
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(
        ["Campus Code", "Department Name", "Designation Name", "Approved Strength", "Effective From (DD-MM-YYYY)", "Remarks"]
    )
    for row in rows:
        writer.writerow(row)

    commit_response = client.post(
        f"{SHARED_ENDPOINT}/bulk-upload/commit",
        files={"file": ("upload.csv", buf.getvalue().encode("utf-8"), "text/csv")},
        headers=auth_headers(client, hr_admin),
    )
    log_id = commit_response.json()["bulk_upload_log_id"]

    response = client.post(f"{SHARED_ENDPOINT}/bulk-uploads/{log_id}/undo", headers=auth_headers(client, hr_admin))
    assert response.status_code == 200
    body = response.json()
    assert body["not_reverted_count"] == 0


# --- DEPARTMENT (Department Master hardening epic, 2026-08-25) -------------


def test_list_bulk_uploads_entity_type_filter_department(client, department_setup):
    _commit_department(client, department_setup["hr_admin"], [_department_row(department_setup)])

    response = client.get(
        f"{SHARED_ENDPOINT}/bulk-uploads",
        params={"entity_type": "DEPARTMENT"},
        headers=auth_headers(client, department_setup["hr_admin"]),
    )
    assert response.status_code == 200
    body = response.json()
    assert body["total"] >= 1
    assert all(item["entity_type"] == "DEPARTMENT" for item in body["items"])


def test_error_report_for_department_batch(client, department_setup):
    good = _department_row(department_setup)
    bad = _department_row(department_setup, code="OTHER")
    bad[0] = "ZZZ"
    commit_response = _commit_department(client, department_setup["hr_admin"], [good, bad])
    log_id = commit_response.json()["bulk_upload_log_id"]

    response = client.get(
        f"{SHARED_ENDPOINT}/bulk-upload/{log_id}/error-report",
        headers=auth_headers(client, department_setup["hr_admin"]),
    )
    assert response.status_code == 200

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    data_rows = list(ws.iter_rows(min_row=5, values_only=True))
    non_empty = [r for r in data_rows if any(c is not None for c in r)]
    assert len(non_empty) == 1
    assert non_empty[0][0] == "ZZZ"


def test_original_file_download_for_department_batch(client, department_setup):
    rows = [_department_row(department_setup)]
    commit_response = _commit_department(client, department_setup["hr_admin"], rows)
    log_id = commit_response.json()["bulk_upload_log_id"]

    response = client.get(
        f"{SHARED_ENDPOINT}/bulk-uploads/{log_id}/original-file",
        headers=auth_headers(client, department_setup["hr_admin"]),
    )
    assert response.status_code == 200
    assert response.content == _csv_bytes(DEPARTMENT_HEADERS, rows)


def test_undo_department_batch_deactivates_created_row(client, department_setup, db_session):
    commit_response = _commit_department(client, department_setup["hr_admin"], [_department_row(department_setup)])
    log_id = commit_response.json()["bulk_upload_log_id"]
    row_id = db_session.query(Department).filter(Department.code == "SHDEPT").one().id

    response = client.post(
        f"{SHARED_ENDPOINT}/bulk-uploads/{log_id}/undo", headers=auth_headers(client, department_setup["hr_admin"])
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["status"] == "UNDONE"
    assert body["reverted_history_count"] == 1
    assert body["not_reverted_count"] == 0

    department = db_session.get(Department, row_id)
    assert department.is_active is False


def test_undo_department_batch_skips_updated_row_and_counts_it(
    client, department_setup, department_factory, db_session
):
    existing = department_factory(
        "SSE", name="Old Name", code="SHDEPT", category=StaffRoleCategoryEnum.TEACHING, parent_group="Old Group",
    )
    db_session.commit()

    commit_response = _commit_department(client, department_setup["hr_admin"], [_department_row(department_setup)])
    log_id = commit_response.json()["bulk_upload_log_id"]

    response = client.post(
        f"{SHARED_ENDPOINT}/bulk-uploads/{log_id}/undo", headers=auth_headers(client, department_setup["hr_admin"])
    )
    assert response.status_code == 200
    body = response.json()
    assert body["reverted_history_count"] == 0
    assert body["not_reverted_count"] == 1

    db_session.refresh(existing)
    # Not reverted -- the re-stamped values stay (no prior-value history to
    # restore, and the row still exists/still active).
    assert existing.name == "Shared Department"
    assert existing.is_active is True


def test_undo_writes_row_log_entries_for_department_batch(client, department_setup, db_session):
    commit_response = _commit_department(client, department_setup["hr_admin"], [_department_row(department_setup)])
    log_id = uuid.UUID(commit_response.json()["bulk_upload_log_id"])

    row_logs = db_session.query(BulkUploadRowLog).filter(BulkUploadRowLog.bulk_upload_log_id == log_id).all()
    assert len(row_logs) == 1
    assert row_logs[0].was_created is True
    assert row_logs[0].entity_type.value == "DEPARTMENT"


def test_undo_forbidden_for_non_write_role_on_department_batch(client, department_setup, user_factory):
    commit_response = _commit_department(client, department_setup["hr_admin"], [_department_row(department_setup)])
    log_id = commit_response.json()["bulk_upload_log_id"]
    hod = user_factory(UserRoleEnum.CAMPUS_HOD, campus_code="SSE")

    response = client.post(f"{SHARED_ENDPOINT}/bulk-uploads/{log_id}/undo", headers=auth_headers(client, hod))
    assert response.status_code == 403


# --- ELIGIBILITY_RULE (starter regulatory-eligibility-rules feature,
# backend Phase 1) -----------------------------------------------------


def test_list_bulk_uploads_entity_type_filter_eligibility_rule(client, eligibility_rule_setup):
    _commit_eligibility_rule(client, eligibility_rule_setup["hr_admin"], [_eligibility_rule_row(eligibility_rule_setup)])

    response = client.get(
        f"{SHARED_ENDPOINT}/bulk-uploads",
        params={"entity_type": "ELIGIBILITY_RULE"},
        headers=auth_headers(client, eligibility_rule_setup["hr_admin"]),
    )
    assert response.status_code == 200
    body = response.json()
    assert body["total"] >= 1
    assert all(item["entity_type"] == "ELIGIBILITY_RULE" for item in body["items"])


def test_error_report_for_eligibility_rule_batch(client, eligibility_rule_setup):
    good = _eligibility_rule_row(eligibility_rule_setup)
    bad = _eligibility_rule_row(eligibility_rule_setup, position_title="Other")
    bad[0] = "ZZZ"
    commit_response = _commit_eligibility_rule(client, eligibility_rule_setup["hr_admin"], [good, bad])
    log_id = commit_response.json()["bulk_upload_log_id"]

    response = client.get(
        f"{SHARED_ENDPOINT}/bulk-upload/{log_id}/error-report",
        headers=auth_headers(client, eligibility_rule_setup["hr_admin"]),
    )
    assert response.status_code == 200

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    data_rows = list(ws.iter_rows(min_row=5, values_only=True))
    non_empty = [r for r in data_rows if any(c is not None for c in r)]
    assert len(non_empty) == 1
    assert non_empty[0][0] == "ZZZ"


def test_original_file_download_for_eligibility_rule_batch(client, eligibility_rule_setup):
    rows = [_eligibility_rule_row(eligibility_rule_setup)]
    commit_response = _commit_eligibility_rule(client, eligibility_rule_setup["hr_admin"], rows)
    log_id = commit_response.json()["bulk_upload_log_id"]

    response = client.get(
        f"{SHARED_ENDPOINT}/bulk-uploads/{log_id}/original-file",
        headers=auth_headers(client, eligibility_rule_setup["hr_admin"]),
    )
    assert response.status_code == 200
    assert response.content == _csv_bytes(ELIGIBILITY_RULE_HEADERS, rows)


def test_undo_eligibility_rule_batch_deactivates_created_row(client, eligibility_rule_setup, db_session):
    commit_response = _commit_eligibility_rule(
        client, eligibility_rule_setup["hr_admin"], [_eligibility_rule_row(eligibility_rule_setup)]
    )
    log_id = commit_response.json()["bulk_upload_log_id"]
    row_id = db_session.query(EligibilityRule).filter(EligibilityRule.position_title == "Shared Position").one().id

    response = client.post(
        f"{SHARED_ENDPOINT}/bulk-uploads/{log_id}/undo",
        headers=auth_headers(client, eligibility_rule_setup["hr_admin"]),
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["status"] == "UNDONE"
    assert body["reverted_history_count"] == 1
    assert body["not_reverted_count"] == 0

    rule = db_session.get(EligibilityRule, row_id)
    assert rule.is_active is False


def test_undo_writes_row_log_entries_for_eligibility_rule_batch(client, eligibility_rule_setup, db_session):
    commit_response = _commit_eligibility_rule(
        client, eligibility_rule_setup["hr_admin"], [_eligibility_rule_row(eligibility_rule_setup)]
    )
    log_id = uuid.UUID(commit_response.json()["bulk_upload_log_id"])

    row_logs = db_session.query(BulkUploadRowLog).filter(BulkUploadRowLog.bulk_upload_log_id == log_id).all()
    assert len(row_logs) == 1
    assert row_logs[0].was_created is True
    assert row_logs[0].entity_type.value == "ELIGIBILITY_RULE"


def test_undo_forbidden_for_non_write_role_on_eligibility_rule_batch(client, eligibility_rule_setup, user_factory):
    commit_response = _commit_eligibility_rule(
        client, eligibility_rule_setup["hr_admin"], [_eligibility_rule_row(eligibility_rule_setup)]
    )
    log_id = commit_response.json()["bulk_upload_log_id"]
    hod = user_factory(UserRoleEnum.CAMPUS_HOD, campus_code="SSE")

    response = client.post(f"{SHARED_ENDPOINT}/bulk-uploads/{log_id}/undo", headers=auth_headers(client, hod))
    assert response.status_code == 403


# --- DESIGNATION (Designation Master bulk-upload epic, backend Phase 1) ----


def test_list_bulk_uploads_entity_type_filter_designation(client, designation_setup):
    _commit_designation(client, designation_setup["super_admin"], [_designation_row(designation_setup)])

    response = client.get(
        f"{SHARED_ENDPOINT}/bulk-uploads",
        params={"entity_type": "DESIGNATION"},
        headers=auth_headers(client, designation_setup["super_admin"]),
    )
    assert response.status_code == 200
    body = response.json()
    assert body["total"] >= 1
    assert all(item["entity_type"] == "DESIGNATION" for item in body["items"])


def test_error_report_for_designation_batch(client, designation_setup):
    good = _designation_row(designation_setup)
    bad = _designation_row(designation_setup, name="Other Role")
    bad[2] = "NOPE"
    commit_response = _commit_designation(client, designation_setup["super_admin"], [good, bad])
    log_id = commit_response.json()["bulk_upload_log_id"]

    response = client.get(
        f"{SHARED_ENDPOINT}/bulk-upload/{log_id}/error-report",
        headers=auth_headers(client, designation_setup["super_admin"]),
    )
    assert response.status_code == 200

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    data_rows = list(ws.iter_rows(min_row=5, values_only=True))
    non_empty = [r for r in data_rows if any(c is not None for c in r)]
    assert len(non_empty) == 1
    assert non_empty[0][0] == "Other Role"


def test_original_file_download_for_designation_batch(client, designation_setup):
    rows = [_designation_row(designation_setup)]
    commit_response = _commit_designation(client, designation_setup["super_admin"], rows)
    log_id = commit_response.json()["bulk_upload_log_id"]

    response = client.get(
        f"{SHARED_ENDPOINT}/bulk-uploads/{log_id}/original-file",
        headers=auth_headers(client, designation_setup["super_admin"]),
    )
    assert response.status_code == 200
    assert response.content == _csv_bytes(DESIGNATION_HEADERS, rows)


def test_undo_designation_batch_deactivates_created_row(client, designation_setup, db_session):
    commit_response = _commit_designation(client, designation_setup["super_admin"], [_designation_row(designation_setup)])
    log_id = commit_response.json()["bulk_upload_log_id"]
    row_id = db_session.query(Designation).filter(Designation.name == "Shared Designation").one().id

    response = client.post(
        f"{SHARED_ENDPOINT}/bulk-uploads/{log_id}/undo", headers=auth_headers(client, designation_setup["super_admin"])
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["status"] == "UNDONE"
    assert body["reverted_history_count"] == 1
    assert body["not_reverted_count"] == 0

    designation = db_session.get(Designation, row_id)
    assert designation.is_active is False


def test_undo_designation_batch_skips_updated_row_and_counts_it(
    client, designation_setup, designation_factory, db_session
):
    existing = designation_factory(
        StaffRoleCategoryEnum.TEACHING, name="Shared Designation", department=designation_setup["department"]
    )
    db_session.commit()

    commit_response = _commit_designation(client, designation_setup["super_admin"], [_designation_row(designation_setup)])
    log_id = commit_response.json()["bulk_upload_log_id"]

    response = client.post(
        f"{SHARED_ENDPOINT}/bulk-uploads/{log_id}/undo", headers=auth_headers(client, designation_setup["super_admin"])
    )
    assert response.status_code == 200
    body = response.json()
    assert body["reverted_history_count"] == 0
    assert body["not_reverted_count"] == 1

    db_session.refresh(existing)
    assert existing.is_active is True


def test_undo_writes_row_log_entries_for_designation_batch(client, designation_setup, db_session):
    commit_response = _commit_designation(client, designation_setup["super_admin"], [_designation_row(designation_setup)])
    log_id = uuid.UUID(commit_response.json()["bulk_upload_log_id"])

    row_logs = db_session.query(BulkUploadRowLog).filter(BulkUploadRowLog.bulk_upload_log_id == log_id).all()
    assert len(row_logs) == 1
    assert row_logs[0].was_created is True
    assert row_logs[0].entity_type.value == "DESIGNATION"


def test_undo_allowed_for_recruitment_coordinator_on_designation_batch(client, designation_setup, user_factory):
    """The exact case that would break if the shared endpoints stayed gated
    on the pre-Designation-epic `_SHARED_BULK_UPLOAD_ROLES` tuple: a
    RECRUITMENT_COORDINATOR who committed a Designation batch must be able
    to undo it through this shared endpoint too (DESIGNATION_WRITE_ROLES
    includes RECRUITMENT_COORDINATOR, not RECRUITMENT_OFFICER)."""
    coordinator = user_factory(UserRoleEnum.RECRUITMENT_COORDINATOR)
    commit_response = _commit_designation(client, coordinator, [_designation_row(designation_setup)])
    log_id = commit_response.json()["bulk_upload_log_id"]

    response = client.post(
        f"{SHARED_ENDPOINT}/bulk-uploads/{log_id}/undo", headers=auth_headers(client, coordinator)
    )
    assert response.status_code == 200


def test_undo_forbidden_for_non_write_role_on_designation_batch(client, designation_setup, user_factory):
    commit_response = _commit_designation(client, designation_setup["super_admin"], [_designation_row(designation_setup)])
    log_id = commit_response.json()["bulk_upload_log_id"]
    hod = user_factory(UserRoleEnum.CAMPUS_HOD, campus_code="SSE")

    response = client.post(f"{SHARED_ENDPOINT}/bulk-uploads/{log_id}/undo", headers=auth_headers(client, hod))
    assert response.status_code == 403


def test_error_report_escapes_formula_text_from_the_rejected_row(client, location_setup):
    """Audit H3 (2026-09-03): an uploader's own text is echoed back in the
    error report, so a rejected row that reads like a formula must come
    back as text, never as an executable cell."""
    payload = "=cmd|' /C calc'!A0"
    bad = _location_row(location_setup, name=payload, block="+Block", floor="-1")
    bad[0] = "ZZZ"  # unknown campus -> rejected -> lands in the error report
    commit_response = _commit_location(client, location_setup["hr_admin"], [bad])
    log_id = commit_response.json()["bulk_upload_log_id"]

    response = client.get(
        f"{SHARED_ENDPOINT}/bulk-upload/{log_id}/error-report",
        headers=auth_headers(client, location_setup["hr_admin"]),
    )
    assert response.status_code == 200

    from openpyxl import load_workbook

    ws = load_workbook(io.BytesIO(response.content)).active
    cells = [c for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert all(c.data_type == "s" for c in cells)
    assert not any(c.value.startswith(("=", "+", "-", "@", "\t", "\r")) for c in cells)
    values = {c.value for c in cells}
    assert "'" + payload in values and "'+Block" in values and "'-1" in values
