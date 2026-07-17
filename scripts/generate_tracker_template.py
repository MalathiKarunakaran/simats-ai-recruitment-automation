"""Generates data/SIMATS_Recruitment_Tracker_TEMPLATE.xlsx -- the real
manual-workflow data-entry template that app/services/tracker_import.py
imports (via POST /migration/import-tracker-workbook). Not part of the app
at runtime; a one-off/rerunnable ops script, same precedent as
scripts/load_test.py and app/db/seed.py.

Usage: venv/Scripts/python.exe scripts/generate_tracker_template.py
"""

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.models.enums import APPLICATION_STATUS_ORDER, CAMPUS_CODES  # noqa: E402

OUTPUT_PATH = Path(__file__).resolve().parents[1] / "data" / "SIMATS_Recruitment_Tracker_TEMPLATE.xlsx"

STAFF_CATEGORIES = ("Teaching", "Non-Teaching", "Housekeeping")
SOURCES = ("Reference", "Job Portal", "FacultyPlus", "Walk-in")
PRIORITIES = ("LOW", "NORMAL", "HIGH", "URGENT")
STATUSES = tuple(s.value for s in APPLICATION_STATUS_ORDER) + ("REJECTED", "WITHDRAWN")

VACANCY_HEADERS = (
    "Request ID",
    "Date Requested",
    "Campus",
    "Department",
    "Staff Category",
    "Position",
    "Vacancies Approved",
    "JD Status",
    "Priority",
    "Vacancy Status",
    "Requested By",
)

CANDIDATE_HEADERS = (
    "Candidate ID",
    "Vacancy Ref",
    "Name",
    "Contact Number",
    "Email",
    "Campus",
    "Department",
    "Staff Category",
    "Position Applied",
    "Source",
    "Reference Name",
    "Resume Received Date",
    "Called Date",
    "Interview Scheduled Date",
    "Panel Members",
    "Panel Result",
    "Panel Remarks",
    "Salary Fixed",
    "Offer Given Date",
    "Joining Confirmed (Y/N)",
    "Expected Joining Date",
    "Actual Joining Date",
    "Department Allotted",
    "Room Allotted",
    "Orientation Date",
    "HOD Assigned",
    "Status",
    "Remarks",
)

HEADER_FILL = PatternFill(start_color="1B5FAA", end_color="1B5FAA", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SAMPLE_FILL = PatternFill(start_color="FFF3D6", end_color="FFF3D6", fill_type="solid")


def _write_header(ws, headers: tuple[str, ...]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20


def _add_list_validation(ws, column_letter: str, values: tuple[str, ...], max_row: int = 500) -> None:
    formula = '"' + ",".join(values) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"{column_letter}2:{column_letter}{max_row}")


def _mark_sample_row(ws, num_columns: int, row: int = 2) -> None:
    for col in range(1, num_columns + 1):
        ws.cell(row=row, column=col).fill = SAMPLE_FILL


def build_workbook() -> Workbook:
    wb = Workbook()

    vacancy_ws = wb.active
    vacancy_ws.title = "Vacancy Tracker"
    _write_header(vacancy_ws, VACANCY_HEADERS)
    vacancy_ws.append(
        [
            "REQ-2026-001",
            "2026-07-01",
            "SSE",
            "Computer Science",
            "Teaching",
            "Assistant Professor",
            2,
            "Posted",
            "NORMAL",
            "Open",
            "Dr. Sample HOD",
        ]
    )
    _mark_sample_row(vacancy_ws, len(VACANCY_HEADERS))
    _add_list_validation(vacancy_ws, "C", CAMPUS_CODES)  # Campus
    _add_list_validation(vacancy_ws, "E", STAFF_CATEGORIES)  # Staff Category
    _add_list_validation(vacancy_ws, "I", PRIORITIES)  # Priority

    candidate_ws = wb.create_sheet("Candidate Pipeline")
    _write_header(candidate_ws, CANDIDATE_HEADERS)
    candidate_ws.append(
        [
            "CAND-0001",
            "REQ-2026-001",
            "Sample Candidate",
            "9999999999",
            "sample.candidate@example.com",
            "SSE",
            "Computer Science",
            "Teaching",
            "Assistant Professor",
            "Reference",
            "Sample Referrer",
            "2026-07-05",
            "2026-07-10",
            "2026-07-12",
            "Dr. A, Dr. B",
            "Selected",
            "Strong candidate",
            75000,
            "2026-07-15",
            "Y",
            "2026-08-01",
            "",
            "",
            "",
            "",
            "",
            "SELECTED",
            "",
        ]
    )
    _mark_sample_row(candidate_ws, len(CANDIDATE_HEADERS))
    _add_list_validation(candidate_ws, "F", CAMPUS_CODES)  # Campus
    _add_list_validation(candidate_ws, "H", STAFF_CATEGORIES)  # Staff Category
    _add_list_validation(candidate_ws, "J", SOURCES)  # Source
    _add_list_validation(candidate_ws, "AA", STATUSES)  # Status (column 27)

    master_ws = wb.create_sheet("Master Lists")
    master_ws.append(["Campus Codes", "Staff Category", "Source", "Priority", "Application Status"])
    for cell in master_ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    columns = [CAMPUS_CODES, STAFF_CATEGORIES, SOURCES, PRIORITIES, STATUSES]
    for row_idx in range(max(len(c) for c in columns)):
        for col_idx, values in enumerate(columns, start=1):
            if row_idx < len(values):
                master_ws.cell(row=row_idx + 2, column=col_idx, value=values[row_idx])
    for col in range(1, 6):
        master_ws.column_dimensions[chr(64 + col)].width = 22
    master_ws.sheet_view.showGridLines = True

    return wb


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
