"""Stage B of the real-manual-workflow rewrite: import
`SIMATS_Recruitment_Tracker_TEMPLATE.xlsx` (see
scripts/generate_tracker_template.py) as real, live data -- not sample data,
and not a DRAFT-only staging area like app/services/migration.py's legacy
CSV importer. Vacancy Tracker rows land as already-published, already-slotted
vacancies (these are real ongoing drives, not new requests awaiting the
approval chain); Candidate Pipeline rows land as Applications with their
real current pipeline status.

Re-runnable by design: rows are matched and upserted by the sheet's own
Request ID / Candidate ID (VacancyRequest.external_ref / Application.external_ref),
not recreated on every import.

Deliberately does NOT create synthetic InterviewSchedule/Offer/JoiningRecord
rows to back a candidate's historical status -- doing so would require
fabricating panel-member User accounts and offer records that don't
correspond to anything real. Panel members/result/remarks, salary, and
joining/allotment/orientation/HOD fields are stored as flat Application
columns instead (see the Stage A model change). What this importer DOES
keep correct is HiringSlot state (reserve/fill/auto-close) since /vacancies,
/candidates, /reports and the dashboard all depend on it being accurate for
real data -- but it does so silently (no "you've been selected!" emails to
real candidates for outcomes that already happened; see _sync_hiring_slot).
"""

import io
import secrets
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.application import Application
from app.models.approved_vacancy import ApprovedVacancy
from app.models.campus import Campus
from app.models.candidate import Candidate
from app.models.department import Department
from app.models.designation import Designation
from app.models.enums import (
    APPLICATION_STATUS_ORDER,
    CAMPUS_CODES,
    ApplicationStatusEnum,
    EmploymentTypeEnum,
    HiringSlotStatusEnum,
    StaffRoleCategoryEnum,
    VacancyPriorityEnum,
    VacancyRequestStatusEnum,
)
from app.models.hiring_slot import HiringSlot
from app.models.job_posting import JobPosting
from app.models.user import User
from app.models.vacancy_request import VacancyRequest
from app.services.audit import log_event

VACANCY_SHEET_NAME = "Vacancy Tracker"
CANDIDATE_SHEET_NAME = "Candidate Pipeline"
SAMPLE_REQUEST_ID = "REQ-2026-001"
SAMPLE_CANDIDATE_ID = "CAND-0001"

SOURCE_VALUES = ("Reference", "Job Portal", "FacultyPlus", "Walk-in")
VALID_APPLICATION_STATUSES = {s.value for s in APPLICATION_STATUS_ORDER} | {"REJECTED", "WITHDRAWN"}

# A candidate at or beyond SELECTED holds a reserved slot; at or beyond
# JOINED, that slot is FILLED. Index-based, same comparison style as
# app/services/pipeline.py::advance_if_behind.
_SELECTED_INDEX = APPLICATION_STATUS_ORDER.index(ApplicationStatusEnum.SELECTED)
_JOINED_INDEX = APPLICATION_STATUS_ORDER.index(ApplicationStatusEnum.JOINED)


def _resolve_designation(db: Session, position_title: str) -> Designation | None:
    """Best-effort, case-insensitive match of the sheet's Position text
    against the Designation Master. No match is NOT a hard failure -- the
    row still imports on position_title alone (designation_id stays null);
    the caller surfaces a warning so a coordinator can fix master data or
    the sheet, matching this file's established flagged-not-dropped pattern
    for every other soft-validation case."""
    if not position_title:
        return None
    return (
        db.query(Designation)
        .filter(Designation.name.ilike(position_title))
        .one_or_none()
    )


def _get_or_create_department(db: Session, campus: Campus, name: str) -> Department:
    department = (
        db.query(Department).filter(Department.campus_id == campus.id, Department.name == name).one_or_none()
    )
    if department is None:
        department = Department(campus_id=campus.id, name=name)
        db.add(department)
        db.flush()
    return department


def _cell(row: dict, key: str) -> str:
    value = row.get(key)
    if value is None:
        return ""
    return str(value).strip()


def _parse_date(row: dict, key: str) -> tuple[date | None, str | None]:
    raw = row.get(key)
    if raw in (None, ""):
        return None, None
    if isinstance(raw, datetime):
        return raw.date(), None
    if isinstance(raw, date):
        return raw, None
    try:
        return datetime.strptime(str(raw).strip(), "%Y-%m-%d").date(), None
    except ValueError:
        return None, f"Invalid date for '{key}': '{raw}' (expected YYYY-MM-DD)"


def _parse_decimal(row: dict, key: str) -> tuple[float | None, str | None]:
    raw = row.get(key)
    if raw in (None, ""):
        return None, None
    try:
        return float(Decimal(str(raw))), None
    except (InvalidOperation, ValueError):
        return None, f"Invalid number for '{key}': '{raw}'"


def _read_sheet_rows(workbook, sheet_name: str) -> list[dict]:
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    except StopIteration:
        return []
    result = []
    for raw_row in rows:
        if all(cell is None or str(cell).strip() == "" for cell in raw_row):
            continue
        result.append(dict(zip(headers, raw_row)))
    return result


def _import_vacancy_row(db: Session, row: dict, row_number: int, actor: User) -> tuple[VacancyRequest | None, list[str]]:
    """Returns (VacancyRequest, warnings) on success -- warnings is empty
    unless the Position text couldn't be matched to Designation Master, in
    which case the row still imports and the caller marks it
    "imported_with_warning" (see _import_candidate_row for the identical
    convention already used for candidate rows). Returns (None, errors) on
    hard validation failure."""
    errors: list[str] = []
    request_id = _cell(row, "Request ID")
    if not request_id:
        errors.append("Missing required column 'Request ID'")

    campus_code = _cell(row, "Campus").upper()
    campus = None
    if not campus_code:
        errors.append("Missing required column 'Campus'")
    elif campus_code not in CAMPUS_CODES:
        errors.append(f"Unknown Campus '{campus_code}' -- must be one of {', '.join(CAMPUS_CODES)}")
    else:
        campus = db.query(Campus).filter(Campus.code == campus_code).one_or_none()
        if campus is None:
            errors.append(f"Campus '{campus_code}' is not seeded in this environment")

    department_name = _cell(row, "Department")
    if not department_name:
        errors.append("Missing required column 'Department'")

    role_category_raw = _cell(row, "Staff Category").upper().replace(" ", "_")
    role_category = None
    if not role_category_raw:
        errors.append("Missing required column 'Staff Category'")
    else:
        try:
            role_category = StaffRoleCategoryEnum[role_category_raw]
        except KeyError:
            errors.append(
                f"Invalid Staff Category '{row.get('Staff Category')}' -- must be Teaching, Non-Teaching, or Housekeeping"
            )

    position_title = _cell(row, "Position")
    if not position_title:
        errors.append("Missing required column 'Position'")

    requested_count_raw = _cell(row, "Vacancies Approved")
    requested_count = None
    if not requested_count_raw:
        errors.append("Missing required column 'Vacancies Approved'")
    else:
        try:
            requested_count = int(float(requested_count_raw))
            if requested_count <= 0:
                errors.append("'Vacancies Approved' must be positive")
        except ValueError:
            errors.append(f"Invalid 'Vacancies Approved' value '{requested_count_raw}'")

    priority = VacancyPriorityEnum.NORMAL
    priority_raw = _cell(row, "Priority").upper()
    if priority_raw:
        try:
            priority = VacancyPriorityEnum[priority_raw]
        except KeyError:
            errors.append(f"Invalid Priority '{row.get('Priority')}'")

    date_requested, date_error = _parse_date(row, "Date Requested")
    if date_error:
        errors.append(date_error)

    if errors:
        return None, errors

    department = _get_or_create_department(db, campus, department_name)

    designation = _resolve_designation(db, position_title)
    warnings: list[str] = []
    if designation is None:
        warnings.append(
            f"Designation '{position_title}' was not found in Designation Master -- imported using "
            "free-text Position only; add it to master data or fix the sheet"
        )

    requested_by_name = _cell(row, "Requested By")
    requested_by = None
    if requested_by_name:
        requested_by = (
            db.query(User)
            .filter(User.campus_id == campus.id, User.full_name.ilike(requested_by_name))
            .one_or_none()
        )
    if requested_by is None:
        requested_by = actor  # best effort -- see module docstring

    existing = db.query(VacancyRequest).filter(VacancyRequest.external_ref == request_id).one_or_none()

    if existing is not None:
        existing.position_title = position_title
        existing.priority = priority
        if designation is not None:
            existing.designation_id = designation.id
        vacancy_request = existing
        # Slot count changes on re-import are a manual admin action, not
        # something this importer resizes automatically -- shrinking a
        # slot pool that already has reservations is a real business
        # decision, not a data-entry correction.
        if vacancy_request.approved_vacancy and requested_count > vacancy_request.approved_vacancy.total_positions:
            additional = requested_count - vacancy_request.approved_vacancy.total_positions
            existing_slot_numbers = [s.slot_number for s in vacancy_request.approved_vacancy.hiring_slots]
            next_slot_number = max(existing_slot_numbers, default=0) + 1
            for i in range(additional):
                db.add(
                    HiringSlot(
                        approved_vacancy_id=vacancy_request.approved_vacancy.id,
                        slot_number=next_slot_number + i,
                    )
                )
            vacancy_request.approved_vacancy.total_positions = requested_count
        db.flush()
        return vacancy_request, warnings

    vacancy_request = VacancyRequest(
        campus_id=campus.id,
        department_id=department.id,
        designation_id=designation.id if designation is not None else None,
        role_category=role_category,
        position_title=position_title,
        employment_type=EmploymentTypeEnum.FULL_TIME,
        requested_count=requested_count,
        qualification="Imported from tracker workbook",
        experience_required="Imported from tracker workbook",
        priority=priority,
        requested_by_id=requested_by.id,
        status=VacancyRequestStatusEnum.PUBLISHED,
        external_ref=request_id,
    )
    if date_requested is not None:
        vacancy_request.created_at = datetime.combine(date_requested, datetime.min.time(), tzinfo=timezone.utc)
    db.add(vacancy_request)
    db.flush()

    now = datetime.now(timezone.utc)
    approved_vacancy = ApprovedVacancy(
        vacancy_request_id=vacancy_request.id,
        campus_id=campus.id,
        total_positions=requested_count,
        approved_by_id=actor.id,
        approved_at=now,
    )
    db.add(approved_vacancy)
    db.flush()

    for slot_number in range(1, requested_count + 1):
        db.add(HiringSlot(approved_vacancy_id=approved_vacancy.id, slot_number=slot_number))

    slug = f"{campus.code}-{position_title}".lower().replace(" ", "-") + f"-{secrets.token_hex(4)}"
    job_posting = JobPosting(
        approved_vacancy_id=approved_vacancy.id,
        campus_id=campus.id,
        role_category=role_category,
        public_apply_slug=slug,
        published_at=now,
    )
    db.add(job_posting)
    db.flush()

    return vacancy_request, warnings


def _sync_hiring_slot(
    db: Session, *, application: Application, new_status: ApplicationStatusEnum, actor: User
) -> str | None:
    """Keep HiringSlot state consistent with an imported historical status,
    without notifications (see module docstring) -- reserve on
    SELECTED-or-later, fill (+ maybe auto-close) on JOINED-or-later, release
    on REJECTED/WITHDRAWN. Returns a warning string if a slot couldn't be
    reserved (data issue worth flagging), else None."""
    approved_vacancy_id = application.job_posting.approved_vacancy_id

    if new_status in (ApplicationStatusEnum.REJECTED, ApplicationStatusEnum.WITHDRAWN):
        slot = db.execute(
            select(HiringSlot).where(HiringSlot.reserved_application_id == application.id)
        ).scalar_one_or_none()
        if slot is not None and slot.status != HiringSlotStatusEnum.FILLED:
            slot.status = HiringSlotStatusEnum.OPEN
            slot.reserved_application_id = None
            slot.released_at = datetime.now(timezone.utc)
        return None

    if new_status not in APPLICATION_STATUS_ORDER or APPLICATION_STATUS_ORDER.index(new_status) < _SELECTED_INDEX:
        return None

    slot = db.execute(
        select(HiringSlot).where(HiringSlot.reserved_application_id == application.id)
    ).scalar_one_or_none()

    if slot is None:
        slot = db.execute(
            select(HiringSlot)
            .where(HiringSlot.approved_vacancy_id == approved_vacancy_id, HiringSlot.status == HiringSlotStatusEnum.OPEN)
            .order_by(HiringSlot.slot_number)
            .limit(1)
        ).scalar_one_or_none()
        if slot is None:
            return "No open hiring slot remains to reserve for this candidate's status"
        slot.status = HiringSlotStatusEnum.RESERVED
        slot.reserved_application_id = application.id
        slot.reserved_at = datetime.now(timezone.utc)

    if APPLICATION_STATUS_ORDER.index(new_status) < _JOINED_INDEX:
        return None

    if slot.status != HiringSlotStatusEnum.FILLED:
        slot.status = HiringSlotStatusEnum.FILLED
        slot.filled_at = datetime.now(timezone.utc)

    approved_vacancy = db.get(ApprovedVacancy, approved_vacancy_id)
    remaining_unfilled = db.execute(
        select(HiringSlot.id).where(
            HiringSlot.approved_vacancy_id == approved_vacancy_id, HiringSlot.status != HiringSlotStatusEnum.FILLED
        )
    ).scalars().all()
    if not remaining_unfilled and approved_vacancy.closed_at is None:
        now = datetime.now(timezone.utc)
        approved_vacancy.closed_at = now
        if application.job_posting is not None:
            application.job_posting.closed_at = now
            application.job_posting.is_active = False
        vacancy_request = db.get(VacancyRequest, approved_vacancy.vacancy_request_id)
        vacancy_request.status = VacancyRequestStatusEnum.CLOSED
    return None


def _import_candidate_row(
    db: Session, row: dict, row_number: int, actor: User
) -> tuple[Application | None, list[str]]:
    errors: list[str] = []

    candidate_ref = _cell(row, "Candidate ID")
    if not candidate_ref:
        errors.append("Missing required column 'Candidate ID'")

    vacancy_ref = _cell(row, "Vacancy Ref")
    vacancy_request = None
    if not vacancy_ref:
        errors.append("Missing required column 'Vacancy Ref'")
    else:
        vacancy_request = db.query(VacancyRequest).filter(VacancyRequest.external_ref == vacancy_ref).one_or_none()
        if vacancy_request is None or vacancy_request.approved_vacancy is None or vacancy_request.approved_vacancy.job_posting is None:
            errors.append(f"Unknown Vacancy Ref '{vacancy_ref}' (import the Vacancy Tracker sheet first)")

    full_name = _cell(row, "Name")
    if not full_name:
        errors.append("Missing required column 'Name'")

    email = _cell(row, "Email")
    if not email:
        errors.append("Missing required column 'Email'")

    source = _cell(row, "Source")
    if source and source not in SOURCE_VALUES:
        errors.append(f"Invalid Source '{source}' -- must be one of {', '.join(SOURCE_VALUES)}")

    reference_name = _cell(row, "Reference Name")
    if source == "Reference" and not reference_name:
        errors.append("Reference Name is required when Source is 'Reference'")

    status_raw = _cell(row, "Status").upper().replace(" ", "_")
    application_status = None
    if not status_raw:
        errors.append("Missing required column 'Status'")
    elif status_raw not in VALID_APPLICATION_STATUSES:
        errors.append(f"Invalid Status '{row.get('Status')}' -- must be one of {', '.join(sorted(VALID_APPLICATION_STATUSES))}")
    else:
        application_status = ApplicationStatusEnum[status_raw]

    date_fields = {}
    for column, field in (
        ("Resume Received Date", "applied_at"),
        ("Called Date", "called_date"),
        ("Interview Scheduled Date", "interview_scheduled_date"),
        ("Offer Given Date", "offer_given_date"),
        ("Expected Joining Date", "expected_joining_date"),
        ("Actual Joining Date", "actual_joining_date"),
    ):
        value, error = _parse_date(row, column)
        if error:
            errors.append(error)
        date_fields[field] = value

    salary_fixed, salary_error = _parse_decimal(row, "Salary Fixed")
    if salary_error:
        errors.append(salary_error)

    if errors:
        return None, errors

    candidate = db.query(Candidate).filter(Candidate.email == email).one_or_none()
    if candidate is None:
        candidate = Candidate(
            full_name=full_name,
            email=email,
            phone_number=_cell(row, "Contact Number") or None,
            source=source or None,
            reference_name=reference_name or None,
        )
        db.add(candidate)
        db.flush()
    else:
        candidate.full_name = full_name
        candidate.source = source or candidate.source
        candidate.reference_name = reference_name or candidate.reference_name

    job_posting = vacancy_request.approved_vacancy.job_posting
    existing = db.query(Application).filter(Application.external_ref == candidate_ref).one_or_none()

    applied_at = date_fields["applied_at"]
    applied_at_dt = (
        datetime.combine(applied_at, datetime.min.time(), tzinfo=timezone.utc)
        if applied_at is not None
        else datetime.now(timezone.utc)
    )

    if existing is not None:
        application = existing
    else:
        if (
            db.query(Application)
            .filter(Application.candidate_id == candidate.id, Application.job_posting_id == job_posting.id)
            .one_or_none()
            is not None
        ):
            return None, [f"Candidate '{email}' has already applied to Vacancy Ref '{vacancy_ref}'"]
        application = Application(
            candidate_id=candidate.id,
            job_posting_id=job_posting.id,
            campus_id=job_posting.campus_id,
            role_category=job_posting.role_category,
            applied_at=applied_at_dt,
            recorded_by_id=actor.id,
            external_ref=candidate_ref,
        )
        db.add(application)
        db.flush()

    application.panel_members = _cell(row, "Panel Members") or None
    application.panel_result = _cell(row, "Panel Result") or None
    application.panel_remarks = _cell(row, "Panel Remarks") or None
    application.salary_fixed = salary_fixed
    application.called_date = date_fields["called_date"]
    application.interview_scheduled_date = date_fields["interview_scheduled_date"]
    application.offer_given_date = date_fields["offer_given_date"]
    application.expected_joining_date = date_fields["expected_joining_date"]
    application.actual_joining_date = date_fields["actual_joining_date"]
    application.room_allotted = _cell(row, "Room Allotted") or None
    department_allotted_name = _cell(row, "Department Allotted")
    if department_allotted_name:
        department_allotted = _get_or_create_department(db, job_posting.campus, department_allotted_name)
        application.department_allotted_id = department_allotted.id
    orientation_date, _ = _parse_date(row, "Orientation Date")
    application.orientation_date = orientation_date
    application.hod_assigned = _cell(row, "HOD Assigned") or None

    if application_status == ApplicationStatusEnum.REJECTED:
        application.rejection_reason = _cell(row, "Remarks") or "Imported from tracker workbook"
    elif application_status == ApplicationStatusEnum.WITHDRAWN:
        application.withdrawn_reason = _cell(row, "Remarks") or "Imported from tracker workbook"
    application.status = application_status
    db.flush()

    warning = _sync_hiring_slot(db, application=application, new_status=application_status, actor=actor)
    return application, [warning] if warning else []


def import_tracker_workbook(
    db: Session, *, workbook_bytes: bytes, actor: User, request=None, include_sample: bool = False
) -> dict:
    workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True)

    vacancy_rows = _read_sheet_rows(workbook, VACANCY_SHEET_NAME)
    candidate_rows = _read_sheet_rows(workbook, CANDIDATE_SHEET_NAME)

    vacancy_results: list[dict] = []
    vacancy_created_or_updated = 0
    for row_number, row in enumerate(vacancy_rows, start=2):
        if not include_sample and _cell(row, "Request ID") == SAMPLE_REQUEST_ID:
            continue
        vacancy_request, warnings = _import_vacancy_row(db, row, row_number, actor)
        if vacancy_request is None:
            vacancy_results.append({"row_number": row_number, "status": "flagged", "errors": warnings})
            continue
        vacancy_created_or_updated += 1
        vacancy_results.append(
            {
                "row_number": row_number,
                "status": "imported" if not warnings else "imported_with_warning",
                "errors": warnings,
                "vacancy_request_id": vacancy_request.id,
            }
        )

    candidate_results: list[dict] = []
    candidate_created_or_updated = 0
    for row_number, row in enumerate(candidate_rows, start=2):
        if not include_sample and _cell(row, "Candidate ID") == SAMPLE_CANDIDATE_ID:
            continue
        application, warnings = _import_candidate_row(db, row, row_number, actor)
        if application is None:
            candidate_results.append({"row_number": row_number, "status": "flagged", "errors": warnings})
            continue
        candidate_created_or_updated += 1
        candidate_results.append(
            {
                "row_number": row_number,
                "status": "imported" if not warnings else "imported_with_warning",
                "errors": warnings,
                "application_id": application.id,
            }
        )

    log_event(
        db,
        actor=actor,
        action="TRACKER_WORKBOOK_IMPORTED",
        entity_type="VacancyRequest",
        after_state={
            "vacancy_rows": len(vacancy_results),
            "vacancy_imported": vacancy_created_or_updated,
            "candidate_rows": len(candidate_results),
            "candidate_imported": candidate_created_or_updated,
        },
        request=request,
    )

    return {
        "vacancy_total_rows": len(vacancy_results),
        "vacancy_imported_count": vacancy_created_or_updated,
        "vacancy_flagged_count": len(vacancy_results) - vacancy_created_or_updated,
        "vacancy_rows": vacancy_results,
        "candidate_total_rows": len(candidate_results),
        "candidate_imported_count": candidate_created_or_updated,
        "candidate_flagged_count": len(candidate_results) - candidate_created_or_updated,
        "candidate_rows": candidate_results,
    }
