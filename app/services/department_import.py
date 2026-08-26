"""Department bulk upload: validate -> preview -> commit (Department Master
hardening epic, backend Phase 1, 2026-08-25). 4th sibling module to
`sanctioned_strength_import.py`/`location_import.py`/
`housekeeping_staff_import.py` -- same dataclass/validate/commit/undo-adjacent
shape, deliberately NOT a shared abstraction between entities (see
`location_import.py`'s own module docstring for the standing rationale, not
repeated per module). `parse_rows` (file bytes -> `list[dict]`, entity-
agnostic) is reused as-is from `sanctioned_strength_import`, same as the
other two siblings.

**UPSERT key**: `(Campus Code, Department Code)` only, normalized (trim +
casefold, same `_normalize` technique as `location_import.py`'s own helper)
-- deliberately narrower than a full-row match: Department Name is NOT part
of the identity (per the brief, "names may change"), so a changed name on an
already-known (campus, code) pair is a legitimate **Updated** row, not a new/
duplicate one. This is the opposite shape from Location's own 2026-08-25 fix
(which *widened* its key) -- Department's key is deliberately narrow,
matching Sanctioned Strength's own existing (narrower-than-full-row) key
precedent.

**No DB-level uniqueness on (campus_id, code) yet**: unlike
`HousekeepingStaff` (which has a real `UniqueConstraint` on its own upsert
key and so can resolve matches via `.one_or_none()`), `Department` does NOT
have a `UniqueConstraint("campus_id", "code")` -- a live-data check before
this epic's own migration found a genuine pre-existing collision (3
departments on campus SHIFT all sharing `code='SHIFT'`), so that constraint
was deliberately deferred (see
`ebafe3ba100c_department_master_description_field.py`'s own docstring).
Existing-row resolution below therefore follows `location_import.py`'s own
defensive pattern instead: every Department for a given campus is fetched
once (not once per row) and matched in Python by normalized code, picking
the first match (ordered by `created_at`) in the rare case pre-existing data
already has a genuine collision -- same accepted-limitation shape Location
uses for its own not-yet-constrained composite key.

**The "Row N: Department Code already exists for campus X" collision
message**: with a key this narrow (campus+code only), a row whose key
matches an existing Department is *always* the same real-world department by
definition -- there is no way for two rows sharing a key to be "two
genuinely different departments colliding," only "the same department,
possibly with some fields changed" (an Updated row). So that literal
rejection message is unreachable from *this* validator and is intentionally
never raised here -- it lives instead on `app/api/v1/routers/departments.py`'s
own `create_department`/`update_department` Code+Campus uniqueness check
(a 400, not a bulk-upload row rejection), which is the one code path that
really can be asked to introduce a second, genuinely different department
under an already-used code. The in-file exact-duplicate-key case (two rows
in the *same upload* both keying to the same normalized campus+code) is a
real, distinct rejection, handled below.

**Undo scope**: same as `location_import.py`/`housekeeping_staff_import.py`
-- Department has no permanent old-value history table, so `commit_rows`
writes one `BulkUploadRowLog` row per non-rejected row (created/updated/
unchanged alike), which `undo_bulk_upload`
(app/api/v1/routers/sanctioned_strength.py) uses to soft-delete only the
rows this batch actually created.
"""

import io
import re
import uuid
from dataclasses import dataclass

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from app.models.bulk_upload_log import BulkUploadLog
from app.models.bulk_upload_row_log import BulkUploadRowLog
from app.models.campus import Campus
from app.models.department import Department
from app.models.enums import CAMPUS_CODES, BulkUploadEntityTypeEnum, StaffRoleCategoryEnum
from app.services.sanctioned_strength_import import parse_rows  # noqa: F401 -- re-exported, see module docstring

MAX_ROWS = 5000

CAMPUS_HEADER = "Campus Code"
CODE_HEADER = "Department Code"
NAME_HEADER = "Department Name"
CATEGORY_HEADER = "Category"
PARENT_GROUP_HEADER = "Parent Group"
DESCRIPTION_HEADER = "Description"
ACTIVE_HEADER = "Active"

TEMPLATE_HEADERS = (
    CAMPUS_HEADER,
    CODE_HEADER,
    NAME_HEADER,
    CATEGORY_HEADER,
    PARENT_GROUP_HEADER,
    DESCRIPTION_HEADER,
    ACTIVE_HEADER,
)

# Same functional backstop as the other 3 importers' own example rows --
# "XXX" is never a real campus code, so a forgotten example row always comes
# back rejected rather than silently importing as real data.
_EXAMPLE_ROWS = (
    ("XXX", "CSE", "EXAMPLE - DELETE THIS ROW - Computer Science Engineering", "TEACHING", "Engineering", "", "TRUE"),
    ("XXX", "IT", "EXAMPLE - DELETE THIS ROW - Information Technology", "TEACHING", "Engineering", "", "TRUE"),
)

_HEADER_FILL = PatternFill(start_color="1B5FAA", end_color="1B5FAA", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SAMPLE_FILL = PatternFill(start_color="FFF3D6", end_color="FFF3D6", fill_type="solid")

_CATEGORY_VALUES = tuple(c.value for c in StaffRoleCategoryEnum)


@dataclass
class ImportRowResult:
    row_number: int
    status: str  # created | updated | unchanged | rejected
    error_reason: str | None = None
    campus_code: str | None = None
    department_code: str | None = None
    department_name: str | None = None
    category: StaffRoleCategoryEnum | None = None
    parent_group: str | None = None
    description: str | None = None
    is_active: bool | None = None
    # Resolved only for non-rejected rows -- commit_rows() uses these
    # directly rather than re-resolving lookups a second time.
    campus_id: uuid.UUID | None = None
    department_id: uuid.UUID | None = None  # the existing match's id, when status != "created"


@dataclass
class ValidationResult:
    rows: list[ImportRowResult]
    total: int
    created_count: int
    updated_count: int
    unchanged_count: int
    rejected_count: int


def _cell(row: dict, key: str) -> str:
    value = row.get(key)
    if value is None:
        return ""
    return str(value).strip()


def _normalize(text: str | None) -> str:
    """Trim, collapse repeated internal whitespace, and casefold -- used
    ONLY to decide whether two rows refer to the same real-world department
    (the in-file duplicate check, and matching against existing DB rows).
    Comparison-only: never mutates what actually gets stored."""
    if not text:
        return ""
    return re.sub(r"\s+", " ", text.strip()).casefold()


def _composite_key(campus_code: str, department_code: str) -> tuple[str, str]:
    """The normalized 2-part identity a Department is matched/deduped by --
    Campus Code + Department Code only, deliberately narrower than a full-row
    match (Department Name is NOT part of the identity -- see this module's
    own docstring)."""
    return (_normalize(campus_code), _normalize(department_code))


def _parse_category(text: str) -> tuple[StaffRoleCategoryEnum | None, str | None]:
    if not text:
        return None, f"Missing required column '{CATEGORY_HEADER}'"
    upper = text.upper()
    if upper not in _CATEGORY_VALUES:
        return None, f"Unknown '{CATEGORY_HEADER}' value '{text}'. Valid values: {', '.join(_CATEGORY_VALUES)}"
    return StaffRoleCategoryEnum(upper), None


def _parse_active(text: str) -> tuple[bool | None, str | None]:
    """Blank -> default True, matching the New Department form's own
    "Active defaults to Yes". Otherwise case-insensitive TRUE/FALSE only --
    anything else is a clear rejection naming the bad value."""
    if not text:
        return True, None
    upper = text.upper()
    if upper == "TRUE":
        return True, None
    if upper == "FALSE":
        return False, None
    return None, f"Invalid '{ACTIVE_HEADER}' value '{text}' -- expected TRUE or FALSE (blank defaults to TRUE)"


def validate_rows(db: Session, raw_rows: list[dict]) -> ValidationResult:
    """Parses+validates every row against the current DB state **without
    writing anything** -- same read-only contract as the other 3 importers'
    `validate_rows`, called by both `/departments/bulk-upload/validate` and
    `/departments/bulk-upload/commit` (which re-validates the re-uploaded
    file before writing).
    """
    if len(raw_rows) > MAX_ROWS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"File has {len(raw_rows)} data rows, exceeding the {MAX_ROWS}-row limit",
        )

    results: list[ImportRowResult] = []
    seen_keys: dict[tuple[str, str], int] = {}
    # Existing Department rows for a given campus, fetched at most once per
    # campus (not once per row) -- same batching technique
    # location_import.py's own `existing_by_campus` cache uses.
    existing_by_campus: dict[uuid.UUID, list[Department]] = {}
    created = updated = unchanged = rejected = 0

    for row_number, raw_row in enumerate(raw_rows, start=2):
        campus_code = _cell(raw_row, CAMPUS_HEADER).upper()
        department_code = _cell(raw_row, CODE_HEADER)
        department_name = _cell(raw_row, NAME_HEADER)
        category_raw = _cell(raw_row, CATEGORY_HEADER)
        parent_group = _cell(raw_row, PARENT_GROUP_HEADER) or None
        description = _cell(raw_row, DESCRIPTION_HEADER) or None
        active_raw = _cell(raw_row, ACTIVE_HEADER)

        errors: list[str] = []

        if not campus_code:
            errors.append(f"Missing required column '{CAMPUS_HEADER}'")
        elif campus_code not in CAMPUS_CODES:
            errors.append(f"Unknown campus code '{campus_code}'")

        if not department_code:
            errors.append(f"Missing required column '{CODE_HEADER}'")

        if not department_name:
            errors.append(f"Missing required column '{NAME_HEADER}'")

        category, category_error = _parse_category(category_raw)
        if category_error:
            errors.append(category_error)

        is_active, active_error = _parse_active(active_raw)
        if active_error:
            errors.append(active_error)

        campus = None
        if campus_code and campus_code in CAMPUS_CODES:
            campus = db.query(Campus).filter(Campus.code == campus_code).one_or_none()
            if campus is None:
                errors.append(f"Campus '{campus_code}' is not seeded in this environment")

        if campus_code and department_code:
            key = _composite_key(campus_code, department_code)
            if key in seen_keys:
                errors.append(
                    f"Duplicate department: same Campus Code and Department Code "
                    f"(already used on row {seen_keys[key]})"
                )
            else:
                seen_keys[key] = row_number

        result = ImportRowResult(
            row_number=row_number,
            status="rejected",
            campus_code=campus_code or None,
            department_code=department_code or None,
            department_name=department_name or None,
            category=category,
            parent_group=parent_group,
            description=description,
            is_active=is_active,
        )

        if errors:
            result.error_reason = "; ".join(errors)
            results.append(result)
            rejected += 1
            continue

        result.campus_id = campus.id

        campus_departments = existing_by_campus.get(campus.id)
        if campus_departments is None:
            campus_departments = (
                db.query(Department).filter(Department.campus_id == campus.id).order_by(Department.created_at).all()
            )
            existing_by_campus[campus.id] = campus_departments

        row_key = _composite_key(campus_code, department_code)
        existing = next(
            (
                dept
                for dept in campus_departments
                if dept.code is not None and _composite_key(campus_code, dept.code) == row_key
            ),
            None,
        )

        if existing is None:
            result.status = "created"
            created += 1
        elif (
            existing.code != department_code
            or existing.name != department_name
            or existing.category != category
            or existing.parent_group != parent_group
            or existing.description != description
            or existing.is_active != is_active
        ):
            result.status = "updated"
            result.department_id = existing.id
            updated += 1
        else:
            result.status = "unchanged"
            result.department_id = existing.id
            unchanged += 1

        results.append(result)

    return ValidationResult(
        rows=results,
        total=len(raw_rows),
        created_count=created,
        updated_count=updated,
        unchanged_count=unchanged,
        rejected_count=rejected,
    )


def commit_rows(db: Session, *, validation: ValidationResult, bulk_upload_log_id: uuid.UUID) -> None:
    """Applies every non-rejected row's UPSERT within the *caller's* existing
    transaction -- no `db.commit()` here, same contract as the other 3
    importers' `commit_rows`. Also writes one `BulkUploadRowLog` row per
    non-rejected row (see this module's own docstring for why -- Department
    has no permanent old-value history table, so undo relies on this
    instead).

    `Department` carries no `created_by_id`/`updated_by_id` columns (unlike
    `HousekeepingStaff`/`SanctionedStrength`), so there is no actor to stamp
    on the row itself here -- same as `location_import.py`'s own commit.
    """
    for row in validation.rows:
        if row.status == "created":
            new_row = Department(
                campus_id=row.campus_id,
                code=row.department_code,
                name=row.department_name,
                category=row.category,
                parent_group=row.parent_group,
                description=row.description,
                is_active=row.is_active,
            )
            db.add(new_row)
            db.flush()
            db.add(
                BulkUploadRowLog(
                    bulk_upload_log_id=bulk_upload_log_id,
                    entity_type=BulkUploadEntityTypeEnum.DEPARTMENT,
                    entity_id=new_row.id,
                    was_created=True,
                )
            )
        elif row.status == "updated":
            existing = db.get(Department, row.department_id)
            existing.code = row.department_code
            existing.name = row.department_name
            existing.category = row.category
            existing.parent_group = row.parent_group
            existing.description = row.description
            existing.is_active = row.is_active
            db.add(
                BulkUploadRowLog(
                    bulk_upload_log_id=bulk_upload_log_id,
                    entity_type=BulkUploadEntityTypeEnum.DEPARTMENT,
                    entity_id=existing.id,
                    was_created=False,
                )
            )
        elif row.status == "unchanged":
            db.add(
                BulkUploadRowLog(
                    bulk_upload_log_id=bulk_upload_log_id,
                    entity_type=BulkUploadEntityTypeEnum.DEPARTMENT,
                    entity_id=row.department_id,
                    was_created=False,
                )
            )
        # "rejected": no write, no row log entry.


def _write_header(ws, headers: tuple[str, ...]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 26


def build_bulk_upload_template_xlsx(db: Session) -> bytes:
    """Live-generated, same styling shape as the other 3 importers' template
    builders: locked/frozen styled header row, highlighted example rows
    marked for deletion (harmless if left in -- see `_EXAMPLE_ROWS`'s
    docstring above), and a "Master Lists" reference sheet of currently-valid
    campus codes + category values so a user can copy exact values in.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Departments"
    _write_header(ws, TEMPLATE_HEADERS)
    for example in _EXAMPLE_ROWS:
        ws.append(list(example))
    for row_idx in range(2, 2 + len(_EXAMPLE_ROWS)):
        for col in range(1, len(TEMPLATE_HEADERS) + 1):
            ws.cell(row=row_idx, column=col).fill = _SAMPLE_FILL

    campus_formula = '"' + ",".join(CAMPUS_CODES) + '"'
    campus_dv = DataValidation(type="list", formula1=campus_formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(campus_dv)
    campus_dv.add("A2:A500")

    category_formula = '"' + ",".join(_CATEGORY_VALUES) + '"'
    category_dv = DataValidation(type="list", formula1=category_formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(category_dv)
    category_dv.add("D2:D500")

    active_formula = '"TRUE,FALSE"'
    active_dv = DataValidation(type="list", formula1=active_formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(active_dv)
    active_dv.add("G2:G500")

    master_ws = wb.create_sheet("Master Lists")
    master_ws.append(["Campus Code", "Category"])
    for cell in master_ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    max_rows = max(len(CAMPUS_CODES), len(_CATEGORY_VALUES))
    for row_idx in range(max_rows):
        if row_idx < len(CAMPUS_CODES):
            master_ws.cell(row=row_idx + 2, column=1, value=CAMPUS_CODES[row_idx])
        if row_idx < len(_CATEGORY_VALUES):
            master_ws.cell(row=row_idx + 2, column=2, value=_CATEGORY_VALUES[row_idx])
    for col in range(1, 3):
        master_ws.column_dimensions[get_column_letter(col)].width = 26

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_error_report_xlsx(log: BulkUploadLog, rejected_rows: list[ImportRowResult]) -> bytes:
    """Same shape as the other 3 importers' `build_error_report_xlsx` --
    only this upload's rejected rows, plus the `error_reason` column."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rejected rows"
    ws.append([f"Bulk upload: {log.filename}"])
    ws.append([f"Uploaded: {log.uploaded_at.isoformat()}"])
    ws.append([])
    headers = list(TEMPLATE_HEADERS) + ["error_reason"]
    ws.append(headers)
    for row in rejected_rows:
        ws.append(
            [
                row.campus_code,
                row.department_code,
                row.department_name,
                row.category.value if row.category else None,
                row.parent_group,
                row.description,
                "TRUE" if row.is_active else ("FALSE" if row.is_active is not None else None),
                row.error_reason,
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
