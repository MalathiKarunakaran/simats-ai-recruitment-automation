"""Sanctioned Strength bulk upload: validate -> preview -> commit
(zany-snuggling-pie.md Phase F). Genuinely new territory in this codebase --
no prior validate-then-commit two-step exists anywhere (`app/services/
tracker_import.py` is one-shot parse+persist in a single call). This module
deliberately reuses, rather than reinvents:

- `app/services/sanctioned_strength.py::current_effective_row` as the single
  UPSERT-target resolver (never reimplemented here) -- the same function the
  Vacancy Register rewrite and the availability strip use.
- `tracker_import.py`'s row-by-row parse/validate loop shape and
  "collect errors, decide status" pattern, adapted to this spec's
  created/updated/unchanged/rejected vocabulary (tracker_import.py uses a
  simpler imported/imported_with_warning/flagged model).
- The plan's deliberate "no server-side cache between validate and commit"
  simplicity choice: `/commit` re-parses+re-validates the *same re-uploaded*
  file bytes rather than trusting a client-held preview -- `validate_rows`
  is the single source of truth both endpoints call.

**Column-naming note**: the plan's prose describes the UPSERT key as
"(campus_code, department_code, designation_code)". This codebase's real
schema doesn't support that literally -- `Designation` (app/models/
designation.py) has no `code` column at all, only `name`; `Department.code`
(app/models/department.py) is optional/free-text and NOT unique (only
`(campus_id, name)` is). So this importer keys rows on (Campus Code,
Department Name, Designation Name) instead: `Campus.code` is the one real
unique code column of the three; Department is resolved by its own
`(campus_id, name)` uniqueness; Designation is resolved case-insensitively
by name, the same convention `tracker_import.py::_resolve_designation`
already established for matching free-text position titles against
Designation Master.
"""

import csv
import io
import uuid
from dataclasses import dataclass
from datetime import date, datetime

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from app.models.bulk_upload_log import BulkUploadLog
from app.models.campus import Campus
from app.models.department import Department
from app.models.designation import Designation
from app.models.enums import CAMPUS_CODES, SanctionedStrengthChangeSourceEnum, StaffRoleCategoryEnum
from app.models.sanctioned_strength import SanctionedStrength, SanctionedStrengthHistory
from app.models.user import User
from app.services.sanctioned_strength import current_effective_row

MAX_ROWS = 5000

CAMPUS_HEADER = "Campus Code"
DEPARTMENT_HEADER = "Department Name"
DESIGNATION_HEADER = "Designation Name"
STRENGTH_HEADER = "Approved Strength"
EFFECTIVE_FROM_HEADER = "Effective From (DD-MM-YYYY)"
REMARKS_HEADER = "Remarks"

TEMPLATE_HEADERS = (
    CAMPUS_HEADER,
    DEPARTMENT_HEADER,
    DESIGNATION_HEADER,
    STRENGTH_HEADER,
    EFFECTIVE_FROM_HEADER,
    REMARKS_HEADER,
)

# Example rows are deliberately built with an invalid campus code -- "XXX"
# is not in CAMPUS_CODES -- so that if a user forgets to delete them before
# uploading, they always come back `rejected` ("Unknown campus code 'XXX'")
# rather than silently being committed as real data. This is the mechanism
# behind the plan's "3 highlighted example rows marked for deletion": the
# highlight is a visual cue, this is the functional backstop.
_EXAMPLE_ROWS = (
    ("XXX", "EXAMPLE - DELETE THIS ROW", "Assistant Professor", 5, "01-04-2026", "Sample only"),
    ("XXX", "EXAMPLE - DELETE THIS ROW", "Lab Assistant", 2, "01-04-2026", ""),
    ("XXX", "EXAMPLE - DELETE THIS ROW", "Office Assistant", 1, "01-04-2026", ""),
)

_HEADER_FILL = PatternFill(start_color="1B5FAA", end_color="1B5FAA", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SAMPLE_FILL = PatternFill(start_color="FFF3D6", end_color="FFF3D6", fill_type="solid")


@dataclass
class ImportRowResult:
    row_number: int
    status: str  # created | updated | unchanged | rejected
    error_reason: str | None = None
    campus_code: str | None = None
    department_name: str | None = None
    designation_name: str | None = None
    approved_strength: int | None = None
    effective_from: date | None = None
    remarks: str | None = None
    # Resolved only for non-rejected rows -- commit_rows() uses these
    # directly rather than re-resolving lookups a second time.
    campus_id: uuid.UUID | None = None
    department_id: uuid.UUID | None = None
    designation_id: uuid.UUID | None = None
    category: StaffRoleCategoryEnum | None = None
    sanctioned_strength_id: uuid.UUID | None = None
    old_value: int | None = None


@dataclass
class ValidationResult:
    rows: list[ImportRowResult]
    total: int
    created_count: int
    updated_count: int
    unchanged_count: int
    rejected_count: int


def _cell(row: dict, key: str) -> str:
    value = row.get(key)
    if value is None:
        return ""
    return str(value).strip()


def _parse_rows_from_xlsx(file_bytes: bytes) -> list[dict]:
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = workbook.worksheets[0]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return []
    result = []
    for raw_row in rows_iter:
        if all(cell is None or str(cell).strip() == "" for cell in raw_row):
            continue
        result.append(dict(zip(headers, raw_row)))
    return result


def _parse_rows_from_csv(file_bytes: bytes) -> list[dict]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    result = []
    for raw_row in reader:
        if all(v is None or str(v).strip() == "" for v in raw_row.values()):
            continue
        result.append(raw_row)
    return result


def parse_rows(file_bytes: bytes, filename: str | None) -> list[dict]:
    lowered = (filename or "").lower()
    if lowered.endswith(".xlsx"):
        return _parse_rows_from_xlsx(file_bytes)
    if lowered.endswith(".csv"):
        return _parse_rows_from_csv(file_bytes)
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only .xlsx or .csv files are accepted")


def _parse_effective_from(raw) -> tuple[date | None, str | None]:
    if isinstance(raw, datetime):
        return raw.date(), None
    if isinstance(raw, date):
        return raw, None
    text = str(raw).strip() if raw is not None else ""
    if not text:
        return None, f"Missing required column '{EFFECTIVE_FROM_HEADER}'"
    try:
        return datetime.strptime(text, "%d-%m-%Y").date(), None
    except ValueError:
        return None, f"Invalid date '{text}' for '{EFFECTIVE_FROM_HEADER}' -- expected DD-MM-YYYY"


def validate_rows(db: Session, raw_rows: list[dict]) -> ValidationResult:
    """Parses+validates every row against the current DB state **without
    writing anything** (only `db.query()`/`SELECT`s below -- no `db.add()`,
    no `db.flush()` of new rows, no `db.commit()`). Called by both
    `/bulk-upload/validate` (read-only preview) and `/bulk-upload/commit`
    (defensive re-validation immediately before `commit_rows()` actually
    writes) so the two endpoints can never disagree on what's valid.
    """
    if len(raw_rows) > MAX_ROWS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"File has {len(raw_rows)} data rows, exceeding the {MAX_ROWS}-row limit",
        )

    results: list[ImportRowResult] = []
    seen_keys: dict[tuple[str, str, str], int] = {}
    created = updated = unchanged = rejected = 0

    for row_number, raw_row in enumerate(raw_rows, start=2):
        campus_code = _cell(raw_row, CAMPUS_HEADER).upper()
        department_name = _cell(raw_row, DEPARTMENT_HEADER)
        designation_name = _cell(raw_row, DESIGNATION_HEADER)
        strength_raw = _cell(raw_row, STRENGTH_HEADER)
        effective_from_raw = raw_row.get(EFFECTIVE_FROM_HEADER)
        remarks = _cell(raw_row, REMARKS_HEADER) or None

        errors: list[str] = []

        if not campus_code:
            errors.append(f"Missing required column '{CAMPUS_HEADER}'")
        elif campus_code not in CAMPUS_CODES:
            errors.append(f"Unknown campus code '{campus_code}'")

        if not department_name:
            errors.append(f"Missing required column '{DEPARTMENT_HEADER}'")

        if not designation_name:
            errors.append(f"Missing required column '{DESIGNATION_HEADER}'")

        approved_strength: int | None = None
        if not strength_raw:
            errors.append(f"Missing required column '{STRENGTH_HEADER}'")
        else:
            try:
                # int(), not int(float()) -- "5.5" must be rejected as a
                # data-entry mistake, not silently truncated to 5.
                approved_strength = int(strength_raw)
                if approved_strength < 0:
                    errors.append(f"'{STRENGTH_HEADER}' must not be negative")
            except ValueError:
                errors.append(f"'{STRENGTH_HEADER}' must be a whole number, got '{strength_raw}'")

        effective_from, date_error = _parse_effective_from(effective_from_raw)
        if date_error:
            errors.append(date_error)

        campus = None
        if campus_code and campus_code in CAMPUS_CODES:
            campus = db.query(Campus).filter(Campus.code == campus_code).one_or_none()
            if campus is None:
                errors.append(f"Campus '{campus_code}' is not seeded in this environment")

        department = None
        if campus is not None and department_name:
            department = (
                db.query(Department)
                .filter(Department.campus_id == campus.id, Department.name.ilike(department_name))
                .one_or_none()
            )
            if department is None:
                errors.append(f"Unknown department '{department_name}' for campus '{campus_code}'")

        designation = None
        if designation_name:
            designation = db.query(Designation).filter(Designation.name.ilike(designation_name)).one_or_none()
            if designation is None:
                errors.append(f"Unknown designation '{designation_name}'")

        if department is not None and designation is not None and designation.category != department.category:
            errors.append(
                f"Designation category ({designation.category.value}) does not match "
                f"department category ({department.category.value})"
            )

        if campus_code and department_name and designation_name:
            key = (campus_code, department_name.lower(), designation_name.lower())
            if key in seen_keys:
                errors.append(f"Duplicate key already used on row {seen_keys[key]}")
            else:
                seen_keys[key] = row_number

        result = ImportRowResult(
            row_number=row_number,
            status="rejected",
            campus_code=campus_code or None,
            department_name=department_name or None,
            designation_name=designation_name or None,
            approved_strength=approved_strength,
            effective_from=effective_from,
            remarks=remarks,
        )

        if errors:
            result.error_reason = "; ".join(errors)
            results.append(result)
            rejected += 1
            continue

        result.campus_id = campus.id
        result.department_id = department.id
        result.designation_id = designation.id
        result.category = designation.category

        existing = current_effective_row(
            db, campus_id=campus.id, department_id=department.id, designation_id=designation.id
        )
        if existing is None:
            result.status = "created"
            created += 1
        elif existing.approved_strength != approved_strength:
            result.status = "updated"
            result.old_value = existing.approved_strength
            result.sanctioned_strength_id = existing.id
            updated += 1
        else:
            result.status = "unchanged"
            result.sanctioned_strength_id = existing.id
            unchanged += 1

        results.append(result)

    return ValidationResult(
        rows=results,
        total=len(raw_rows),
        created_count=created,
        updated_count=updated,
        unchanged_count=unchanged,
        rejected_count=rejected,
    )


def commit_rows(
    db: Session, *, validation: ValidationResult, actor: User, bulk_upload_log_id: uuid.UUID
) -> None:
    """Applies every non-rejected row's UPSERT within the *caller's* existing
    transaction -- no `db.commit()` here. The router commits exactly once,
    after this returns, so a failure partway through (e.g. an unexpected
    IntegrityError -- defense in depth, since `validate_rows` should already
    have caught anything that would fail here) leaves nothing persisted once
    the router rolls back / the session closes without committing.

    `rejected` rows are skipped entirely (never written). `unchanged` rows
    are true no-ops -- no row mutation, no history row -- which is what makes
    re-uploading an identical file idempotent (zero new history rows the
    second time).
    """
    for row in validation.rows:
        if row.status == "created":
            new_row = SanctionedStrength(
                campus_id=row.campus_id,
                department_id=row.department_id,
                designation_id=row.designation_id,
                category=row.category,
                approved_strength=row.approved_strength,
                effective_from=row.effective_from,
                remarks=row.remarks,
                created_by_id=actor.id,
            )
            db.add(new_row)
            db.flush()
            db.add(
                SanctionedStrengthHistory(
                    sanctioned_strength_id=new_row.id,
                    old_value=None,
                    new_value=row.approved_strength,
                    changed_by_id=actor.id,
                    source=SanctionedStrengthChangeSourceEnum.BULK_UPLOAD,
                    bulk_upload_log_id=bulk_upload_log_id,
                )
            )
        elif row.status == "updated":
            existing = db.get(SanctionedStrength, row.sanctioned_strength_id)
            old_value = existing.approved_strength
            existing.approved_strength = row.approved_strength
            existing.effective_from = row.effective_from
            existing.remarks = row.remarks
            existing.updated_by_id = actor.id
            db.add(
                SanctionedStrengthHistory(
                    sanctioned_strength_id=existing.id,
                    old_value=old_value,
                    new_value=row.approved_strength,
                    changed_by_id=actor.id,
                    source=SanctionedStrengthChangeSourceEnum.BULK_UPLOAD,
                    bulk_upload_log_id=bulk_upload_log_id,
                )
            )
        # "unchanged"/"rejected": no write.


def _write_header(ws, headers: tuple[str, ...]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 26


def build_bulk_upload_template_xlsx(db: Session) -> bytes:
    """Live-generated (not a static file, unlike migration.py's tracker
    template) -- mirrors scripts/generate_tracker_template.py's styling:
    locked/frozen styled header row, 3 highlighted example rows marked for
    deletion (see `_EXAMPLE_ROWS`'s docstring above for how leaving them in
    is made harmless rather than just visually discouraged), and a second
    "Master Lists" reference sheet of currently-valid campus/department/
    designation names + categories so a user can copy exact values in.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sanctioned Strength"
    _write_header(ws, TEMPLATE_HEADERS)
    for example in _EXAMPLE_ROWS:
        ws.append(list(example))
    for row_idx in range(2, 2 + len(_EXAMPLE_ROWS)):
        for col in range(1, len(TEMPLATE_HEADERS) + 1):
            ws.cell(row=row_idx, column=col).fill = _SAMPLE_FILL

    # Campus Code is the one bounded-vocabulary column -- Department/
    # Designation names are open-ended free text (matching the Master Lists
    # sheet), same restraint scripts/generate_tracker_template.py already
    # applies (it doesn't dropdown-validate its own free-text "Department"
    # column either).
    formula = '"' + ",".join(CAMPUS_CODES) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"A2:A{500}")

    master_ws = wb.create_sheet("Master Lists")
    master_ws.append(["Campus Code", "Department Name (Campus)", "Designation Name (Category)"])
    for cell in master_ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    departments = (
        db.query(Department)
        .filter(Department.is_active.is_(True))
        .join(Campus, Department.campus_id == Campus.id)
        .order_by(Campus.code, Department.name)
        .all()
    )
    designations = db.query(Designation).filter(Designation.is_active.is_(True)).order_by(Designation.name).all()

    department_labels = [f"{d.name} ({d.campus.code})" for d in departments]
    designation_labels = [f"{d.name} ({d.category.value})" for d in designations]

    max_rows = max(len(CAMPUS_CODES), len(department_labels), len(designation_labels))
    for row_idx in range(max_rows):
        if row_idx < len(CAMPUS_CODES):
            master_ws.cell(row=row_idx + 2, column=1, value=CAMPUS_CODES[row_idx])
        if row_idx < len(department_labels):
            master_ws.cell(row=row_idx + 2, column=2, value=department_labels[row_idx])
        if row_idx < len(designation_labels):
            master_ws.cell(row=row_idx + 2, column=3, value=designation_labels[row_idx])
    for col in range(1, 4):
        master_ws.column_dimensions[get_column_letter(col)].width = 34

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_error_report_xlsx(log: BulkUploadLog, rejected_rows: list[ImportRowResult]) -> bytes:
    """Plain Workbook-append pattern, same shape as `app/services/
    exports.py::build_report_excel` -- only this upload's rejected rows,
    plus the `error_reason` column."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rejected rows"
    ws.append([f"Bulk upload: {log.filename}"])
    ws.append([f"Uploaded: {log.uploaded_at.isoformat()}"])
    ws.append([])
    headers = list(TEMPLATE_HEADERS) + ["error_reason"]
    ws.append(headers)
    for row in rejected_rows:
        ws.append(
            [
                row.campus_code,
                row.department_name,
                row.designation_name,
                row.approved_strength,
                row.effective_from.strftime("%d-%m-%Y") if row.effective_from else None,
                row.remarks,
                row.error_reason,
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
