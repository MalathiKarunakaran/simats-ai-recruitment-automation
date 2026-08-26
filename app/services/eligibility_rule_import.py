"""EligibilityRule bulk upload: validate -> preview -> commit (starter
regulatory-eligibility-rules feature, backend Phase 1). 5th sibling module to
`sanctioned_strength_import.py`/`location_import.py`/
`housekeeping_staff_import.py`/`department_import.py` -- same dataclass/
validate/commit/undo-adjacent shape, deliberately NOT a shared abstraction
between entities (see `location_import.py`'s own module docstring for the
standing rationale, not repeated per module). `parse_rows` (file bytes ->
`list[dict]`, entity-agnostic) is reused as-is from
`sanctioned_strength_import`, same as the other 4 siblings.

**UPSERT key**: `(Campus Code, Department Code, Position Title, Regulatory
Authority, Effective From)`, normalized (trim + casefold for the 3 text
parts, same `_normalize` technique as `location_import.py`'s own helper;
Regulatory Authority compared by its raw enum value; Effective From compared
by ISO date, with a blank/None cell normalized to the same empty-string key
component as the other optional parts) -- the same 5-field natural key
`app/api/v1/routers/eligibility_rules.py`'s own application-level uniqueness
check uses for manual create/update, so a bulk upload and a manual create can
never quietly diverge on what counts as "the same rule".

**No DB-level uniqueness on this 5-field key**: same reasoning as
Department's own Code+Campus story (see `department_import.py`'s own
docstring) -- 4 of the 5 key columns are nullable, and Postgres treats every
NULL as distinct in a unique index, which would defeat a real DB constraint's
guarantee. Existing-row resolution below therefore fetches every
EligibilityRule for a given campus once (not once per row) and matches in
Python by the normalized composite key, same defensive pattern
`department_import.py`/`location_import.py` already use for their own
not-yet-constrained composite keys.

**Undo scope**: same as `department_import.py`'s own -- EligibilityRule has
no permanent old-value history table, so `commit_rows` writes one
`BulkUploadRowLog` row per non-rejected row (created/updated/unchanged
alike), which `undo_bulk_upload` (app/api/v1/routers/sanctioned_strength.py)
uses to soft-delete only the rows this batch actually created.
"""

import io
import re
import uuid
from dataclasses import dataclass
from datetime import date

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from app.models.bulk_upload_log import BulkUploadLog
from app.models.bulk_upload_row_log import BulkUploadRowLog
from app.models.campus import Campus
from app.models.department import Department
from app.models.eligibility_rule import EligibilityRule
from app.models.enums import (
    CAMPUS_CODES,
    BulkUploadEntityTypeEnum,
    EligibilityRuleStatusEnum,
    RegulatoryAuthorityEnum,
    StaffRoleCategoryEnum,
)
from app.services.sanctioned_strength_import import parse_rows  # noqa: F401 -- re-exported, see module docstring

MAX_ROWS = 5000

CAMPUS_HEADER = "Campus Code"
DEPARTMENT_HEADER = "Department Code"
STAFF_CATEGORY_HEADER = "Staff Category"
POSITION_TITLE_HEADER = "Position Title"
REQUIRED_QUALIFICATION_KEYWORD_HEADER = "Required Qualification Keyword"
NET_SET_REQUIRED_HEADER = "NET/SET/SLET Required"
SUBJECT_HEADER = "Subject"
SKILLS_KEYWORD_HEADER = "Skills Keyword"
ID_PROOF_REQUIRED_HEADER = "ID Proof Required"
SHIFT_PREFERENCE_HEADER = "Shift Preference"
REGULATORY_AUTHORITY_HEADER = "Regulatory Authority"
SCHOOL_OR_COLLEGE_HEADER = "School/College"
PROGRAMME_DISCIPLINE_HEADER = "Programme/Discipline"
MINIMUM_QUALIFICATION_HEADER = "Minimum Qualification"
MINIMUM_PERCENTAGE_HEADER = "Minimum Percentage"
REQUIRED_EXPERIENCE_HEADER = "Required Experience"
REQUIRED_CREDENTIAL_HEADER = "Required Credential"
REQUIRED_KEYWORDS_HEADER = "Required Keywords (informational only)"
PREFERRED_KEYWORDS_HEADER = "Preferred Keywords (informational only)"
PHD_REQUIRED_HEADER = "PhD Required"
PROFESSIONAL_REGISTRATION_HEADER = "Professional Registration"
INDUSTRY_EXPERIENCE_HEADER = "Industry Experience"
PRIORITY_HEADER = "Priority"
EFFECTIVE_FROM_HEADER = "Effective From (YYYY-MM-DD)"
EFFECTIVE_TO_HEADER = "Effective To (YYYY-MM-DD)"
SOURCE_REGULATION_HEADER = "Source Regulation"
STATUS_HEADER = "Status"
VERIFICATION_REQUIRED_HEADER = "Verification Required"
ACTIVE_HEADER = "Active"
NOTES_HEADER = "Notes"

TEMPLATE_HEADERS = (
    CAMPUS_HEADER,
    DEPARTMENT_HEADER,
    STAFF_CATEGORY_HEADER,
    POSITION_TITLE_HEADER,
    REQUIRED_QUALIFICATION_KEYWORD_HEADER,
    NET_SET_REQUIRED_HEADER,
    SUBJECT_HEADER,
    SKILLS_KEYWORD_HEADER,
    ID_PROOF_REQUIRED_HEADER,
    SHIFT_PREFERENCE_HEADER,
    REGULATORY_AUTHORITY_HEADER,
    SCHOOL_OR_COLLEGE_HEADER,
    PROGRAMME_DISCIPLINE_HEADER,
    MINIMUM_QUALIFICATION_HEADER,
    MINIMUM_PERCENTAGE_HEADER,
    REQUIRED_EXPERIENCE_HEADER,
    REQUIRED_CREDENTIAL_HEADER,
    REQUIRED_KEYWORDS_HEADER,
    PREFERRED_KEYWORDS_HEADER,
    PHD_REQUIRED_HEADER,
    PROFESSIONAL_REGISTRATION_HEADER,
    INDUSTRY_EXPERIENCE_HEADER,
    PRIORITY_HEADER,
    EFFECTIVE_FROM_HEADER,
    EFFECTIVE_TO_HEADER,
    SOURCE_REGULATION_HEADER,
    STATUS_HEADER,
    VERIFICATION_REQUIRED_HEADER,
    ACTIVE_HEADER,
    NOTES_HEADER,
)

# Same functional backstop as the other 4 importers' own example rows --
# "XXX" is never a real campus code, so a forgotten example row always comes
# back rejected rather than silently importing as real data.
_EXAMPLE_ROWS = (
    (
        "XXX", "CSE", "TEACHING", "Assistant Professor", "PHD", "TRUE", "Computer Science", "", "", "",
        "AICTE_UGC", "EXAMPLE - DELETE THIS ROW", "B.Tech Computer Science and Engineering",
        "PhD in relevant discipline as per AICTE/UGC norms", "55% or equivalent CGPA of 6.25/10, relaxable for SC/ST/PWD",
        "", "NET/SET/SLET or PhD as per UGC 2018 Regulations", "", "", "TRUE", "", "", "", "", "",
        "AICTE + applicable UGC rules -- starter regulatory mapping, verify before activation", "DRAFT", "TRUE",
        "TRUE", "EXAMPLE - DELETE THIS ROW",
    ),
)

_HEADER_FILL = PatternFill(start_color="1B5FAA", end_color="1B5FAA", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SAMPLE_FILL = PatternFill(start_color="FFF3D6", end_color="FFF3D6", fill_type="solid")

_CATEGORY_VALUES = tuple(c.value for c in StaffRoleCategoryEnum)
_AUTHORITY_VALUES = tuple(a.value for a in RegulatoryAuthorityEnum)
_STATUS_VALUES = tuple(s.value for s in EligibilityRuleStatusEnum)


@dataclass
class ImportRowResult:
    row_number: int
    status: str  # created | updated | unchanged | rejected
    error_reason: str | None = None
    campus_code: str | None = None
    department_code: str | None = None
    staff_category: StaffRoleCategoryEnum | None = None
    position_title: str | None = None
    required_qualification_keyword: str | None = None
    net_set_required: bool | None = None
    subject: str | None = None
    skills_keyword: str | None = None
    id_proof_required: bool | None = None
    shift_preference: str | None = None
    regulatory_authority: RegulatoryAuthorityEnum | None = None
    school_or_college: str | None = None
    programme_discipline: str | None = None
    minimum_qualification: str | None = None
    minimum_percentage: str | None = None
    required_experience: str | None = None
    required_credential: str | None = None
    required_keywords: str | None = None
    preferred_keywords: str | None = None
    phd_required: bool | None = None
    professional_registration: str | None = None
    industry_experience: str | None = None
    priority: str | None = None
    effective_from: date | None = None
    effective_to: date | None = None
    source_regulation: str | None = None
    # Named rule_status (not `status`) to avoid colliding with this same
    # dataclass's own created/updated/unchanged/rejected `status` field.
    rule_status: EligibilityRuleStatusEnum | None = None
    verification_required: bool | None = None
    is_active: bool | None = None
    notes: str | None = None
    # Resolved only for non-rejected rows -- commit_rows() uses these
    # directly rather than re-resolving lookups a second time.
    campus_id: uuid.UUID | None = None
    department_id: uuid.UUID | None = None
    eligibility_rule_id: uuid.UUID | None = None  # the existing match's id, when status != "created"


@dataclass
class ValidationResult:
    rows: list[ImportRowResult]
    total: int
    created_count: int
    updated_count: int
    unchanged_count: int
    rejected_count: int


def _cell(row: dict, key: str) -> str:
    value = row.get(key)
    if value is None:
        return ""
    return str(value).strip()


def _normalize(text: str | None) -> str:
    """Trim, collapse repeated internal whitespace, and casefold -- used
    ONLY to decide whether two rows refer to the same real-world eligibility
    rule (the in-file duplicate check, and matching against existing DB
    rows). Comparison-only: never mutates what actually gets stored."""
    if not text:
        return ""
    return re.sub(r"\s+", " ", text.strip()).casefold()


def _composite_key(
    campus_code: str,
    department_code: str | None,
    position_title: str | None,
    regulatory_authority: RegulatoryAuthorityEnum | None,
    effective_from: date | None,
) -> tuple[str, str, str, str, str]:
    """The normalized 5-part identity an EligibilityRule is matched/deduped
    by -- same natural key as
    app/api/v1/routers/eligibility_rules.py::_check_uniqueness_conflict's own
    application-level uniqueness check (campus, department, position title,
    regulatory authority, effective_from)."""
    return (
        _normalize(campus_code),
        _normalize(department_code),
        _normalize(position_title),
        regulatory_authority.value if regulatory_authority else "",
        effective_from.isoformat() if effective_from else "",
    )


def _parse_category(text: str) -> tuple[StaffRoleCategoryEnum | None, str | None]:
    if not text:
        return None, f"Missing required column '{STAFF_CATEGORY_HEADER}'"
    upper = text.upper()
    if upper not in _CATEGORY_VALUES:
        return None, f"Unknown '{STAFF_CATEGORY_HEADER}' value '{text}'. Valid values: {', '.join(_CATEGORY_VALUES)}"
    return StaffRoleCategoryEnum(upper), None


def _parse_authority(text: str) -> tuple[RegulatoryAuthorityEnum | None, str | None]:
    if not text:
        return None, None
    upper = text.upper()
    if upper not in _AUTHORITY_VALUES:
        return None, f"Unknown '{REGULATORY_AUTHORITY_HEADER}' value '{text}'. Valid values: {', '.join(_AUTHORITY_VALUES)}"
    return RegulatoryAuthorityEnum(upper), None


def _parse_rule_status(text: str) -> tuple[EligibilityRuleStatusEnum, str | None]:
    """Blank -> DRAFT, matching the model's own server-side default."""
    if not text:
        return EligibilityRuleStatusEnum.DRAFT, None
    upper = text.upper()
    if upper not in _STATUS_VALUES:
        return EligibilityRuleStatusEnum.DRAFT, f"Unknown '{STATUS_HEADER}' value '{text}'. Valid values: {', '.join(_STATUS_VALUES)}"
    return EligibilityRuleStatusEnum(upper), None


def _parse_bool_default_true(text: str, header: str) -> tuple[bool | None, str | None]:
    """Blank -> default True. Otherwise case-insensitive TRUE/FALSE only."""
    if not text:
        return True, None
    upper = text.upper()
    if upper == "TRUE":
        return True, None
    if upper == "FALSE":
        return False, None
    return None, f"Invalid '{header}' value '{text}' -- expected TRUE or FALSE (blank defaults to TRUE)"


def _parse_tristate_bool(text: str, header: str) -> tuple[bool | None, str | None]:
    """Blank -> None (genuinely unset, not a default) -- for the category-
    specific optional flags (net_set_required/id_proof_required/
    phd_required), which are nullable and mean something different from
    False when absent."""
    if not text:
        return None, None
    upper = text.upper()
    if upper == "TRUE":
        return True, None
    if upper == "FALSE":
        return False, None
    return None, f"Invalid '{header}' value '{text}' -- expected TRUE or FALSE (blank leaves it unset)"


def _parse_date(text: str, header: str) -> tuple[date | None, str | None]:
    if not text:
        return None, None
    try:
        return date.fromisoformat(text), None
    except ValueError:
        return None, f"Invalid '{header}' value '{text}' -- expected YYYY-MM-DD"


def validate_rows(db: Session, raw_rows: list[dict]) -> ValidationResult:
    """Parses+validates every row against the current DB state **without
    writing anything** -- same read-only contract as the other 4 importers'
    `validate_rows`, called by both `/eligibility-rules/bulk-upload/validate`
    and `/eligibility-rules/bulk-upload/commit` (which re-validates the
    re-uploaded file before writing).
    """
    if len(raw_rows) > MAX_ROWS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"File has {len(raw_rows)} data rows, exceeding the {MAX_ROWS}-row limit",
        )

    results: list[ImportRowResult] = []
    seen_keys: dict[tuple[str, str, str, str, str], int] = {}
    existing_by_campus: dict[uuid.UUID, list[EligibilityRule]] = {}
    created = updated = unchanged = rejected = 0

    for row_number, raw_row in enumerate(raw_rows, start=2):
        campus_code = _cell(raw_row, CAMPUS_HEADER).upper()
        department_code = _cell(raw_row, DEPARTMENT_HEADER) or None
        staff_category_raw = _cell(raw_row, STAFF_CATEGORY_HEADER)
        position_title = _cell(raw_row, POSITION_TITLE_HEADER) or None
        required_qualification_keyword = _cell(raw_row, REQUIRED_QUALIFICATION_KEYWORD_HEADER)
        subject = _cell(raw_row, SUBJECT_HEADER) or None
        skills_keyword = _cell(raw_row, SKILLS_KEYWORD_HEADER) or None
        shift_preference = _cell(raw_row, SHIFT_PREFERENCE_HEADER) or None
        school_or_college = _cell(raw_row, SCHOOL_OR_COLLEGE_HEADER) or None
        programme_discipline = _cell(raw_row, PROGRAMME_DISCIPLINE_HEADER) or None
        minimum_qualification = _cell(raw_row, MINIMUM_QUALIFICATION_HEADER) or None
        minimum_percentage = _cell(raw_row, MINIMUM_PERCENTAGE_HEADER) or None
        required_experience = _cell(raw_row, REQUIRED_EXPERIENCE_HEADER) or None
        required_credential = _cell(raw_row, REQUIRED_CREDENTIAL_HEADER) or None
        required_keywords = _cell(raw_row, REQUIRED_KEYWORDS_HEADER) or None
        preferred_keywords = _cell(raw_row, PREFERRED_KEYWORDS_HEADER) or None
        professional_registration = _cell(raw_row, PROFESSIONAL_REGISTRATION_HEADER) or None
        industry_experience = _cell(raw_row, INDUSTRY_EXPERIENCE_HEADER) or None
        priority = _cell(raw_row, PRIORITY_HEADER) or None
        source_regulation = _cell(raw_row, SOURCE_REGULATION_HEADER) or None
        notes = _cell(raw_row, NOTES_HEADER) or None

        errors: list[str] = []

        if not campus_code:
            errors.append(f"Missing required column '{CAMPUS_HEADER}'")
        elif campus_code not in CAMPUS_CODES:
            errors.append(f"Unknown campus code '{campus_code}'")

        staff_category, category_error = _parse_category(staff_category_raw)
        if category_error:
            errors.append(category_error)

        if not required_qualification_keyword:
            errors.append(f"Missing required column '{REQUIRED_QUALIFICATION_KEYWORD_HEADER}'")

        net_set_required, net_set_error = _parse_tristate_bool(_cell(raw_row, NET_SET_REQUIRED_HEADER), NET_SET_REQUIRED_HEADER)
        if net_set_error:
            errors.append(net_set_error)

        id_proof_required, id_proof_error = _parse_tristate_bool(_cell(raw_row, ID_PROOF_REQUIRED_HEADER), ID_PROOF_REQUIRED_HEADER)
        if id_proof_error:
            errors.append(id_proof_error)

        phd_required, phd_error = _parse_tristate_bool(_cell(raw_row, PHD_REQUIRED_HEADER), PHD_REQUIRED_HEADER)
        if phd_error:
            errors.append(phd_error)

        regulatory_authority, authority_error = _parse_authority(_cell(raw_row, REGULATORY_AUTHORITY_HEADER))
        if authority_error:
            errors.append(authority_error)

        rule_status, status_error = _parse_rule_status(_cell(raw_row, STATUS_HEADER))
        if status_error:
            errors.append(status_error)

        verification_required, verification_error = _parse_bool_default_true(
            _cell(raw_row, VERIFICATION_REQUIRED_HEADER), VERIFICATION_REQUIRED_HEADER
        )
        if verification_error:
            errors.append(verification_error)

        is_active, active_error = _parse_bool_default_true(_cell(raw_row, ACTIVE_HEADER), ACTIVE_HEADER)
        if active_error:
            errors.append(active_error)

        effective_from, effective_from_error = _parse_date(_cell(raw_row, EFFECTIVE_FROM_HEADER), EFFECTIVE_FROM_HEADER)
        if effective_from_error:
            errors.append(effective_from_error)

        effective_to, effective_to_error = _parse_date(_cell(raw_row, EFFECTIVE_TO_HEADER), EFFECTIVE_TO_HEADER)
        if effective_to_error:
            errors.append(effective_to_error)

        campus = None
        if campus_code and campus_code in CAMPUS_CODES:
            campus = db.query(Campus).filter(Campus.code == campus_code).one_or_none()
            if campus is None:
                errors.append(f"Campus '{campus_code}' is not seeded in this environment")

        department = None
        if campus is not None and department_code:
            campus_departments = (
                db.query(Department)
                .filter(Department.campus_id == campus.id, Department.code.isnot(None))
                .all()
            )
            department = next(
                (dept for dept in campus_departments if _normalize(dept.code) == _normalize(department_code)), None
            )
            if department is None:
                errors.append(f"Department code '{department_code}' is not known for campus '{campus_code}'")

        if campus_code and not errors:
            key = _composite_key(campus_code, department_code, position_title, regulatory_authority, effective_from)
            if key in seen_keys:
                errors.append(
                    "Duplicate eligibility rule: same Campus + Department + Position Title + Regulatory "
                    f"Authority + Effective From (already used on row {seen_keys[key]})"
                )
            else:
                seen_keys[key] = row_number

        result = ImportRowResult(
            row_number=row_number,
            status="rejected",
            campus_code=campus_code or None,
            department_code=department_code,
            staff_category=staff_category,
            position_title=position_title,
            required_qualification_keyword=required_qualification_keyword or None,
            net_set_required=net_set_required,
            subject=subject,
            skills_keyword=skills_keyword,
            id_proof_required=id_proof_required,
            shift_preference=shift_preference,
            regulatory_authority=regulatory_authority,
            school_or_college=school_or_college,
            programme_discipline=programme_discipline,
            minimum_qualification=minimum_qualification,
            minimum_percentage=minimum_percentage,
            required_experience=required_experience,
            required_credential=required_credential,
            required_keywords=required_keywords,
            preferred_keywords=preferred_keywords,
            phd_required=phd_required,
            professional_registration=professional_registration,
            industry_experience=industry_experience,
            priority=priority,
            effective_from=effective_from,
            effective_to=effective_to,
            source_regulation=source_regulation,
            rule_status=rule_status,
            verification_required=verification_required,
            is_active=is_active,
            notes=notes,
        )

        if errors:
            result.error_reason = "; ".join(errors)
            results.append(result)
            rejected += 1
            continue

        result.campus_id = campus.id
        result.department_id = department.id if department is not None else None

        campus_rules = existing_by_campus.get(campus.id)
        if campus_rules is None:
            campus_rules = (
                db.query(EligibilityRule)
                .filter(EligibilityRule.campus_id == campus.id)
                .order_by(EligibilityRule.created_at)
                .all()
            )
            existing_by_campus[campus.id] = campus_rules

        row_key = _composite_key(campus_code, department_code, position_title, regulatory_authority, effective_from)

        def _existing_key(rule: EligibilityRule) -> tuple[str, str, str, str, str]:
            dept_code = rule.department.code if rule.department is not None else None
            return _composite_key(campus_code, dept_code, rule.position_title, rule.regulatory_authority, rule.effective_from)

        existing = next((rule for rule in campus_rules if _existing_key(rule) == row_key), None)

        if existing is None:
            result.status = "created"
            created += 1
        elif (
            existing.department_id != result.department_id
            or existing.staff_category != staff_category
            or existing.position_title != position_title
            or existing.required_qualification_keyword != required_qualification_keyword
            or existing.net_set_required != net_set_required
            or existing.subject != subject
            or existing.skills_keyword != skills_keyword
            or existing.id_proof_required != id_proof_required
            or existing.shift_preference != shift_preference
            or existing.regulatory_authority != regulatory_authority
            or existing.school_or_college != school_or_college
            or existing.programme_discipline != programme_discipline
            or existing.minimum_qualification != minimum_qualification
            or existing.minimum_percentage != minimum_percentage
            or existing.required_experience != required_experience
            or existing.required_credential != required_credential
            or existing.required_keywords != required_keywords
            or existing.preferred_keywords != preferred_keywords
            or existing.phd_required != phd_required
            or existing.professional_registration != professional_registration
            or existing.industry_experience != industry_experience
            or existing.priority != priority
            or existing.effective_from != effective_from
            or existing.effective_to != effective_to
            or existing.source_regulation != source_regulation
            or existing.status != rule_status
            or existing.verification_required != verification_required
            or existing.is_active != is_active
            or existing.notes != notes
        ):
            result.status = "updated"
            result.eligibility_rule_id = existing.id
            updated += 1
        else:
            result.status = "unchanged"
            result.eligibility_rule_id = existing.id
            unchanged += 1

        results.append(result)

    return ValidationResult(
        rows=results,
        total=len(raw_rows),
        created_count=created,
        updated_count=updated,
        unchanged_count=unchanged,
        rejected_count=rejected,
    )


def commit_rows(db: Session, *, validation: ValidationResult, bulk_upload_log_id: uuid.UUID) -> None:
    """Applies every non-rejected row's UPSERT within the *caller's* existing
    transaction -- no `db.commit()` here, same contract as the other 4
    importers' `commit_rows`. Also writes one `BulkUploadRowLog` row per
    non-rejected row (see this module's own docstring for why -- EligibilityRule
    has no permanent old-value history table, so undo relies on this
    instead).
    """
    for row in validation.rows:
        if row.status == "created":
            new_row = EligibilityRule(
                campus_id=row.campus_id,
                department_id=row.department_id,
                staff_category=row.staff_category,
                position_title=row.position_title,
                required_qualification_keyword=row.required_qualification_keyword,
                net_set_required=row.net_set_required,
                subject=row.subject,
                skills_keyword=row.skills_keyword,
                id_proof_required=row.id_proof_required,
                shift_preference=row.shift_preference,
                regulatory_authority=row.regulatory_authority,
                school_or_college=row.school_or_college,
                programme_discipline=row.programme_discipline,
                minimum_qualification=row.minimum_qualification,
                minimum_percentage=row.minimum_percentage,
                required_experience=row.required_experience,
                required_credential=row.required_credential,
                required_keywords=row.required_keywords,
                preferred_keywords=row.preferred_keywords,
                phd_required=row.phd_required,
                professional_registration=row.professional_registration,
                industry_experience=row.industry_experience,
                priority=row.priority,
                effective_from=row.effective_from,
                effective_to=row.effective_to,
                source_regulation=row.source_regulation,
                status=row.rule_status,
                verification_required=row.verification_required,
                is_active=row.is_active,
                notes=row.notes,
            )
            db.add(new_row)
            db.flush()
            db.add(
                BulkUploadRowLog(
                    bulk_upload_log_id=bulk_upload_log_id,
                    entity_type=BulkUploadEntityTypeEnum.ELIGIBILITY_RULE,
                    entity_id=new_row.id,
                    was_created=True,
                )
            )
        elif row.status == "updated":
            existing = db.get(EligibilityRule, row.eligibility_rule_id)
            existing.department_id = row.department_id
            existing.staff_category = row.staff_category
            existing.position_title = row.position_title
            existing.required_qualification_keyword = row.required_qualification_keyword
            existing.net_set_required = row.net_set_required
            existing.subject = row.subject
            existing.skills_keyword = row.skills_keyword
            existing.id_proof_required = row.id_proof_required
            existing.shift_preference = row.shift_preference
            existing.regulatory_authority = row.regulatory_authority
            existing.school_or_college = row.school_or_college
            existing.programme_discipline = row.programme_discipline
            existing.minimum_qualification = row.minimum_qualification
            existing.minimum_percentage = row.minimum_percentage
            existing.required_experience = row.required_experience
            existing.required_credential = row.required_credential
            existing.required_keywords = row.required_keywords
            existing.preferred_keywords = row.preferred_keywords
            existing.phd_required = row.phd_required
            existing.professional_registration = row.professional_registration
            existing.industry_experience = row.industry_experience
            existing.priority = row.priority
            existing.effective_from = row.effective_from
            existing.effective_to = row.effective_to
            existing.source_regulation = row.source_regulation
            existing.status = row.rule_status
            existing.verification_required = row.verification_required
            existing.is_active = row.is_active
            existing.notes = row.notes
            db.add(
                BulkUploadRowLog(
                    bulk_upload_log_id=bulk_upload_log_id,
                    entity_type=BulkUploadEntityTypeEnum.ELIGIBILITY_RULE,
                    entity_id=existing.id,
                    was_created=False,
                )
            )
        elif row.status == "unchanged":
            db.add(
                BulkUploadRowLog(
                    bulk_upload_log_id=bulk_upload_log_id,
                    entity_type=BulkUploadEntityTypeEnum.ELIGIBILITY_RULE,
                    entity_id=row.eligibility_rule_id,
                    was_created=False,
                )
            )
        # "rejected": no write, no row log entry.


def _write_header(ws, headers: tuple[str, ...]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 24


def build_bulk_upload_template_xlsx(db: Session) -> bytes:
    """Live-generated, same styling shape as the other 4 importers' template
    builders: locked/frozen styled header row, a highlighted example row
    marked for deletion (harmless if left in -- see `_EXAMPLE_ROWS`'s
    docstring above), and a "Master Lists" reference sheet of currently-valid
    campus codes/category/authority/status values so a user can copy exact
    values in.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Eligibility Rules"
    _write_header(ws, TEMPLATE_HEADERS)
    for example in _EXAMPLE_ROWS:
        ws.append(list(example))
    for row_idx in range(2, 2 + len(_EXAMPLE_ROWS)):
        for col in range(1, len(TEMPLATE_HEADERS) + 1):
            ws.cell(row=row_idx, column=col).fill = _SAMPLE_FILL

    header_index = {header: idx for idx, header in enumerate(TEMPLATE_HEADERS, start=1)}

    def _add_list_validation(header: str, values: tuple[str, ...]):
        col_letter = get_column_letter(header_index[header])
        formula = '"' + ",".join(values) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}500")

    _add_list_validation(CAMPUS_HEADER, CAMPUS_CODES)
    _add_list_validation(STAFF_CATEGORY_HEADER, _CATEGORY_VALUES)
    _add_list_validation(REGULATORY_AUTHORITY_HEADER, _AUTHORITY_VALUES)
    _add_list_validation(STATUS_HEADER, _STATUS_VALUES)
    for boolean_header in (
        NET_SET_REQUIRED_HEADER,
        ID_PROOF_REQUIRED_HEADER,
        PHD_REQUIRED_HEADER,
        VERIFICATION_REQUIRED_HEADER,
        ACTIVE_HEADER,
    ):
        _add_list_validation(boolean_header, ("TRUE", "FALSE"))

    master_ws = wb.create_sheet("Master Lists")
    master_ws.append(["Campus Code", "Staff Category", "Regulatory Authority", "Status"])
    for cell in master_ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    max_rows = max(len(CAMPUS_CODES), len(_CATEGORY_VALUES), len(_AUTHORITY_VALUES), len(_STATUS_VALUES))
    for row_idx in range(max_rows):
        if row_idx < len(CAMPUS_CODES):
            master_ws.cell(row=row_idx + 2, column=1, value=CAMPUS_CODES[row_idx])
        if row_idx < len(_CATEGORY_VALUES):
            master_ws.cell(row=row_idx + 2, column=2, value=_CATEGORY_VALUES[row_idx])
        if row_idx < len(_AUTHORITY_VALUES):
            master_ws.cell(row=row_idx + 2, column=3, value=_AUTHORITY_VALUES[row_idx])
        if row_idx < len(_STATUS_VALUES):
            master_ws.cell(row=row_idx + 2, column=4, value=_STATUS_VALUES[row_idx])
    for col in range(1, 5):
        master_ws.column_dimensions[get_column_letter(col)].width = 26

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_error_report_xlsx(log: BulkUploadLog, rejected_rows: list[ImportRowResult]) -> bytes:
    """Same shape as the other 4 importers' `build_error_report_xlsx` --
    only this upload's rejected rows, plus the `error_reason` column."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rejected rows"
    ws.append([f"Bulk upload: {log.filename}"])
    ws.append([f"Uploaded: {log.uploaded_at.isoformat()}"])
    ws.append([])
    headers = list(TEMPLATE_HEADERS) + ["error_reason"]
    ws.append(headers)
    for row in rejected_rows:
        ws.append(
            [
                row.campus_code,
                row.department_code,
                row.staff_category.value if row.staff_category else None,
                row.position_title,
                row.required_qualification_keyword,
                _bool_cell(row.net_set_required),
                row.subject,
                row.skills_keyword,
                _bool_cell(row.id_proof_required),
                row.shift_preference,
                row.regulatory_authority.value if row.regulatory_authority else None,
                row.school_or_college,
                row.programme_discipline,
                row.minimum_qualification,
                row.minimum_percentage,
                row.required_experience,
                row.required_credential,
                row.required_keywords,
                row.preferred_keywords,
                _bool_cell(row.phd_required),
                row.professional_registration,
                row.industry_experience,
                row.priority,
                row.effective_from.isoformat() if row.effective_from else None,
                row.effective_to.isoformat() if row.effective_to else None,
                row.source_regulation,
                row.rule_status.value if row.rule_status else None,
                "TRUE" if row.verification_required else ("FALSE" if row.verification_required is not None else None),
                "TRUE" if row.is_active else ("FALSE" if row.is_active is not None else None),
                row.notes,
                row.error_reason,
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _bool_cell(value: bool | None) -> str | None:
    if value is None:
        return None
    return "TRUE" if value else "FALSE"
