"""Designation bulk upload: validate -> preview -> commit (Designation Master
bulk-upload epic, backend Phase 1). 6th sibling module to
`sanctioned_strength_import.py`/`location_import.py`/`housekeeping_staff_import.py`/
`department_import.py`/`eligibility_rule_import.py` -- same dataclass/
validate/commit/undo-adjacent shape, deliberately NOT a shared abstraction
between entities (see `location_import.py`'s own module docstring for the
standing rationale, not repeated per module). `parse_rows` (file bytes ->
`list[dict]`, entity-agnostic) is reused as-is from
`sanctioned_strength_import`, same as every other sibling.

**Natural key: (Designation Name, Category) -- NOT "Name + Category +
Department"**. The plan's prose describes the duplicate key as "Designation
Name + Category + Department", but `Designation` (app/models/designation.py)
is genuinely many-to-many with `Department` via `designation_departments` --
one `Designation` row can, and by design does, span multiple departments
simultaneously (e.g. "Assistant Professor" across several teaching
departments). There is no per-department "Designation row" to key against,
so a literal "+ Department" key doesn't map onto this schema at all. The
faithful interpretation, matching how a Designation is actually identified
today (its own `id`, addressed in practice by name+category): a row's
"Department Codes" column is not part of the identity key, it's a
*property* of the identified designation -- the full set of departments that
designation should be linked to after this row is applied.

**Update semantics: the uploaded row's department list REPLACES the
existing linked-department set, it is never merged/unioned.** This matches
every other bulk-upload UPSERT in this app (e.g. `department_import.py`'s
own full-row overwrite-on-update) -- the uploaded file is the source of
truth for what it explicitly states. A row with a blank "Department Codes"
cell on an UPDATE therefore clears the designation's department links
entirely (a designation with zero linked departments is valid today, per
`DesignationCreate.department_ids`'s own `default_factory=list`) -- not a
special case, just the natural consequence of "replace, don't merge".

**Multiple in-file rows sharing one (Name, Category) are UNIONED, not
rejected as duplicates** (changed 2026-08-27 after real user testing -- see
below). Writing one row per department is the natural reading of the
original "Name + Category + Department" spec wording, and it is what a real
user actually did on first use: `Assistant Professor (SG)/TEACHING/AIDS` on
one row and `Assistant Professor (SG)/TEACHING/CSE` on the next. The first
implementation rejected the second row as a duplicate, which silently
dropped the CSE link and reported "unchanged" -- wrong in the most damaging
direction, since the upload looked like it had nothing to do.

So rows sharing a key are treated as *contributions to the same
designation*: their `Department Codes` are unioned, the FIRST such row
carries the resulting created/updated/unchanged status (its preview shows
the full combined code set), and each later row is marked `merged` with
`merged_into_row` pointing at that first row. Nothing is silently dropped
and nothing is hidden from the preview.

This does NOT weaken "the uploaded file is the source of truth": the union
is computed within the file, then still REPLACES whatever department set the
designation had in the DB. Union-within-file, replace-against-DB.

The one case still rejected is a genuine contradiction: two rows sharing a
(Name, Category) but disagreeing on a field that describes the designation
itself (Qualification / Min Experience / Employment Type / Required Skills /
Active -- `_GROUPED_CONFLICT_FIELDS`). Those cannot both be true, so the
later row is rejected naming exactly which columns disagree, rather than
silently letting one row win.

**"Allow the same designation to exist in different categories/departments."**
The same Name under a *different* Category remains a genuinely distinct
designation and is never unioned with it -- only an exact (Name, Category)
match groups.

**Department Code resolution is campus-agnostic and multi-match-tolerant.**
Unlike `eligibility_rule_import.py`'s own single-campus-scoped Department
Code column, a Designation's department mapping has no campus dimension at
all (`designation_departments` carries no `campus_id`) -- a single "Department
Codes" cell can legitimately resolve to departments across *multiple*
campuses that happen to share the same code (real, pre-existing collisions
of this kind already exist -- see `department_import.py`'s own docstring for
the SHIFT-campus example). So each code in this column is resolved against
*every* Department row carrying that code, not scoped to any one campus, and
ALL departments a code resolves to are linked. If a code matches zero real
departments, or if *any* department it resolves to has a category that
doesn't match the row's own Category, the whole row is rejected (never
partially applied) -- naming which code failed and why, mirroring
`app/api/v1/routers/designations.py::_validate_department_categories`'s own
manual-create/update enforcement of the same "every linked department must
share the designation's category" rule.

**Undo scope**: same as `department_import.py`/`location_import.py`/
`housekeeping_staff_import.py` -- Designation has no permanent old-value
history table, so `commit_rows` writes one `BulkUploadRowLog` row per
non-rejected row (created/updated/unchanged alike), which `undo_bulk_upload`
(app/api/v1/routers/sanctioned_strength.py) uses to soft-delete only the
rows this batch actually created.
"""

import io
import re
import uuid
from dataclasses import dataclass, field

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session, selectinload

from app.models.bulk_upload_log import BulkUploadLog
from app.models.bulk_upload_row_log import BulkUploadRowLog
from app.models.department import Department
from app.models.designation import Designation
from app.models.enums import BulkUploadEntityTypeEnum, EmploymentTypeEnum, StaffRoleCategoryEnum
from app.services.sanctioned_strength_import import parse_rows  # noqa: F401 -- re-exported, see module docstring

MAX_ROWS = 5000

NAME_HEADER = "Designation Name"
CATEGORY_HEADER = "Category"
DEPARTMENT_CODES_HEADER = "Department Codes"
QUALIFICATION_HEADER = "Minimum Qualification"
MIN_EXPERIENCE_HEADER = "Minimum Experience"
EMPLOYMENT_TYPE_HEADER = "Employment Type"
REQUIRED_SKILLS_HEADER = "Required Skills"
ACTIVE_HEADER = "Active"

TEMPLATE_HEADERS = (
    NAME_HEADER,
    CATEGORY_HEADER,
    DEPARTMENT_CODES_HEADER,
    QUALIFICATION_HEADER,
    MIN_EXPERIENCE_HEADER,
    EMPLOYMENT_TYPE_HEADER,
    REQUIRED_SKILLS_HEADER,
    ACTIVE_HEADER,
)

# Same functional backstop as the other siblings' own example rows -- a
# guaranteed-unknown Department Code ("XXXNOPE") means a forgotten example
# row always comes back rejected rather than silently importing as real
# data, even though every other column is otherwise a valid value.
_EXAMPLE_ROWS = (
    (
        "EXAMPLE - DELETE THIS ROW - Assistant Professor",
        "TEACHING",
        "XXXNOPE",
        "PhD in relevant field",
        "3+ years",
        "FULL_TIME",
        "Curriculum design, Research methodology",
        "TRUE",
    ),
    (
        "EXAMPLE - DELETE THIS ROW - Office Assistant",
        "NON_TEACHING",
        "XXXNOPE",
        "Bachelor's degree",
        "1+ years",
        "FULL_TIME",
        "MS Office, Communication",
        "TRUE",
    ),
)

_HEADER_FILL = PatternFill(start_color="1B5FAA", end_color="1B5FAA", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SAMPLE_FILL = PatternFill(start_color="FFF3D6", end_color="FFF3D6", fill_type="solid")

_CATEGORY_VALUES = tuple(c.value for c in StaffRoleCategoryEnum)
_EMPLOYMENT_TYPE_VALUES = tuple(e.value for e in EmploymentTypeEnum)


# Fields that describe the DESIGNATION ITSELF (as opposed to
# `Department Codes`, which describes its links). Two rows sharing a
# (Name, Category) key may contribute different department codes -- that is
# the whole point of the union -- but they cannot disagree on these without
# one of them being wrong, so a mismatch rejects the later row instead of
# silently letting the first row win. `(attribute, template header)` pairs so
# the rejection message names the real spreadsheet column.
_GROUPED_CONFLICT_FIELDS = (
    ("qualification", QUALIFICATION_HEADER),
    ("min_experience", MIN_EXPERIENCE_HEADER),
    ("employment_type", EMPLOYMENT_TYPE_HEADER),
    ("required_skills", REQUIRED_SKILLS_HEADER),
    ("is_active", ACTIVE_HEADER),
)


@dataclass
class ImportRowResult:
    row_number: int
    status: str  # created | updated | unchanged | merged | rejected
    error_reason: str | None = None
    # Set only on a `merged` row: the row_number of the earlier row for the
    # same (Name, Category) that this row's Department Codes were folded
    # into, and which carries the real created/updated/unchanged status.
    merged_into_row: int | None = None
    name: str | None = None
    category: StaffRoleCategoryEnum | None = None
    department_codes: list[str] = field(default_factory=list)
    qualification: str | None = None
    min_experience: str | None = None
    employment_type: EmploymentTypeEnum | None = None
    required_skills: str | None = None
    is_active: bool | None = None
    # Resolved only for non-rejected rows -- commit_rows() uses these
    # directly rather than re-resolving lookups a second time.
    department_ids: list[uuid.UUID] = field(default_factory=list)
    designation_id: uuid.UUID | None = None  # the existing match's id, when status != "created"


@dataclass
class ValidationResult:
    rows: list[ImportRowResult]
    total: int
    created_count: int
    updated_count: int
    unchanged_count: int
    rejected_count: int
    # Rows folded into an earlier row for the same designation. Deliberately
    # counted separately from created/updated/unchanged: a merged row writes
    # nothing of its own, so counting it as an update would double-count the
    # single designation its group actually touches.
    merged_count: int = 0


def _cell(row: dict, key: str) -> str:
    value = row.get(key)
    if value is None:
        return ""
    return str(value).strip()


def _normalize(text: str | None) -> str:
    """Trim, collapse repeated internal whitespace, and casefold -- used
    ONLY to decide whether two rows/values refer to the same real-world
    designation or department code (the in-file duplicate check, matching
    against existing DB rows, and department-code lookup). Comparison-only:
    never mutates what actually gets stored."""
    if not text:
        return ""
    return re.sub(r"\s+", " ", text.strip()).casefold()


def _parse_category(text: str) -> tuple[StaffRoleCategoryEnum | None, str | None]:
    """Case-insensitive and hyphen/space/underscore-tolerant -- accepts
    'NON-TEACHING', 'NON_TEACHING', or 'Non Teaching' alike, all normalizing
    to the same real enum value."""
    if not text:
        return None, f"Missing required column '{CATEGORY_HEADER}'"
    normalized = re.sub(r"[\s\-]+", "_", text.strip()).upper()
    if normalized not in _CATEGORY_VALUES:
        return None, f"Unknown '{CATEGORY_HEADER}' value '{text}'. Valid values: {', '.join(_CATEGORY_VALUES)}"
    return StaffRoleCategoryEnum(normalized), None


def _parse_employment_type(text: str) -> tuple[EmploymentTypeEnum | None, str | None]:
    """Same case/hyphen/space tolerance as `_parse_category` above (e.g.
    'Full Time' / 'full-time' / 'FULL_TIME' all resolve to FULL_TIME)."""
    if not text:
        return None, f"Missing required column '{EMPLOYMENT_TYPE_HEADER}'"
    normalized = re.sub(r"[\s\-]+", "_", text.strip()).upper()
    if normalized not in _EMPLOYMENT_TYPE_VALUES:
        return None, (
            f"Unknown '{EMPLOYMENT_TYPE_HEADER}' value '{text}'. Valid values: {', '.join(_EMPLOYMENT_TYPE_VALUES)}"
        )
    return EmploymentTypeEnum(normalized), None


def _parse_active(text: str) -> tuple[bool | None, str | None]:
    """Blank -> default True, matching the New Designation form's own
    `is_active: bool = True` default. Otherwise case-insensitive TRUE/FALSE
    only -- anything else is a clear rejection naming the bad value."""
    if not text:
        return True, None
    upper = text.upper()
    if upper == "TRUE":
        return True, None
    if upper == "FALSE":
        return False, None
    return None, f"Invalid '{ACTIVE_HEADER}' value '{text}' -- expected TRUE or FALSE (blank defaults to TRUE)"


def _resolve_department_codes(
    departments_by_code: dict[str, list[Department]],
    codes: list[str],
    expected_category: StaffRoleCategoryEnum | None,
) -> tuple[list[Department], list[str]]:
    """Resolves every code in `codes` against the campus-agnostic
    `departments_by_code` cache (see module docstring for why this is never
    scoped to a single campus). Every department a code resolves to must
    share `expected_category` (skipped entirely if `expected_category` is
    None -- i.e. the row's own Category was itself invalid/missing, already
    a separate error) -- a single mismatched department anywhere in a code's
    matches rejects that code (and therefore the whole row), naming exactly
    which department(s) and why, same as
    `designations.py::_validate_department_categories`'s own manual-create
    enforcement of this rule.
    """
    errors: list[str] = []
    resolved: list[Department] = []
    seen_ids: set[uuid.UUID] = set()
    for code in codes:
        matches = departments_by_code.get(_normalize(code), [])
        if not matches:
            errors.append(f"Unknown department code '{code}'")
            continue
        if expected_category is not None:
            mismatched = [department for department in matches if department.category != expected_category]
            if mismatched:
                names = ", ".join(
                    f"'{department.name}' ({department.campus.code if department.campus else '?'}) is "
                    f"{department.category.value}"
                    for department in mismatched
                )
                errors.append(
                    f"Department code '{code}' matches {names} but designation category is "
                    f"{expected_category.value}"
                )
                continue
        for department in matches:
            if department.id not in seen_ids:
                seen_ids.add(department.id)
                resolved.append(department)
    return resolved, errors


def validate_rows(db: Session, raw_rows: list[dict]) -> ValidationResult:
    """Parses+validates every row against the current DB state **without
    writing anything** -- same read-only contract as every other importer's
    `validate_rows`, called by both `/designations/bulk-upload/validate` and
    `/designations/bulk-upload/commit` (which re-validates the re-uploaded
    file before writing).
    """
    if len(raw_rows) > MAX_ROWS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"File has {len(raw_rows)} data rows, exceeding the {MAX_ROWS}-row limit",
        )

    results: list[ImportRowResult] = []

    # Department code cache, built once -- global, NOT per-campus (see module
    # docstring: Designation<->Department mapping is campus-agnostic).
    departments_by_code: dict[str, list[Department]] = {}
    for department in (
        db.query(Department).filter(Department.code.isnot(None)).options(selectinload(Department.campus)).all()
    ):
        departments_by_code.setdefault(_normalize(department.code), []).append(department)

    # Existing Designation rows, fetched once (not once per row), matched in
    # Python by normalized (name, category) -- same defensive "fetch once,
    # match in Python" batching technique department_import.py's own
    # existing_by_campus cache uses. Ordered by created_at so, in the rare
    # case pre-existing data already has more than one row sharing a
    # (name, category) pair, the earliest one is treated as canonical.
    existing_by_key: dict[tuple[str, str], Designation] = {}
    for designation in (
        db.query(Designation).options(selectinload(Designation.departments)).order_by(Designation.created_at).all()
    ):
        key = (_normalize(designation.name), designation.category.value)
        existing_by_key.setdefault(key, designation)

    for row_number, raw_row in enumerate(raw_rows, start=2):
        name = _cell(raw_row, NAME_HEADER)
        category_raw = _cell(raw_row, CATEGORY_HEADER)
        department_codes_raw = _cell(raw_row, DEPARTMENT_CODES_HEADER)
        qualification = _cell(raw_row, QUALIFICATION_HEADER)
        min_experience = _cell(raw_row, MIN_EXPERIENCE_HEADER)
        employment_type_raw = _cell(raw_row, EMPLOYMENT_TYPE_HEADER)
        required_skills = _cell(raw_row, REQUIRED_SKILLS_HEADER) or None
        active_raw = _cell(raw_row, ACTIVE_HEADER)

        errors: list[str] = []

        if not name:
            errors.append(f"Missing required column '{NAME_HEADER}'")

        category, category_error = _parse_category(category_raw)
        if category_error:
            errors.append(category_error)

        if not qualification:
            errors.append(f"Missing required column '{QUALIFICATION_HEADER}'")

        if not min_experience:
            errors.append(f"Missing required column '{MIN_EXPERIENCE_HEADER}'")

        employment_type, employment_type_error = _parse_employment_type(employment_type_raw)
        if employment_type_error:
            errors.append(employment_type_error)

        is_active, active_error = _parse_active(active_raw)
        if active_error:
            errors.append(active_error)

        department_codes = [code.strip() for code in re.split(r"[;,]", department_codes_raw) if code.strip()]
        resolved_departments, dept_errors = _resolve_department_codes(departments_by_code, department_codes, category)
        errors.extend(dept_errors)
        department_ids = [department.id for department in resolved_departments]

        result = ImportRowResult(
            row_number=row_number,
            # Placeholder. Rows with errors keep "rejected"; every other row
            # has its real status assigned by the grouping pass below.
            status="rejected",
            name=name or None,
            category=category,
            department_codes=department_codes,
            qualification=qualification or None,
            min_experience=min_experience or None,
            employment_type=employment_type,
            required_skills=required_skills,
            is_active=is_active,
        )

        if errors:
            result.error_reason = "; ".join(errors)
        else:
            result.department_ids = department_ids

        results.append(result)

    # --- Grouping pass: fold rows describing the same designation ---------
    #
    # Every row that survived its own field validation is grouped by
    # (normalized Name, Category). A group of one behaves exactly as before.
    # A group of several is the "one row per department" shape a real user
    # naturally writes -- unioned rather than rejected (see module docstring).
    groups: dict[tuple[str, str], list[ImportRowResult]] = {}
    for result in results:
        if result.error_reason is not None:
            continue
        groups.setdefault((_normalize(result.name), result.category.value), []).append(result)

    for lookup_key, group in groups.items():
        primary, followers = group[0], group[1:]

        # A follower that contradicts the primary on a field describing the
        # designation itself is rejected -- it cannot be folded in, because
        # there is no honest way to decide which value wins.
        contributors = [primary]
        for follower in followers:
            mismatched = [
                header
                for attribute, header in _GROUPED_CONFLICT_FIELDS
                if getattr(follower, attribute) != getattr(primary, attribute)
            ]
            if mismatched:
                follower.status = "rejected"
                follower.error_reason = (
                    f"Row {primary.row_number} describes this same designation (same "
                    f"'{NAME_HEADER}' and '{CATEGORY_HEADER}') but disagrees on: "
                    f"{', '.join(mismatched)}. Make these columns match on every row for "
                    f"this designation, or put all its '{DEPARTMENT_CODES_HEADER}' on one row."
                )
            else:
                contributors.append(follower)

        # Union of every contributing row's departments, order-preserving.
        union_ids: list[uuid.UUID] = []
        union_codes: list[str] = []
        seen_ids: set[uuid.UUID] = set()
        seen_codes: set[str] = set()
        for contributor in contributors:
            for department_id in contributor.department_ids:
                if department_id not in seen_ids:
                    seen_ids.add(department_id)
                    union_ids.append(department_id)
            for code in contributor.department_codes:
                if _normalize(code) not in seen_codes:
                    seen_codes.add(_normalize(code))
                    union_codes.append(code)

        # Followers write nothing themselves; the primary applies the union.
        # Their own `department_codes` are left untouched so the preview still
        # shows what each row contributed.
        for follower in contributors[1:]:
            follower.status = "merged"
            follower.merged_into_row = primary.row_number
            follower.department_ids = []

        primary.department_ids = union_ids
        primary.department_codes = union_codes

        existing = existing_by_key.get(lookup_key)
        if existing is None:
            primary.status = "created"
        else:
            primary.designation_id = existing.id
            existing_department_ids = {department.id for department in existing.departments}
            if (
                existing.name != primary.name
                or existing.qualification != primary.qualification
                or existing.min_experience != primary.min_experience
                or existing.employment_type != primary.employment_type
                or (existing.required_skills or None) != (primary.required_skills or None)
                or existing.is_active != primary.is_active
                or existing_department_ids != set(union_ids)
            ):
                primary.status = "updated"
            else:
                primary.status = "unchanged"

    return ValidationResult(
        rows=results,
        total=len(raw_rows),
        created_count=sum(1 for row in results if row.status == "created"),
        updated_count=sum(1 for row in results if row.status == "updated"),
        unchanged_count=sum(1 for row in results if row.status == "unchanged"),
        rejected_count=sum(1 for row in results if row.status == "rejected"),
        merged_count=sum(1 for row in results if row.status == "merged"),
    )


def commit_rows(db: Session, *, validation: ValidationResult, bulk_upload_log_id: uuid.UUID) -> None:
    """Applies every non-rejected row's UPSERT within the *caller's* existing
    transaction -- no `db.commit()` here, same contract as every other
    importer's `commit_rows`. Also writes one `BulkUploadRowLog` row per
    non-rejected row (see this module's own docstring for why -- Designation
    has no permanent old-value history table, so undo relies on this
    instead).

    On an "updated" row, `existing.departments` is REPLACED wholesale with
    the row's resolved department set (never merged) -- see this module's
    own docstring for why that's the correct reading of an ambiguous spec
    against this schema's real many-to-many shape.
    """
    for row in validation.rows:
        if row.status == "created":
            departments = (
                db.query(Department).filter(Department.id.in_(row.department_ids)).all() if row.department_ids else []
            )
            new_row = Designation(
                name=row.name,
                category=row.category,
                qualification=row.qualification,
                min_experience=row.min_experience,
                employment_type=row.employment_type,
                required_skills=row.required_skills,
                is_active=row.is_active,
            )
            new_row.departments = departments
            db.add(new_row)
            db.flush()
            db.add(
                BulkUploadRowLog(
                    bulk_upload_log_id=bulk_upload_log_id,
                    entity_type=BulkUploadEntityTypeEnum.DESIGNATION,
                    entity_id=new_row.id,
                    was_created=True,
                )
            )
        elif row.status == "updated":
            existing = db.get(Designation, row.designation_id)
            existing.name = row.name
            existing.category = row.category
            existing.qualification = row.qualification
            existing.min_experience = row.min_experience
            existing.employment_type = row.employment_type
            existing.required_skills = row.required_skills
            existing.is_active = row.is_active
            departments = (
                db.query(Department).filter(Department.id.in_(row.department_ids)).all() if row.department_ids else []
            )
            existing.departments = departments
            db.add(
                BulkUploadRowLog(
                    bulk_upload_log_id=bulk_upload_log_id,
                    entity_type=BulkUploadEntityTypeEnum.DESIGNATION,
                    entity_id=existing.id,
                    was_created=False,
                )
            )
        elif row.status == "unchanged":
            db.add(
                BulkUploadRowLog(
                    bulk_upload_log_id=bulk_upload_log_id,
                    entity_type=BulkUploadEntityTypeEnum.DESIGNATION,
                    entity_id=row.designation_id,
                    was_created=False,
                )
            )
        # "rejected": no write, no row log entry.
        # "merged": also no write and no row log entry -- this row's
        # departments were already folded into its group's primary row above,
        # which does the single write for the whole group. Giving a merged row
        # its own row log would make undo try to revert the same designation
        # more than once.


def _write_header(ws, headers: tuple[str, ...]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 26


def build_bulk_upload_template_xlsx(db: Session) -> bytes:
    """Live-generated, same styling shape as every other importer's template
    builder: locked/frozen styled header row, highlighted example rows
    marked for deletion (harmless if left in -- see `_EXAMPLE_ROWS`'s
    docstring above), and a "Master Lists" reference sheet of currently-valid
    Category/Employment Type values plus real, currently-known Department
    Codes, so a user can copy exact values in.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Designations"
    _write_header(ws, TEMPLATE_HEADERS)
    for example in _EXAMPLE_ROWS:
        ws.append(list(example))
    for row_idx in range(2, 2 + len(_EXAMPLE_ROWS)):
        for col in range(1, len(TEMPLATE_HEADERS) + 1):
            ws.cell(row=row_idx, column=col).fill = _SAMPLE_FILL

    category_formula = '"' + ",".join(_CATEGORY_VALUES) + '"'
    category_dv = DataValidation(type="list", formula1=category_formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(category_dv)
    category_dv.add("B2:B500")

    employment_type_formula = '"' + ",".join(_EMPLOYMENT_TYPE_VALUES) + '"'
    employment_type_dv = DataValidation(
        type="list", formula1=employment_type_formula, allow_blank=True, showDropDown=False
    )
    ws.add_data_validation(employment_type_dv)
    employment_type_dv.add("F2:F500")

    active_formula = '"TRUE,FALSE"'
    active_dv = DataValidation(type="list", formula1=active_formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(active_dv)
    active_dv.add("H2:H500")

    department_codes = sorted(
        {code for (code,) in db.query(Department.code).filter(Department.code.isnot(None)).distinct().all() if code}
    )

    master_ws = wb.create_sheet("Master Lists")
    master_ws.append(["Category", "Employment Type", "Department Code"])
    for cell in master_ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    max_rows = max(len(_CATEGORY_VALUES), len(_EMPLOYMENT_TYPE_VALUES), len(department_codes))
    for row_idx in range(max_rows):
        if row_idx < len(_CATEGORY_VALUES):
            master_ws.cell(row=row_idx + 2, column=1, value=_CATEGORY_VALUES[row_idx])
        if row_idx < len(_EMPLOYMENT_TYPE_VALUES):
            master_ws.cell(row=row_idx + 2, column=2, value=_EMPLOYMENT_TYPE_VALUES[row_idx])
        if row_idx < len(department_codes):
            master_ws.cell(row=row_idx + 2, column=3, value=department_codes[row_idx])
    for col in range(1, 4):
        master_ws.column_dimensions[get_column_letter(col)].width = 26

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_error_report_xlsx(log: BulkUploadLog, rejected_rows: list[ImportRowResult]) -> bytes:
    """Same shape as every other importer's `build_error_report_xlsx` --
    only this upload's rejected rows, plus the `error_reason` column."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rejected rows"
    ws.append([f"Bulk upload: {log.filename}"])
    ws.append([f"Uploaded: {log.uploaded_at.isoformat()}"])
    ws.append([])
    headers = list(TEMPLATE_HEADERS) + ["error_reason"]
    ws.append(headers)
    for row in rejected_rows:
        ws.append(
            [
                row.name,
                row.category.value if row.category else None,
                ", ".join(row.department_codes),
                row.qualification,
                row.min_experience,
                row.employment_type.value if row.employment_type else None,
                row.required_skills,
                "TRUE" if row.is_active else ("FALSE" if row.is_active is not None else None),
                row.error_reason,
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
