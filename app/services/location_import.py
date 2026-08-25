"""Location bulk upload: validate -> preview -> commit (Phase J,
glowing-zooming-hamming.md). Sibling module to `sanctioned_strength_import.py`
(same dataclass/validate/commit/undo-adjacent shape, deliberately NOT a
shared abstraction between entities per this phase's dispatch brief), with
one exception: `parse_rows` (file-bytes -> `list[dict]`, dispatching on
`.xlsx`/`.csv` extension) is genuinely entity-agnostic already -- it never
references Sanctioned-Strength-specific columns -- so this module imports
and reuses it directly from `sanctioned_strength_import` rather than keeping
3 byte-identical copies. Everything else here (headers, validation, the
UPSERT resolution, the two xlsx builders) is its own copy, matching the
dispatch brief's explicit "your call, document which and why."

**UPSERT key** (fixed 2026-08-25 -- see the bug report this change closes):
the full composite `(campus_code, name, block_building, floor_venue,
category)`, normalized before comparison (trimmed, repeated internal
whitespace collapsed, case-insensitive). A building's separate floors/venues
are distinct real-world locations, not one location with a mutable floor
attribute -- matching on `(campus_id, name)` alone (the previous key) let a
row for "RB Block / First Floor" silently match and overwrite an existing
"RB Block / Ground Floor" row for the same building name, and rejected a
second floor as an in-file duplicate of the first. `Location`
(app/models/location.py) has no DB-level unique constraint on this
composite -- unlike `Department`, which has a real `uq_department_campus_name`
constraint and so can be resolved with `.one_or_none()` in
`sanctioned_strength_import.py`. Here, matching iterates every existing
Location for the row's campus (fetched once per campus, not once per row)
ordered by `created_at`, picking the first whose own normalized composite
key equals the row's -- a deterministic pick for the rare case where
pre-existing data already has a genuine composite-key collision, same
accepted-limitation shape as before this fix, just scoped to the wider key.
A row is `updated` only when its composite key already matches an existing
row (same real-world location) but the raw stored text differs from what
was just uploaded (pure case/whitespace normalization drift, since every
other business field is now part of the identity itself); `unchanged` if
byte-identical; `created` if no existing row shares that composite key.

**Undo scope**: unlike Sanctioned Strength (whose undo replays
`SanctionedStrengthHistory.old_value`), Location has no permanent
old-value history table, so `commit_rows` below additionally writes one
`BulkUploadRowLog` row per non-rejected row (created/updated/unchanged
alike -- see that model's own docstring for why "per non-rejected row",
not just "per created row"), which is what `undo_bulk_upload`
(app/api/v1/routers/sanctioned_strength.py) uses to soft-delete only the
rows this batch actually created, skipping (and counting) the rest.
"""

import io
import re
import uuid
from dataclasses import dataclass

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from app.models.bulk_upload_log import BulkUploadLog
from app.models.bulk_upload_row_log import BulkUploadRowLog
from app.models.campus import Campus
from app.models.enums import CAMPUS_CODES, BulkUploadEntityTypeEnum, StaffRoleCategoryEnum
from app.models.location import Location
from app.services.sanctioned_strength_import import parse_rows  # noqa: F401 -- re-exported, see module docstring

MAX_ROWS = 5000

CAMPUS_HEADER = "Campus Code"
NAME_HEADER = "Location Name"
BLOCK_HEADER = "Block/Building"
FLOOR_HEADER = "Floor/Venue"
CATEGORY_HEADER = "Category"

TEMPLATE_HEADERS = (CAMPUS_HEADER, NAME_HEADER, BLOCK_HEADER, FLOOR_HEADER, CATEGORY_HEADER)

# Same functional backstop as sanctioned_strength_import.py's own example
# rows -- "XXX" is never a real campus code, so a forgotten example row
# always comes back rejected rather than silently importing as real data.
_EXAMPLE_ROWS = (
    ("XXX", "EXAMPLE - DELETE THIS ROW", "Block A", "Ground Floor", "TEACHING"),
    ("XXX", "EXAMPLE - DELETE THIS ROW", "", "", ""),
)

_HEADER_FILL = PatternFill(start_color="1B5FAA", end_color="1B5FAA", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SAMPLE_FILL = PatternFill(start_color="FFF3D6", end_color="FFF3D6", fill_type="solid")

_CATEGORY_VALUES = tuple(c.value for c in StaffRoleCategoryEnum)


@dataclass
class ImportRowResult:
    row_number: int
    status: str  # created | updated | unchanged | rejected
    error_reason: str | None = None
    campus_code: str | None = None
    location_name: str | None = None
    block_building: str | None = None
    floor_venue: str | None = None
    category: StaffRoleCategoryEnum | None = None
    # Resolved only for non-rejected rows -- commit_rows() uses these
    # directly rather than re-resolving lookups a second time.
    campus_id: uuid.UUID | None = None
    location_id: uuid.UUID | None = None  # the existing match's id, when status != "created"


@dataclass
class ValidationResult:
    rows: list[ImportRowResult]
    total: int
    created_count: int
    updated_count: int
    unchanged_count: int
    rejected_count: int


def _cell(row: dict, key: str) -> str:
    value = row.get(key)
    if value is None:
        return ""
    return str(value).strip()


def _parse_optional_category(text: str) -> tuple[StaffRoleCategoryEnum | None, str | None]:
    if not text:
        return None, None
    upper = text.upper()
    if upper not in _CATEGORY_VALUES:
        return None, f"Unknown '{CATEGORY_HEADER}' value '{text}'. Valid values: {', '.join(_CATEGORY_VALUES)}"
    return StaffRoleCategoryEnum(upper), None


def _normalize(text: str | None) -> str:
    """Trim, collapse repeated internal whitespace, and casefold -- used
    ONLY to decide whether two rows refer to the same real-world location
    (the in-file duplicate check, and matching against existing DB rows).
    Comparison-only: never mutates what actually gets stored, so "RB  Block"
    and "rb block" are treated as identical for matching purposes but the
    row's own trimmed-but-original-case text is still what's written on
    create/update."""
    if not text:
        return ""
    return re.sub(r"\s+", " ", text.strip()).casefold()


def _composite_key(
    campus_code: str,
    location_name: str,
    block_building: str | None,
    floor_venue: str | None,
    category: StaffRoleCategoryEnum | None,
) -> tuple[str, str, str, str, str]:
    """The full, normalized 5-part identity a Location is matched/deduped
    by: Campus Code + Location Name + Block/Building + Floor/Venue +
    Category. All five must match for two rows to be considered the same
    location -- see this module's own docstring for why (a building's
    separate floors/venues are distinct real-world locations, not one
    location with a mutable floor attribute)."""
    return (
        _normalize(campus_code),
        _normalize(location_name),
        _normalize(block_building),
        _normalize(floor_venue),
        category.value if category else "",
    )


def validate_rows(db: Session, raw_rows: list[dict]) -> ValidationResult:
    """Parses+validates every row against the current DB state **without
    writing anything** -- same read-only contract as
    `sanctioned_strength_import.validate_rows`, called by both
    `/locations/bulk-upload/validate` and `/locations/bulk-upload/commit`
    (which re-validates the re-uploaded file before writing).
    """
    if len(raw_rows) > MAX_ROWS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"File has {len(raw_rows)} data rows, exceeding the {MAX_ROWS}-row limit",
        )

    results: list[ImportRowResult] = []
    seen_keys: dict[tuple[str, str, str, str, str], int] = {}
    # Existing Location rows for a given campus, fetched at most once per
    # campus (not once per row) and reused for every row on that campus --
    # matching is now a full in-Python composite-key comparison (see
    # `_composite_key`), not something a single ILIKE query can express once
    # whitespace-collapsing is part of "normalized".
    existing_by_campus: dict[uuid.UUID, list[Location]] = {}
    created = updated = unchanged = rejected = 0

    for row_number, raw_row in enumerate(raw_rows, start=2):
        campus_code = _cell(raw_row, CAMPUS_HEADER).upper()
        location_name = _cell(raw_row, NAME_HEADER)
        block_building = _cell(raw_row, BLOCK_HEADER) or None
        floor_venue = _cell(raw_row, FLOOR_HEADER) or None
        category_raw = _cell(raw_row, CATEGORY_HEADER)

        errors: list[str] = []

        if not campus_code:
            errors.append(f"Missing required column '{CAMPUS_HEADER}'")
        elif campus_code not in CAMPUS_CODES:
            errors.append(f"Unknown campus code '{campus_code}'")

        if not location_name:
            errors.append(f"Missing required column '{NAME_HEADER}'")

        category, category_error = _parse_optional_category(category_raw)
        if category_error:
            errors.append(category_error)

        campus = None
        if campus_code and campus_code in CAMPUS_CODES:
            campus = db.query(Campus).filter(Campus.code == campus_code).one_or_none()
            if campus is None:
                errors.append(f"Campus '{campus_code}' is not seeded in this environment")

        if campus_code and location_name:
            key = _composite_key(campus_code, location_name, block_building, floor_venue, category)
            if key in seen_keys:
                errors.append(
                    "Duplicate location: same Campus, Location, Block, Floor/Venue and Category "
                    f"(already used on row {seen_keys[key]})"
                )
            else:
                seen_keys[key] = row_number

        result = ImportRowResult(
            row_number=row_number,
            status="rejected",
            campus_code=campus_code or None,
            location_name=location_name or None,
            block_building=block_building,
            floor_venue=floor_venue,
            category=category,
        )

        if errors:
            result.error_reason = "; ".join(errors)
            results.append(result)
            rejected += 1
            continue

        result.campus_id = campus.id

        campus_locations = existing_by_campus.get(campus.id)
        if campus_locations is None:
            campus_locations = (
                db.query(Location).filter(Location.campus_id == campus.id).order_by(Location.created_at).all()
            )
            existing_by_campus[campus.id] = campus_locations

        row_key = _composite_key(campus_code, location_name, block_building, floor_venue, category)
        existing = next(
            (
                loc
                for loc in campus_locations
                if _composite_key(campus_code, loc.name, loc.block_building, loc.floor_venue, loc.category)
                == row_key
            ),
            None,
        )

        if existing is None:
            result.status = "created"
            created += 1
        elif (
            existing.name != location_name
            or existing.block_building != block_building
            or existing.floor_venue != floor_venue
            or existing.category != category
        ):
            # Same real-world location (normalized composite key matches)
            # but the raw stored text differs only in case/whitespace/
            # formatting from what was just uploaded -- re-stamp the
            # canonical uploaded formatting.
            result.status = "updated"
            result.location_id = existing.id
            updated += 1
        else:
            result.status = "unchanged"
            result.location_id = existing.id
            unchanged += 1

        results.append(result)

    return ValidationResult(
        rows=results,
        total=len(raw_rows),
        created_count=created,
        updated_count=updated,
        unchanged_count=unchanged,
        rejected_count=rejected,
    )


def commit_rows(
    db: Session, *, validation: ValidationResult, bulk_upload_log_id: uuid.UUID
) -> None:
    """Applies every non-rejected row's UPSERT within the *caller's* existing
    transaction -- no `db.commit()` here, same contract as
    `sanctioned_strength_import.commit_rows`. Also writes one
    `BulkUploadRowLog` row per non-rejected row (see this module's own
    docstring for why -- Location has no permanent old-value history table,
    so undo relies on this instead).

    `Location` carries no `created_by_id`/`updated_by_id` columns (unlike
    `HousekeepingStaff`), so there is no actor to stamp on the row itself
    here.
    """
    for row in validation.rows:
        if row.status == "created":
            new_row = Location(
                campus_id=row.campus_id,
                name=row.location_name,
                block_building=row.block_building,
                floor_venue=row.floor_venue,
                category=row.category,
            )
            db.add(new_row)
            db.flush()
            db.add(
                BulkUploadRowLog(
                    bulk_upload_log_id=bulk_upload_log_id,
                    entity_type=BulkUploadEntityTypeEnum.LOCATION,
                    entity_id=new_row.id,
                    was_created=True,
                )
            )
        elif row.status == "updated":
            existing = db.get(Location, row.location_id)
            # `name` too, not just block/floor/category -- "updated" can now
            # ONLY happen when the composite key already matched (same
            # real-world location) but some field's raw text differs purely
            # in case/whitespace from what was just uploaded, and that can
            # be `name` itself just as easily as block/floor/category.
            existing.name = row.location_name
            existing.block_building = row.block_building
            existing.floor_venue = row.floor_venue
            existing.category = row.category
            db.add(
                BulkUploadRowLog(
                    bulk_upload_log_id=bulk_upload_log_id,
                    entity_type=BulkUploadEntityTypeEnum.LOCATION,
                    entity_id=existing.id,
                    was_created=False,
                )
            )
        elif row.status == "unchanged":
            db.add(
                BulkUploadRowLog(
                    bulk_upload_log_id=bulk_upload_log_id,
                    entity_type=BulkUploadEntityTypeEnum.LOCATION,
                    entity_id=row.location_id,
                    was_created=False,
                )
            )
        # "rejected": no write, no row log entry.


def _write_header(ws, headers: tuple[str, ...]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 26


def build_bulk_upload_template_xlsx(db: Session) -> bytes:
    """Live-generated, same styling shape as
    `sanctioned_strength_import.build_bulk_upload_template_xlsx`: locked/
    frozen styled header row, highlighted example rows marked for deletion
    (harmless if left in -- see `_EXAMPLE_ROWS`'s docstring above), and a
    "Master Lists" reference sheet of currently-valid campus codes so a
    user can copy exact values in. No real production workbook exists yet
    for this entity (Phase J is scaffolding, per the plan) -- this is a
    from-scratch template, not derived from a real file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Locations"
    _write_header(ws, TEMPLATE_HEADERS)
    for example in _EXAMPLE_ROWS:
        ws.append(list(example))
    for row_idx in range(2, 2 + len(_EXAMPLE_ROWS)):
        for col in range(1, len(TEMPLATE_HEADERS) + 1):
            ws.cell(row=row_idx, column=col).fill = _SAMPLE_FILL

    campus_formula = '"' + ",".join(CAMPUS_CODES) + '"'
    campus_dv = DataValidation(type="list", formula1=campus_formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(campus_dv)
    campus_dv.add("A2:A500")

    category_formula = '"' + ",".join(_CATEGORY_VALUES) + '"'
    category_dv = DataValidation(type="list", formula1=category_formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(category_dv)
    category_dv.add("E2:E500")

    master_ws = wb.create_sheet("Master Lists")
    master_ws.append(["Campus Code", "Category (optional)"])
    for cell in master_ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    max_rows = max(len(CAMPUS_CODES), len(_CATEGORY_VALUES))
    for row_idx in range(max_rows):
        if row_idx < len(CAMPUS_CODES):
            master_ws.cell(row=row_idx + 2, column=1, value=CAMPUS_CODES[row_idx])
        if row_idx < len(_CATEGORY_VALUES):
            master_ws.cell(row=row_idx + 2, column=2, value=_CATEGORY_VALUES[row_idx])
    for col in range(1, 3):
        master_ws.column_dimensions[get_column_letter(col)].width = 26

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_error_report_xlsx(log: BulkUploadLog, rejected_rows: list[ImportRowResult]) -> bytes:
    """Same shape as `sanctioned_strength_import.build_error_report_xlsx` --
    only this upload's rejected rows, plus the `error_reason` column."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rejected rows"
    ws.append([f"Bulk upload: {log.filename}"])
    ws.append([f"Uploaded: {log.uploaded_at.isoformat()}"])
    ws.append([])
    headers = list(TEMPLATE_HEADERS) + ["error_reason"]
    ws.append(headers)
    for row in rejected_rows:
        ws.append(
            [
                row.campus_code,
                row.location_name,
                row.block_building,
                row.floor_venue,
                row.category.value if row.category else None,
                row.error_reason,
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
