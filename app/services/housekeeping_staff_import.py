"""HousekeepingStaff bulk upload: validate -> preview -> commit (Phase J,
glowing-zooming-hamming.md). Sibling module to `sanctioned_strength_import.py`
and `location_import.py` (same shape, deliberately not a shared
abstraction). `parse_rows` is reused as-is from `sanctioned_strength_import`
(genuinely entity-agnostic -- see that import's own comment in
`location_import.py` for the reasoning, not repeated per module).

**UPSERT key**: `(campus_id, bio_id)` -- unlike Location, `HousekeepingStaff`
DOES have a real DB `UniqueConstraint` on this pair
(`uq_housekeeping_staff_campus_bio_id`, app/models/housekeeping_staff.py),
so key resolution below is a real equality lookup via `.one_or_none()`, not
fuzzy/ordered matching. A row is `updated` if any of
designation_id/location_id/block/floor_venue/shift/supervisor differs from
the existing match; `unchanged` if identical; `created` if no existing
bio_id match for that campus.

**Designation/Location resolution**: `Designation Name` must resolve
(case-insensitively, matching `sanctioned_strength_import.py`'s own
designation lookup) to a real `Designation` whose `category == HOUSEKEEPING`
-- the error message deliberately mirrors
`app/api/v1/routers/housekeeping_staff.py::_validate_housekeeping_designation`'s
own wording. `Location Name` must resolve (case-insensitively, scoped to the
row's own campus -- same convention as Department resolution in
`sanctioned_strength_import.py`) to a real `Location` for that campus.

**Undo scope**: same as `location_import.py` -- no permanent old-value
history table exists for HousekeepingStaff, so `commit_rows` writes one
`BulkUploadRowLog` row per non-rejected row (created/updated/unchanged
alike), which `undo_bulk_upload` uses to soft-delete only batch-created rows.
"""

import io
import uuid
from dataclasses import dataclass

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from app.models.bulk_upload_log import BulkUploadLog
from app.models.bulk_upload_row_log import BulkUploadRowLog
from app.models.campus import Campus
from app.models.designation import Designation
from app.models.enums import CAMPUS_CODES, BulkUploadEntityTypeEnum, HousekeepingShiftEnum, StaffRoleCategoryEnum
from app.models.housekeeping_staff import HousekeepingStaff
from app.models.location import Location
from app.models.user import User
from app.services.sanctioned_strength_import import parse_rows  # noqa: F401 -- re-exported, see module docstring

MAX_ROWS = 5000

CAMPUS_HEADER = "Campus Code"
BIO_ID_HEADER = "Bio ID"
NAME_HEADER = "Name"
DESIGNATION_HEADER = "Designation Name"
LOCATION_HEADER = "Location Name"
BLOCK_HEADER = "Block"
FLOOR_HEADER = "Floor/Venue"
SHIFT_HEADER = "Shift"
SUPERVISOR_HEADER = "Supervisor"

TEMPLATE_HEADERS = (
    CAMPUS_HEADER,
    BIO_ID_HEADER,
    NAME_HEADER,
    DESIGNATION_HEADER,
    LOCATION_HEADER,
    BLOCK_HEADER,
    FLOOR_HEADER,
    SHIFT_HEADER,
    SUPERVISOR_HEADER,
)

# Same functional backstop as the other two importers' own example rows.
_EXAMPLE_ROWS = (
    ("XXX", "BIO-0001", "EXAMPLE - DELETE THIS ROW", "Housekeeping Staff", "Main Block", "", "", "MORNING", ""),
)

_HEADER_FILL = PatternFill(start_color="1B5FAA", end_color="1B5FAA", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SAMPLE_FILL = PatternFill(start_color="FFF3D6", end_color="FFF3D6", fill_type="solid")

_SHIFT_VALUES = tuple(s.value for s in HousekeepingShiftEnum)


@dataclass
class ImportRowResult:
    row_number: int
    status: str  # created | updated | unchanged | rejected
    error_reason: str | None = None
    campus_code: str | None = None
    bio_id: str | None = None
    name: str | None = None
    designation_name: str | None = None
    location_name: str | None = None
    block: str | None = None
    floor_venue: str | None = None
    shift: HousekeepingShiftEnum | None = None
    supervisor: str | None = None
    # Resolved only for non-rejected rows -- commit_rows() uses these
    # directly rather than re-resolving lookups a second time.
    campus_id: uuid.UUID | None = None
    designation_id: uuid.UUID | None = None
    location_id: uuid.UUID | None = None
    housekeeping_staff_id: uuid.UUID | None = None  # the existing match's id, when status != "created"


@dataclass
class ValidationResult:
    rows: list[ImportRowResult]
    total: int
    created_count: int
    updated_count: int
    unchanged_count: int
    rejected_count: int


def _cell(row: dict, key: str) -> str:
    value = row.get(key)
    if value is None:
        return ""
    return str(value).strip()


def _parse_shift(text: str) -> tuple[HousekeepingShiftEnum | None, str | None]:
    if not text:
        return None, f"Missing required column '{SHIFT_HEADER}'"
    upper = text.upper()
    if upper not in _SHIFT_VALUES:
        return None, f"Unknown '{SHIFT_HEADER}' value '{text}'. Valid values: {', '.join(_SHIFT_VALUES)}"
    return HousekeepingShiftEnum(upper), None


def validate_rows(db: Session, raw_rows: list[dict]) -> ValidationResult:
    """Parses+validates every row against the current DB state **without
    writing anything** -- same read-only contract as the other two
    importers' `validate_rows`.
    """
    if len(raw_rows) > MAX_ROWS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"File has {len(raw_rows)} data rows, exceeding the {MAX_ROWS}-row limit",
        )

    results: list[ImportRowResult] = []
    seen_keys: dict[tuple[str, str], int] = {}
    created = updated = unchanged = rejected = 0

    for row_number, raw_row in enumerate(raw_rows, start=2):
        campus_code = _cell(raw_row, CAMPUS_HEADER).upper()
        bio_id = _cell(raw_row, BIO_ID_HEADER)
        name = _cell(raw_row, NAME_HEADER)
        designation_name = _cell(raw_row, DESIGNATION_HEADER)
        location_name = _cell(raw_row, LOCATION_HEADER)
        block = _cell(raw_row, BLOCK_HEADER) or None
        floor_venue = _cell(raw_row, FLOOR_HEADER) or None
        shift_raw = _cell(raw_row, SHIFT_HEADER)
        supervisor = _cell(raw_row, SUPERVISOR_HEADER) or None

        errors: list[str] = []

        if not campus_code:
            errors.append(f"Missing required column '{CAMPUS_HEADER}'")
        elif campus_code not in CAMPUS_CODES:
            errors.append(f"Unknown campus code '{campus_code}'")

        if not bio_id:
            errors.append(f"Missing required column '{BIO_ID_HEADER}'")
        if not name:
            errors.append(f"Missing required column '{NAME_HEADER}'")
        if not designation_name:
            errors.append(f"Missing required column '{DESIGNATION_HEADER}'")
        if not location_name:
            errors.append(f"Missing required column '{LOCATION_HEADER}'")

        shift, shift_error = _parse_shift(shift_raw)
        if shift_error:
            errors.append(shift_error)

        campus = None
        if campus_code and campus_code in CAMPUS_CODES:
            campus = db.query(Campus).filter(Campus.code == campus_code).one_or_none()
            if campus is None:
                errors.append(f"Campus '{campus_code}' is not seeded in this environment")

        designation = None
        if designation_name:
            designation = db.query(Designation).filter(Designation.name.ilike(designation_name)).one_or_none()
            if designation is None:
                errors.append(f"Unknown designation '{designation_name}'")
            elif designation.category != StaffRoleCategoryEnum.HOUSEKEEPING:
                errors.append(
                    f"Designation category ({designation.category.value}) is not HOUSEKEEPING -- "
                    "Housekeeping staff must be assigned a Housekeeping designation."
                )

        location = None
        if campus is not None and location_name:
            location = (
                db.query(Location)
                .filter(Location.campus_id == campus.id, Location.name.ilike(location_name))
                .order_by(Location.created_at)
                .first()
            )
            if location is None:
                errors.append(f"Unknown location '{location_name}' for campus '{campus_code}'")

        if campus_code and bio_id:
            key = (campus_code, bio_id.lower())
            if key in seen_keys:
                errors.append(f"Duplicate key already used on row {seen_keys[key]}")
            else:
                seen_keys[key] = row_number

        result = ImportRowResult(
            row_number=row_number,
            status="rejected",
            campus_code=campus_code or None,
            bio_id=bio_id or None,
            name=name or None,
            designation_name=designation_name or None,
            location_name=location_name or None,
            block=block,
            floor_venue=floor_venue,
            shift=shift,
            supervisor=supervisor,
        )

        if errors:
            result.error_reason = "; ".join(errors)
            results.append(result)
            rejected += 1
            continue

        result.campus_id = campus.id
        result.designation_id = designation.id
        result.location_id = location.id

        existing = (
            db.query(HousekeepingStaff)
            .filter(HousekeepingStaff.campus_id == campus.id, HousekeepingStaff.bio_id == bio_id)
            .one_or_none()
        )
        if existing is None:
            result.status = "created"
            created += 1
        elif (
            existing.designation_id != designation.id
            or existing.location_id != location.id
            or existing.block != block
            or existing.floor_venue != floor_venue
            or existing.shift != shift
            or existing.supervisor != supervisor
            or existing.name != name
        ):
            result.status = "updated"
            result.housekeeping_staff_id = existing.id
            updated += 1
        else:
            result.status = "unchanged"
            result.housekeeping_staff_id = existing.id
            unchanged += 1

        results.append(result)

    return ValidationResult(
        rows=results,
        total=len(raw_rows),
        created_count=created,
        updated_count=updated,
        unchanged_count=unchanged,
        rejected_count=rejected,
    )


def commit_rows(
    db: Session, *, validation: ValidationResult, actor: User, bulk_upload_log_id: uuid.UUID
) -> None:
    """Applies every non-rejected row's UPSERT within the *caller's* existing
    transaction -- no `db.commit()` here, same contract as the other two
    importers' `commit_rows`. Also writes one `BulkUploadRowLog` row per
    non-rejected row (see this module's own docstring for why).
    """
    for row in validation.rows:
        if row.status == "created":
            new_row = HousekeepingStaff(
                campus_id=row.campus_id,
                bio_id=row.bio_id,
                name=row.name,
                designation_id=row.designation_id,
                location_id=row.location_id,
                block=row.block,
                floor_venue=row.floor_venue,
                shift=row.shift,
                supervisor=row.supervisor,
                created_by_id=actor.id,
            )
            db.add(new_row)
            db.flush()
            db.add(
                BulkUploadRowLog(
                    bulk_upload_log_id=bulk_upload_log_id,
                    entity_type=BulkUploadEntityTypeEnum.HOUSEKEEPING_STAFF,
                    entity_id=new_row.id,
                    was_created=True,
                )
            )
        elif row.status == "updated":
            existing = db.get(HousekeepingStaff, row.housekeeping_staff_id)
            existing.name = row.name
            existing.designation_id = row.designation_id
            existing.location_id = row.location_id
            existing.block = row.block
            existing.floor_venue = row.floor_venue
            existing.shift = row.shift
            existing.supervisor = row.supervisor
            existing.updated_by_id = actor.id
            db.add(
                BulkUploadRowLog(
                    bulk_upload_log_id=bulk_upload_log_id,
                    entity_type=BulkUploadEntityTypeEnum.HOUSEKEEPING_STAFF,
                    entity_id=existing.id,
                    was_created=False,
                )
            )
        elif row.status == "unchanged":
            db.add(
                BulkUploadRowLog(
                    bulk_upload_log_id=bulk_upload_log_id,
                    entity_type=BulkUploadEntityTypeEnum.HOUSEKEEPING_STAFF,
                    entity_id=row.housekeeping_staff_id,
                    was_created=False,
                )
            )
        # "rejected": no write, no row log entry.


def _write_header(ws, headers: tuple[str, ...]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 24


def build_bulk_upload_template_xlsx(db: Session) -> bytes:
    """Live-generated, same styling shape as the other two importers'
    template builders. Master Lists sheet lists campus codes, valid shifts,
    and currently-active Housekeeping designations/locations so a user can
    copy exact free-text values in (Designation Name/Location Name are not
    dropdown-validated -- open-ended free text matched case-insensitively,
    same restraint `sanctioned_strength_import.py`'s own template takes for
    Department/Designation names).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Housekeeping Staff"
    _write_header(ws, TEMPLATE_HEADERS)
    for example in _EXAMPLE_ROWS:
        ws.append(list(example))
    for row_idx in range(2, 2 + len(_EXAMPLE_ROWS)):
        for col in range(1, len(TEMPLATE_HEADERS) + 1):
            ws.cell(row=row_idx, column=col).fill = _SAMPLE_FILL

    campus_formula = '"' + ",".join(CAMPUS_CODES) + '"'
    campus_dv = DataValidation(type="list", formula1=campus_formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(campus_dv)
    campus_dv.add("A2:A500")

    shift_formula = '"' + ",".join(_SHIFT_VALUES) + '"'
    shift_dv = DataValidation(type="list", formula1=shift_formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(shift_dv)
    shift_dv.add("H2:H500")

    master_ws = wb.create_sheet("Master Lists")
    master_ws.append(["Campus Code", "Shift", "Housekeeping Designation Name", "Location Name (Campus)"])
    for cell in master_ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    designations = (
        db.query(Designation)
        .filter(Designation.is_active.is_(True), Designation.category == StaffRoleCategoryEnum.HOUSEKEEPING)
        .order_by(Designation.name)
        .all()
    )
    # Location has no `campus` relationship (unlike Department) -- select
    # Campus.code alongside it explicitly rather than relying on lazy-load.
    location_rows = (
        db.query(Location, Campus.code)
        .join(Campus, Location.campus_id == Campus.id)
        .filter(Location.is_active.is_(True))
        .order_by(Campus.code, Location.name)
        .all()
    )
    location_labels = [f"{location.name} ({campus_code})" for location, campus_code in location_rows]

    max_rows = max(len(CAMPUS_CODES), len(_SHIFT_VALUES), len(designations), len(location_labels))
    for row_idx in range(max_rows):
        if row_idx < len(CAMPUS_CODES):
            master_ws.cell(row=row_idx + 2, column=1, value=CAMPUS_CODES[row_idx])
        if row_idx < len(_SHIFT_VALUES):
            master_ws.cell(row=row_idx + 2, column=2, value=_SHIFT_VALUES[row_idx])
        if row_idx < len(designations):
            master_ws.cell(row=row_idx + 2, column=3, value=designations[row_idx].name)
        if row_idx < len(location_labels):
            master_ws.cell(row=row_idx + 2, column=4, value=location_labels[row_idx])
    for col in range(1, 5):
        master_ws.column_dimensions[get_column_letter(col)].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_error_report_xlsx(log: BulkUploadLog, rejected_rows: list[ImportRowResult]) -> bytes:
    """Same shape as the other two importers' `build_error_report_xlsx` --
    only this upload's rejected rows, plus the `error_reason` column."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rejected rows"
    ws.append([f"Bulk upload: {log.filename}"])
    ws.append([f"Uploaded: {log.uploaded_at.isoformat()}"])
    ws.append([])
    headers = list(TEMPLATE_HEADERS) + ["error_reason"]
    ws.append(headers)
    for row in rejected_rows:
        ws.append(
            [
                row.campus_code,
                row.bio_id,
                row.name,
                row.designation_name,
                row.location_name,
                row.block,
                row.floor_venue,
                row.shift.value if row.shift else None,
                row.supervisor,
                row.error_reason,
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
