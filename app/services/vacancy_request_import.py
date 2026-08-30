"""Vacancy Request bulk upload: validate -> preview -> commit (2026-08-30).

Sibling to `location_import.py` / `department_import.py` / the other four, and
follows their dataclass/validate/commit shape deliberately rather than
inventing a new one. `parse_rows` is reused from `sanctioned_strength_import`
for the same reason those modules reuse it: it is genuinely entity-agnostic.

**Two things make this one different from every other importer here, and both
are deliberate.**

1. **It is CREATE-ONLY. There is no upsert.** Every other entity in
   `BulkUploadEntityTypeEnum` is master data, where re-uploading a row for an
   existing record should update it. A vacancy request is an *event*: two
   identical rows are two genuine requests for two posts, not a duplicate to
   be collapsed. Silently matching a new row onto an existing request would
   lose a real request. So `updated_count` and `unchanged_count` are always 0
   -- reported for response-shape parity with the other five, not because
   they can ever be non-zero.

   This has a second benefit worth stating: the shared row-log undo can only
   revert rows a batch *created*, so a create-only importer is the only one
   whose batches are fully revertible.

2. **Rows are created as DRAFT, not submitted.** `vacancy_workflow.submit()`
   enforces the Sanctioned Strength ceiling per request and would refuse some
   rows and accept others, leaving a batch half in the approval queue -- and
   commit is meant to be all-or-nothing. Creating DRAFTs matches what the
   in-app single-entry form does (it also creates a DRAFT and requires an
   explicit submit), so the uploader reviews and submits them afterwards
   through the normal choke point. The importer never touches `.status`
   itself beyond the initial default.

Validation mirrors the authenticated create path, including the
`Department.supports(category)` MEMBERSHIP rule -- not
`designation.category == department.category`, which was the original bug
(CLAUDE.md).
"""

import io
import uuid
from dataclasses import dataclass
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from app.models.bulk_upload_log import BulkUploadLog
from app.models.bulk_upload_row_log import BulkUploadRowLog
from app.models.campus import Campus
from app.models.department import Department
from app.models.designation import Designation
from app.models.enums import (
    CAMPUS_CODES,
    BulkUploadEntityTypeEnum,
    VacancyPriorityEnum,
    VacancyRequestSourceEnum,
)
from app.models.vacancy_request import VacancyRequest
from app.services.sanctioned_strength_import import parse_rows  # noqa: F401 -- re-exported, see module docstring

MAX_ROWS = 5000

CAMPUS_HEADER = "Campus Code"
DEPARTMENT_HEADER = "Department Name"
DESIGNATION_HEADER = "Designation"
COUNT_HEADER = "Number of Positions"
PRIORITY_HEADER = "Priority"
REQUIRED_BY_HEADER = "Required By (DD-MM-YYYY)"
JUSTIFICATION_HEADER = "Justification"

TEMPLATE_HEADERS = (
    CAMPUS_HEADER,
    DEPARTMENT_HEADER,
    DESIGNATION_HEADER,
    COUNT_HEADER,
    PRIORITY_HEADER,
    REQUIRED_BY_HEADER,
    JUSTIFICATION_HEADER,
)

# Same functional backstop every other importer uses: "XXX" is never a real
# campus code, so a forgotten example row always comes back rejected rather
# than silently importing as a real vacancy request.
_EXAMPLE_ROWS = (
    ("XXX", "EXAMPLE - DELETE THIS ROW", "Assistant Professor", 2, "NORMAL", "01-04-2026", "Sample only"),
    ("XXX", "EXAMPLE - DELETE THIS ROW", "Lab Assistant", 1, "HIGH", "", "Sample only"),
)

_HEADER_FILL = PatternFill(start_color="1B5FAA", end_color="1B5FAA", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SAMPLE_FILL = PatternFill(start_color="FFF3D6", end_color="FFF3D6", fill_type="solid")

_PRIORITY_VALUES = tuple(p.value for p in VacancyPriorityEnum)

# Upper bound on one row's requested_count. Matches the public QR form's own
# cap: a four-digit typo in a spreadsheet cell is a far more likely
# explanation than a genuine request for 5000 posts, and the resulting
# HiringSlot rows would be created for real at HR approval.
MAX_POSITIONS_PER_ROW = 100


@dataclass
class ImportRowResult:
    row_number: int
    status: str  # created | rejected  -- see the module docstring: never updated/unchanged
    error_reason: str | None = None
    campus_code: str | None = None
    department_name: str | None = None
    designation_name: str | None = None
    requested_count: int | None = None
    priority: str | None = None
    required_by: date | None = None
    justification: str | None = None
    # Resolved only for non-rejected rows -- commit_rows() uses these directly
    # rather than re-running the lookups a second time.
    campus_id: uuid.UUID | None = None
    department_id: uuid.UUID | None = None
    designation_id: uuid.UUID | None = None


@dataclass
class ValidationResult:
    rows: list[ImportRowResult]
    total: int
    created_count: int
    # Always 0. Kept for response-shape parity with the other five importers;
    # see the module docstring for why they can never be non-zero here.
    updated_count: int
    unchanged_count: int
    rejected_count: int


def _cell(row: dict, key: str) -> str:
    value = row.get(key)
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _normalize(text: str | None) -> str:
    """Trimmed, whitespace-collapsed, lower-cased -- used only for MATCHING
    master-data names, never for storage."""
    return " ".join((text or "").split()).lower()


def _parse_date(text: str) -> tuple[date | None, str | None]:
    """DD-MM-YYYY, the format the template header states and every other
    importer in this codebase uses. ISO is also accepted because openpyxl
    hands back real dates from a date-formatted cell, which `_cell` renders
    as ISO -- rejecting those would fail rows that are actually correct."""
    if not text:
        return None, None
    for fmt in ("%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date(), None
        except ValueError:
            continue
    return None, f"Could not read '{text}' as a date. Use DD-MM-YYYY."


def validate_rows(db: Session, raw_rows: list[dict]) -> ValidationResult:
    """Pure preview -- writes nothing. Mirrors the authenticated create
    path's own validation so a row that previews as `created` here would also
    have been accepted through the form."""
    if len(raw_rows) > MAX_ROWS:
        from fastapi import HTTPException, status as http_status

        raise HTTPException(
            status_code=http_status.HTTP_400_BAD_REQUEST,
            detail=f"File has {len(raw_rows)} rows; the limit is {MAX_ROWS}.",
        )

    campuses_by_code = {c.code.upper(): c for c in db.query(Campus).filter(Campus.is_active.is_(True)).all()}
    departments = db.query(Department).filter(Department.is_active.is_(True)).all()
    designations = db.query(Designation).filter(Designation.is_active.is_(True)).all()

    # Department names are unique per campus, so key on (campus_id, name).
    departments_by_key = {(d.campus_id, _normalize(d.name)): d for d in departments}
    # Designation names are NOT guaranteed unique across the table, so this
    # keeps the FIRST match by name and rejects nothing on that basis -- the
    # department-supports check below is what actually constrains the choice.
    designations_by_name: dict[str, Designation] = {}
    for designation in designations:
        designations_by_name.setdefault(_normalize(designation.name), designation)

    rows: list[ImportRowResult] = []
    for index, raw in enumerate(raw_rows, start=2):  # row 1 is the header
        campus_code = _cell(raw, CAMPUS_HEADER).upper()
        department_name = _cell(raw, DEPARTMENT_HEADER)
        designation_name = _cell(raw, DESIGNATION_HEADER)
        count_text = _cell(raw, COUNT_HEADER)
        priority_text = _cell(raw, PRIORITY_HEADER).upper() or VacancyPriorityEnum.NORMAL.value
        required_by_text = _cell(raw, REQUIRED_BY_HEADER)
        justification = _cell(raw, JUSTIFICATION_HEADER)

        result = ImportRowResult(
            row_number=index,
            status="rejected",
            campus_code=campus_code or None,
            department_name=department_name or None,
            designation_name=designation_name or None,
            priority=priority_text,
            justification=justification or None,
        )

        if not campus_code or not department_name or not designation_name:
            result.error_reason = "Campus Code, Department Name and Designation are all required."
            rows.append(result)
            continue

        if campus_code not in CAMPUS_CODES:
            result.error_reason = f"Unknown campus code '{campus_code}'."
            rows.append(result)
            continue
        campus = campuses_by_code.get(campus_code)
        if campus is None:
            result.error_reason = f"Campus '{campus_code}' is not active."
            rows.append(result)
            continue

        department = departments_by_key.get((campus.id, _normalize(department_name)))
        if department is None:
            result.error_reason = f"Unknown department '{department_name}' on campus {campus_code}."
            rows.append(result)
            continue

        designation = designations_by_name.get(_normalize(designation_name))
        if designation is None:
            result.error_reason = f"Unknown designation '{designation_name}'."
            rows.append(result)
            continue

        # MEMBERSHIP, not equality -- a department supports a SET of
        # categories (CLAUDE.md).
        if not department.supports(designation.category):
            result.error_reason = (
                f"{department.name} does not support {designation.category.value} designations."
            )
            rows.append(result)
            continue

        if not count_text.isdigit() or int(count_text) < 1:
            result.error_reason = "Number of Positions must be a whole number of 1 or more."
            rows.append(result)
            continue
        requested_count = int(count_text)
        if requested_count > MAX_POSITIONS_PER_ROW:
            result.error_reason = f"Number of Positions cannot exceed {MAX_POSITIONS_PER_ROW}."
            rows.append(result)
            continue

        if priority_text not in _PRIORITY_VALUES:
            result.error_reason = f"Unknown priority '{priority_text}'. Use one of: {', '.join(_PRIORITY_VALUES)}."
            rows.append(result)
            continue

        required_by, date_error = _parse_date(required_by_text)
        if date_error:
            result.error_reason = date_error
            rows.append(result)
            continue

        result.status = "created"
        result.error_reason = None
        result.requested_count = requested_count
        result.required_by = required_by
        result.campus_id = campus.id
        result.department_id = department.id
        result.designation_id = designation.id
        rows.append(result)

    created = sum(1 for r in rows if r.status == "created")
    rejected = sum(1 for r in rows if r.status == "rejected")
    return ValidationResult(
        rows=rows,
        total=len(rows),
        created_count=created,
        updated_count=0,
        unchanged_count=0,
        rejected_count=rejected,
    )


def commit_rows(
    db: Session,
    *,
    validation: ValidationResult,
    bulk_upload_log_id: uuid.UUID,
    requested_by_id: uuid.UUID,
) -> None:
    """Creates one DRAFT VacancyRequest per non-rejected row, plus one
    `BulkUploadRowLog` so the batch can be undone.

    Rejected rows write nothing at all -- not a request, not a row log --
    matching the brief's "do not commit rejected rows" and every other
    importer's behaviour.

    `requested_by_id` is the UPLOADER. They are authenticated and permitted to
    raise requests, so attributing the rows to them is accurate; `source` is
    BULK_UPLOAD, which is what distinguishes these from what that person
    raised by hand.
    """
    for row in validation.rows:
        if row.status == "rejected":
            continue

        designation = db.get(Designation, row.designation_id)
        vr = VacancyRequest(
            campus_id=row.campus_id,
            department_id=row.department_id,
            designation_id=row.designation_id,
            role_category=designation.category,
            position_title=designation.name,
            employment_type=designation.employment_type,
            requested_count=row.requested_count,
            # NOT NULL on the model and not asked for in the template -- taken
            # from Designation Master, exactly as the public QR intake does,
            # so a bulk-created request carries the same detail an in-app one
            # would for that designation.
            qualification=designation.qualification,
            experience_required=designation.min_experience,
            remarks=row.justification,
            priority=VacancyPriorityEnum(row.priority),
            required_by=row.required_by,
            source=VacancyRequestSourceEnum.BULK_UPLOAD,
            requested_by_id=requested_by_id,
        )
        db.add(vr)
        db.flush()  # assigns vr.id for the row log below

        db.add(
            BulkUploadRowLog(
                bulk_upload_log_id=bulk_upload_log_id,
                entity_type=BulkUploadEntityTypeEnum.VACANCY_REQUEST,
                entity_id=vr.id,
                # Always True here -- this importer never updates (see the
                # module docstring), which is precisely what makes its
                # batches fully revertible.
                was_created=True,
            )
        )


def _write_header(ws, headers: tuple[str, ...]) -> None:
    for column_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=column_index, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(column_index)].width = max(16, len(header) + 4)


def build_bulk_upload_template_xlsx(db: Session) -> bytes:
    """Template workbook. Headers match TEMPLATE_HEADERS exactly, so an
    export -> edit -> re-upload round trip stays possible -- the property the
    master-data exports deliberately preserve."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vacancy Requests"
    _write_header(ws, TEMPLATE_HEADERS)

    for row_index, example in enumerate(_EXAMPLE_ROWS, start=2):
        for column_index, value in enumerate(example, start=1):
            cell = ws.cell(row=row_index, column=column_index, value=value)
            cell.fill = _SAMPLE_FILL

    priority_validation = DataValidation(
        type="list", formula1=f'"{",".join(_PRIORITY_VALUES)}"', allow_blank=True
    )
    ws.add_data_validation(priority_validation)
    priority_validation.add(f"E2:E{MAX_ROWS + 1}")

    reference = wb.create_sheet("Reference")
    reference["A1"] = "Campus Codes"
    reference["A1"].font = Font(bold=True)
    for offset, code in enumerate(sorted(CAMPUS_CODES), start=2):
        reference[f"A{offset}"] = code
    reference["B1"] = "Priorities"
    reference["B1"].font = Font(bold=True)
    for offset, value in enumerate(_PRIORITY_VALUES, start=2):
        reference[f"B{offset}"] = value

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_error_report_xlsx(log: BulkUploadLog, rejected_rows: list[ImportRowResult]) -> bytes:
    """One sheet of rejected rows with their reasons -- same shape as the
    other importers' error reports, so the shared download endpoint needs no
    special-casing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rejected Rows"
    headers = ("Row", *TEMPLATE_HEADERS, "Error Reason")
    _write_header(ws, headers)

    for offset, row in enumerate(rejected_rows, start=2):
        ws.cell(row=offset, column=1, value=row.row_number)
        ws.cell(row=offset, column=2, value=row.campus_code)
        ws.cell(row=offset, column=3, value=row.department_name)
        ws.cell(row=offset, column=4, value=row.designation_name)
        ws.cell(row=offset, column=5, value=row.requested_count)
        ws.cell(row=offset, column=6, value=row.priority)
        ws.cell(row=offset, column=7, value=row.required_by.isoformat() if row.required_by else None)
        ws.cell(row=offset, column=8, value=row.justification)
        ws.cell(row=offset, column=9, value=row.error_reason)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
